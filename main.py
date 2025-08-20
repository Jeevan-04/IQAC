from pymongo import MongoClient
from bson import ObjectId
from datetime import datetime, timezone
from typing import Dict, List, Optional, Union, Any
from nicegui.events import UploadEventArguments
from fastapi import FastAPI, Request, Depends, HTTPException, status
from fastapi import Request as FastAPIRequest
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import hashlib
import secrets
import re
from nicegui import app, ui
import urllib.parse
import pandas as pd
import io
import base64
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import HTTPException
from pymongo import MongoClient, ASCENDING, DESCENDING
from bson import ObjectId
from datetime import datetime, timedelta
import os
import json
import string
import uuid
import logging
from typing import Dict, List, Optional, Any
# Import MongoClient for MongoDB connection
from pymongo import MongoClient
# MongoDB connection functions removed

# Supabase table creation functions removed

# MongoDB to Supabase migration function removed

# MongoDB Setup
client = MongoClient('mongodb://localhost:27017/')
db = client['iqac_portal']

# Collections
institutions_col = db['institutions']
users_col = db['users']
academic_years_col = db['academic_years']
schools_col = db['schools']
programs_col = db['programs']
audit_logs_col = db['audit_logs']
criterias_col = db['criterias']
extended_profiles_col = db['extended_profiles']
criteria_submissions_col = db['criteria_submissions']
# Define the data_table_container as a NiceGUI container
data_table_container = ui.column().style('display: block;')
print("DEBUG: Data table container display set to block")
with data_table_container:
    print("DEBUG: Creating table content")
    # Define main_color with a default value
    main_color = '#667eea'
    # Define data_rows and file_name with default values
    data_rows = []
    file_name = 'N/A'
    # Define headers with a default value
    headers = ['Name', 'Email', 'Role']
    ui.label('📊 Data Table').style(
        f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
    )
    ui.label(f'Parsed {len(data_rows)} rows from {file_name}').style(
        'color: var(--text-secondary); margin-bottom: 1.5rem;'
    )
    # Create editable data table
    if data_rows:
        print(f"DEBUG: Creating {len(data_rows)} row cards")
        # Create a grid of input fields for each row
        for row_index, row_data in enumerate(data_rows):
            print(f"DEBUG: Creating row {row_index + 1}")
            with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                ui.label(f'Row {row_index + 1}').style(
                    'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                )
                # Create input fields for each header
                with ui.row().style('flex-wrap: wrap; gap: 1rem;'):
                    for header in headers:
                        with ui.column().style('min-width: 200px;'):
                            ui.label(header).style(
                                'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                            )
                            # Check if cell is empty and highlight it
                            cell_value = row_data.get(header, '')
                            is_empty = not cell_value.strip()
                            input_field = ui.input(
                                label='',
                                value=cell_value,
                                placeholder=f'Enter {header}'
                            ).style(f'width: 100%; border: {"2px solid #dc3545" if is_empty else "1px solid #e9ecef"};')
                            # Store reference to input field for later data collection
                            row_data[f'input_{header}'] = input_field
    try:
    # ...existing code reformatted for correct indentation and block structure...
        # Always show at least one empty row for manual data entry
        if not data_rows:
            print("DEBUG: No data rows found, creating empty row for manual entry")
            ui.label('📝 Manual Data Entry').style(
                f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
            )
            ui.label('No data found in the file. You can fill the data manually below:').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem; text-align: center;'
            )
            # Create empty row for manual entry
            with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                ui.label('Row 1 (Manual Entry)').style(
                    'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                )
                # Create input fields for each header
                with ui.row().style('flex-wrap: wrap; gap: 1rem;'):
                    for header in headers:
                        with ui.column().style('min-width: 200px;'):
                            ui.label(header).style(
                                'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                            )
                            input_field = ui.input(
                                label='',
                                value='',
                                placeholder=f'Enter {header}'
                            ).style('width: 100%; border: 2px dashed #e9ecef;')
                            # Store reference to input field for later data collection
                            if not data_rows:
                                data_rows = [{}]
                            data_rows[0][f'input_{header}'] = input_field
        global entry_method, entry_data, current_profile
        print("DEBUG: Creating action buttons")
        # Action buttons (always show)
        with ui.row().style('gap: 1rem; margin-top: 1.5rem; justify-content: center;'):
            ui.button('💾 Save as Draft', on_click=lambda: save_as_draft(data_rows, headers)).style(
                'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
            )
            ui.button('✅ Submit', on_click=lambda: submit_data(data_rows, headers)).style(
                f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
            )
        print("DEBUG: Table creation complete, showing success notification")
        ui.notify(f'File "{file_name}" parsed successfully! {len(data_rows)} rows loaded. Please review and edit the data.', color='positive')
    except Exception as e:
        ui.notify(f'Error processing file: {str(e)}', color='negative')
        import traceback
        traceback.print_exc()

def program_admin_sidebar(program_id, institution=None):
    """Reusable sidebar for program admin pages"""
    if not institution:
        from pymongo import MongoClient
        from bson import ObjectId
        program = programs_col.find_one({'_id': ObjectId(program_id)})
        institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])}) if program else None
    with ui.column().style('width: 200px; background: var(--surface); border-right: 2px solid var(--border); padding: 0; height: 100vh; overflow-y: auto;'):
        with ui.card().style('background: linear-gradient(135deg, var(--primary-color), var(--primary-dark)); color: white; padding: 1.5rem; margin: 1rem; border: none; border-radius: 12px; box-shadow: var(--shadow);'):
            ui.element('div').style('width: 60px; height: 60px; background: rgba(255,255,255,0.2); border-radius: 50%; margin: 0 auto 1rem; display: flex; align-items: center; justify-content: center; font-size: 1.5rem; font-weight: bold;').inner_html = '🎓'
            ui.label(institution.get('name', 'Institution') if institution else 'Institution').style('font-size: 1.2rem; font-weight: bold; text-align: center; margin-bottom: 0.5rem;')
            ui.separator().style('margin: 1rem 0; background: rgba(255,255,255,0.3);')
            current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
            ui.label(current_user_email).style('font-size: 0.9rem; text-align: center; opacity: 0.9; margin-bottom: 0.5rem;')
            ui.label('Program Admin').style('font-size: 0.9rem; text-align: center; opacity: 0.8;')
        with ui.column().style('padding: 1rem; gap: 0.5rem;'):
            with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 1rem; margin-bottom: 1rem;'):
                nav_items = [
                    ('📊 Criteria Management', f'/program_admin/{program_id}/criterias'),
                    ('👤 Extended Profiles', f'/program_admin/{program_id}/profiles'),
                    ('📥 My Submissions', f'/program_admin/{program_id}/submissions'),
                ]
                for label, url in nav_items:
                    ui.button(label, on_click=lambda u=url: ui.navigate.to(u)).style(
                        'width: 100%; justify-content: flex-start; margin-bottom: 0.5rem; text-align: left; background: white; color: var(--text-primary); border: 1px solid #e9ecef; padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 500;'
                    ).on('mouseenter', lambda e: ui.run_javascript('event.target.style.background = "rgb(154, 44, 84)"; event.target.style.color = "white"; event.target.style.borderColor = "rgb(154, 44, 84)";')).on('mouseleave', lambda e: ui.run_javascript('event.target.style.background = "white"; event.target.style.color = "var(--text-primary)"; event.target.style.borderColor = "#e9ecef";'))
                ui.separator().style('margin: 1rem 0; background: #e9ecef;')
                ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                    'width: 100%; justify-content: center; background: #dc3545; color: white; border: 1px solid #dc3545; padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 600;'
                ).on('mouseenter', lambda e: ui.run_javascript('event.target.style.background = "#c82333"; event.target.style.borderColor = "#c82333";')).on('mouseleave', lambda e: ui.run_javascript('event.target.style.background = "#dc3545"; event.target.style.borderColor = "#dc3545";'))


def add_beautiful_global_styles():
    """Add comprehensive styling with theme color support"""
    ui.add_head_html("""
    <style>
        :root {
            --primary-color: rgb(154, 44, 84);
            --primary-dark: rgb(124, 35, 67);
            --primary-light: rgba(154, 44, 84, 0.1);
            --success-color: #4caf50;
            --warning-color: #ff9800;
            --error-color: #f44336;
            --info-color: #2196f3;
            --text-primary: #2c3e50;
            --text-secondary: #546e7a;
            --background: #fafafa;
            --surface: #ffffff;
            --border: rgba(154, 44, 84, 0.2);
            --shadow: 0 2px 8px rgba(154, 44, 84, 0.12);
            --shadow-hover: 0 4px 16px rgba(154, 44, 84, 0.16);
            --border-radius: 8px;
            --transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        body {
            font-family: 'Inter', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: var(--background);
            color: var(--text-primary);
            line-height: 1.6;
        }
        
        /* Beautiful Cards */
        .beautiful-card {
            background: var(--surface);
            border-radius: var(--border-radius);
            box-shadow: var(--shadow);
            border: 1px solid var(--border);
            transition: var(--transition);
            overflow: hidden;
        }
        
        .beautiful-card:hover {
            box-shadow: var(--shadow-hover);
            transform: translateY(-2px);
        }
        
        /* Themed Buttons */
        .btn-primary {
            background: var(--primary-color);
            color: white;
            border: 2px solid var(--primary-color);
            border-radius: var(--border-radius);
            padding: 0.75rem 1.5rem;
            font-weight: 600;
            transition: var(--transition);
            cursor: pointer;
        }
        
        .btn-primary:hover {
            background: var(--primary-dark);
            border-color: var(--primary-dark);
            transform: translateY(-1px);
            box-shadow: var(--shadow-hover);
        }
        
        .btn-secondary {
            background: transparent;
            color: var(--primary-color);
            border: 2px solid var(--primary-color);
            border-radius: var(--border-radius);
            padding: 0.75rem 1.5rem;
            font-weight: 600;
            transition: var(--transition);
            cursor: pointer;
        }
        
        .btn-secondary:hover {
            background: var(--primary-light);
            transform: translateY(-1px);
        }
        
        /* Sidebar Button Styling */
        .sidebar button {
            color: var(--text-primary) !important;
            background: transparent !important;
            border: none !important;
            transition: all 0.3s ease !important;
            border-radius: 8px !important;
        }
        
        .sidebar button:hover {
            background: var(--primary-light) !important;
            color: var(--primary-color) !important;
            transform: translateX(4px);
        }
        
        .btn-success {
            background: var(--success-color);
            color: white;
            border: 2px solid var(--success-color);
            border-radius: var(--border-radius);
            padding: 0.75rem 1.5rem;
            font-weight: 600;
        .btn-warning {
            background: var(--warning-color);
            color: white;
            border: 2px solid var(--warning-color);
            border-radius: var(--border-radius);
            padding: 0.75rem 1.5rem;
            font-weight: 600;
            transition: var(--transition);
            cursor: pointer;
        }
        
        .btn-danger {
            background: var(--error-color);
            color: white;
            border: 2px solid var(--error-color);
            border-radius: var(--border-radius);
            padding: 0.75rem 1.5rem;
            font-weight: 600;
            transition: var(--transition);
            cursor: pointer;
        }
        
        /* Beautiful Inputs */
        .beautiful-input {
            border: 2px solid var(--border);
            border-radius: var(--border-radius);
            padding: 0.75rem;
            transition: var(--transition);
            background: var(--surface);
        }
        
        .beautiful-input:focus {
            border-color: var(--primary-color);
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
            outline: none;
        }
        
        /* Layout Components */
        .sidebar {
            position: fixed !important;
            top: 0;
            left: 0;
            width: 280px;
            background: var(--surface);
            border-right: 1px solid var(--border);
            height: 100vh;
            max-height: 100vh;
            overflow-y: auto;
            overflow-x: hidden;
            padding: 1.5rem;
            z-index: 1000;
        }
        
        .main-content {
            margin-left: 280px;
            background: var(--background);
            min-height: 100vh;
            padding: 2rem;
        }
        
        /* Status Indicators */
        .status-success {
            background: #e8f5e9;
            color: #2e7d32;
            border: 1px solid #c8e6c9;
            padding: 0.5rem 1rem;
            border-radius: var(--border-radius);
            font-weight: 500;
        }
        
        .status-warning {
            background: #fff3e0;
            color: #f57c00;
            border: 1px solid #ffcc02;
            padding: 0.5rem 1rem;
            border-radius: var(--border-radius);
            font-weight: 500;
        }
        
        .status-error {
            background: #ffebee;
            color: #c62828;
            border: 1px solid #ef9a9a;
            padding: 0.5rem 1rem;
            border-radius: var(--border-radius);
            font-weight: 500;
        }
        
        /* Animations */
        .fade-in {
            animation: fadeIn 0.5s ease-in-out;
        }
        
        .animate-fade-in-up {
            animation: fadeInUp 0.6s ease-out;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        
        /* Form Styling */
        .form-group {
            margin-bottom: 1.5rem;
        }
        
        .form-label {
            display: block;
            margin-bottom: 0.5rem;
            font-weight: 600;
            color: var(--text-primary);
            font-size: 0.9rem;
        }
        
        .input-beautiful {
            width: 100%;
            border: 2px solid var(--border);
            border-radius: var(--border-radius);
            padding: 0.875rem 1rem;
            transition: var(--transition);
            background: var(--surface);
            font-size: 1rem;
        }
        
        .input-beautiful:focus {
            border-color: var(--primary-color);
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
            outline: none;
        }
        
        .btn-beautiful {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--primary-dark) 100%);
            color: white;
            border: none;
            border-radius: var(--border-radius);
            padding: 0.875rem 1.5rem;
            font-weight: 600;
            transition: var(--transition);
            cursor: pointer;
            font-size: 1rem;
        }
        
        .btn-beautiful:hover {
            background: linear-gradient(135deg, var(--primary-dark) 0%, #4c63d2 100%);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
        }
        
        .page-title {
            font-size: 2rem;
            font-weight: 700;
            color: var(--text-primary);
            margin-bottom: 0.5rem;
        }
        
        .slide-up {
            animation: slideUp 0.3s ease-out;
        }
        
        @keyframes slideUp {
            from { transform: translateY(20px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
    </style>
    """)

# Color utility functions
def hex_to_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def rgb_to_hex(rgb):
    return '#%02x%02x%02x' % rgb

def lighten_color(color, percent):
    """Lighten a color by a percentage"""
    if color.startswith('rgb'):
        # Extract RGB values from rgb() string
        import re
        nums = re.findall(r'\d+', color)
        if len(nums) >= 3:
            r, g, b = int(nums[0]), int(nums[1]), int(nums[2])
        else:
            r, g, b = 102, 126, 234  # fallback
    else:
        # Assume hex color
        try:
            r, g, b = hex_to_rgb(color)
        except:
            r, g, b = 102, 126, 234  # fallback
    
    # Lighten each component
    r = min(255, int(r + (255 - r) * percent))
    g = min(255, int(g + (255 - g) * percent))
    b = min(255, int(b + (255 - b) * percent))
    
    return f'rgb({r}, {g}, {b})'

def darken_color(color, percent):
    """Darken a color by a percentage"""
    if color.startswith('rgb'):
        import re
        nums = re.findall(r'\d+', color)
        if len(nums) >= 3:
            r, g, b = int(nums[0]), int(nums[1]), int(nums[2])
        else:
            r, g, b = 102, 126, 234  # fallback
    else:
        try:
            r, g, b = hex_to_rgb(color)
        except:
            r, g, b = 102, 126, 234  # fallback
    
    # Darken each component
    r = max(0, int(r * (1 - percent)))
    g = max(0, int(g * (1 - percent)))
    b = max(0, int(b * (1 - percent)))
    
    return f'rgb({r}, {g}, {b})'

def set_theme_colors(primary_color):
    """Set CSS custom properties for theme colors"""
    light_color = lighten_color(primary_color, 0.8)
    dark_color = darken_color(primary_color, 0.1)
    
    ui.add_head_html(f"""
    <style>
        :root {{
            --primary-color: {primary_color};
            --primary-dark: {dark_color};
            --primary-light: {light_color};
        }}
    </style>
    """)

# Helper Functions
def check_auth():
    """Check if user is authenticated and has a valid session"""
    global current_user
    
    # Check if user is in storage
    if not hasattr(app.storage, 'user') or not app.storage.user.get('user'):
        ui.navigate.to('/')
        return False
        
    # Update current_user from storage
    current_user = app.storage.user.get('user')
    
    # Additional check for required fields
    if not current_user or not current_user.get('email') or not current_user.get('role'):
        ui.navigate.to('/')
        return False
        
    return True
def log_audit_action(action, details, user_email=None, institution_id=None, entity_type=None, entity_id=None):
    """Log an audit action"""
    global current_user
    if not user_email and current_user:
        user_email = current_user.get('email', 'Unknown')
    
    audit_log = {
        'timestamp': datetime.now(timezone.utc),
        'user_email': user_email or current_user.get('email'),
        'institution_id': institution_id or current_user.get('institution_id'),
        'action': action,
        'details': details,
        'entity_type': entity_type,
        'entity_id': entity_id,
        'ip_address': None  # TODO: Get real IP address from request
    }
    audit_logs_col.insert_one(audit_log)

def hash_password(password, salt):
    """Hash a password with salt"""
    return hashlib.sha256((password + salt).encode()).hexdigest()

# Program Code Generation
def generate_program_code(institution_id, academic_year_name, program_type, program_name):
    """
    Generate a unique program code in format: ITM_2023_UG_BAC001
    Format: INSTITUTION_YEAR_TYPE_PROGRAM001
    - ITM = Institution abbreviation (first 3 letters)
    - 2023 = Academic year (extracted from format like "2023-24")
    - UG/PG/PHD/DP = Program type
    - BAC = First 3 letters of program name
    - 001 = Sequential number (ensures uniqueness)
    """
    from bson import ObjectId
    
    # Get institution abbreviation (first 3 letters of name, uppercase)
    inst = institutions_col.find_one({'_id': ObjectId(institution_id)})
    inst_name = inst.get('name', 'ITM') if inst else 'ITM'
    inst_abbr = ''.join([c for c in inst_name if c.isalpha()])[:3].upper()
    if len(inst_abbr) < 3:
        inst_abbr = inst_abbr + 'M' * (3 - len(inst_abbr))  # Pad with M if needed
    
    # Extract year from academic year name (2023-24 -> 2023)
    year_abbr = academic_year_name.split('-')[0] if '-' in academic_year_name else academic_year_name[:4]
    # Clean year to ensure it's numeric
    year_abbr = ''.join([c for c in year_abbr if c.isdigit()])[:4]
    if len(year_abbr) < 4:
        year_abbr = '2023'  # Default fallback
    
    # Program type mapping
    type_mapping = {
        'Undergraduate': 'UG',
        'Postgraduate': 'PG', 
        'Diploma': 'DP',
        'Certificate': 'CT',
        'Doctorate': 'PHD',
        'PhD': 'PHD'
    }
    type_abbr = type_mapping.get(program_type, program_type.upper()[:3])
    
    # Program abbreviation (first 3 letters of program name, uppercase)
    prog_name_clean = ''.join([c for c in program_name if c.isalpha()])
    prog_abbr = prog_name_clean[:3].upper() if prog_name_clean else 'PRG'
    if len(prog_abbr) < 3:
        prog_abbr = prog_abbr + 'G' * (3 - len(prog_abbr))  # Pad with G if needed
    
    # Base code format: ITM_2023_UG_BAC
    base_code = f"{inst_abbr}_{year_abbr}_{type_abbr}_{prog_abbr}"
    
    # Find all existing codes with this base pattern to determine next sequence number
    existing_codes = list(programs_col.find(
        {'code': {'$regex': f'^{base_code}\\d{{3}}$'}},  # Match exact pattern with 3 digits
        {'code': 1}
    ))
    
    # Extract sequence numbers and find the next available one
    used_sequences = []
    for doc in existing_codes:
        code = doc.get('code', '')
        if code.startswith(base_code) and len(code) == len(base_code) + 3:
            try:
                seq_num = int(code[-3:])  # Last 3 digits
                used_sequences.append(seq_num)
            except ValueError:
                continue
    
    # Find next available sequence number
    sequence = 1
    while sequence in used_sequences:
        sequence += 1
    
    # Ensure sequence doesn't exceed 999
    if sequence > 999:
        sequence = 999
    
    # Format final code: ITM_2023_UG_BAC001
    final_code = f"{base_code}{sequence:03d}"
    
    return final_code
    if existing_codes:
        # Extract sequence numbers from existing codes
        sequences = []
        for prog in existing_codes:
            code = prog.get('code', '')
            if code:
                try:
                    seq_part = code.split('_')[-1]  # Get last part
                    seq_num = ''.join(filter(str.isdigit, seq_part))  # Extract digits
                    if seq_num:
                        sequences.append(int(seq_num))
                except:
                    pass
        if sequences:
            sequence = max(sequences) + 1
    
    # Generate final code
    code = f"{inst_abbr}_{year_abbr}_{type_abbr}_{prog_abbr}{sequence:03d}"
    return code

# Create User Group Page
@ui.page('/institution_admin/{inst_id}/create_user_group')
def create_user_group_page(inst_id: str, school_id: str = None, program_id: str = None, department_id: str = None):
    """Create user group with hierarchy context"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    from nicegui import app
    from bson import ObjectId
    
    # Initialize app.storage.user if needed
    if not hasattr(app.storage, 'user'):
        app.storage.user = {}
    
    # Get query parameters from the current request - try multiple approaches
    # First check if parameters were passed directly to the function
    if school_id or program_id or department_id:
        print(f"DEBUG: Parameters passed directly to function - school_id: '{school_id}', program_id: '{program_id}', department_id: '{department_id}'")
        full_query = f"school_id={school_id}&program_id={program_id}&department_id={department_id}" if any([school_id, program_id, department_id]) else ""
    else:
        # Try to get from query parameters
        school_id = ''
        program_id = ''
        department_id = ''
        full_query = ''
    
    try:
        # Method 1: Try NiceGUI context approach
        from nicegui import context
        req = context.get().request
        query_params = req.query_params
        school_id = query_params.get('school_id','')
        program_id = query_params.get('program_id','')
        department_id = query_params.get('department_id','')
        full_query = req.url.query if hasattr(req.url, 'query') else str(query_params)
        
        print(f"DEBUG: Method 1 - Raw query_params: {query_params}")
        print(f"DEBUG: Method 1 - req.url: {req.url}")
        print(f"DEBUG: Method 1 - req.url.query: {getattr(req.url, 'query', 'N/A')}")
        
    except Exception as e:
        print(f"DEBUG: Method 1 failed: {e}")
        
        try:
            # Method 2: Try direct URL parsing
            from nicegui import context
            req = context.get().request
            url_str = str(req.url)
            print(f"DEBUG: Method 2 - Full URL string: {url_str}")
            
            # Parse URL manually
            if '?' in url_str:
                query_part = url_str.split('?')[1]
                full_query = query_part
                print(f"DEBUG: Method 2 - Query part: {query_part}")
                
                # Parse individual parameters
                params = query_part.split('&')
                for param in params:
                    if '=' in param:
                        key, value = param.split('=', 1)
                        if key == 'school_id':
                            school_id = value
                        elif key == 'program_id':
                            program_id = value
                        elif key == 'department_id':
                            department_id = value
                
                print(f"DEBUG: Method 2 - Parsed params: school_id='{school_id}', program_id='{program_id}', department_id='{department_id}'")
            
        except Exception as e2:
            print(f"DEBUG: Method 2 also failed: {e2}")
            
            try:
                # Method 3: Try Flask-style request.args
                from nicegui import context
                req = context.get().request
                if hasattr(req, 'args'):
                    school_id = req.args.get('school_id', '')
                    program_id = req.args.get('program_id', '')
                    department_id = req.args.get('department_id', '')
                    print(f"DEBUG: Method 3 - Flask args: school_id='{school_id}', program_id='{program_id}', department_id='{department_id}'")
                else:
                    print("DEBUG: Method 3 - No Flask args available")
            except Exception as e3:
                print(f"DEBUG: Method 3 also failed: {e3}")
                
        # Method 4: Try to get URL from browser if server methods failed
        if not any([school_id, program_id, department_id]):
            print("DEBUG: All server methods failed, trying browser URL...")
            try:
                # This will be executed in the browser
                browser_url = ui.run_javascript('return window.location.href;')
                print(f"DEBUG: Browser URL: {browser_url}")
                
                # Parse the URL manually
                if browser_url and '?' in browser_url:
                    query_part = browser_url.split('?')[1]
                    full_query = query_part
                    print(f"DEBUG: Browser query part: {query_part}")
                    
                    # Parse individual parameters
                    params = query_part.split('&')
                    for param in params:
                        if '=' in param:
                            key, value = param.split('=', 1)
                            if key == 'school_id':
                                school_id = value
                            elif key == 'program_id':
                                program_id = value
                            elif key == 'department_id':
                                department_id = value
                    
                    print(f"DEBUG: Browser parsed params: school_id='{school_id}', program_id='{program_id}', department_id='{department_id}'")
            except Exception as e4:
                print(f"DEBUG: Browser method also failed: {e4}")
    
    # Final debug logging
    print(f"DEBUG: Final extracted params - school_id: '{school_id}', program_id: '{program_id}', department_id: '{department_id}'")
    print(f"DEBUG: Full query string: {full_query}")
    
    # Validate that we got at least one parameter
    if not any([school_id, program_id, department_id]):
        print("DEBUG: WARNING - No query parameters detected!")
        print("DEBUG: This usually means the button navigation is not working properly")
        print("DEBUG: Trying to get current URL from browser...")
    else:
        print("DEBUG: Query parameters detected successfully")
    
    # Enhanced role detection with better error handling and validation
    default_role = ''
    context_info = ''
    
    # Debug logging for troubleshooting
    print(f"DEBUG: Query params - school_id: '{school_id}', program_id: '{program_id}', department_id: '{department_id}'")
    
    # Check program_id first (highest priority)
    if program_id and program_id.strip() and program_id.strip().lower() not in ['none', 'null', '']:
        try:
            program_doc = programs_col.find_one({'_id': ObjectId(program_id)})
            if program_doc:
                program_name = program_doc.get('name', 'Unknown Program')
                default_role = 'Program Admin'
                context_info = f'Program: {program_name}'
                print(f"DEBUG: Detected Program Admin for: {program_name}")
            else:
                print(f"DEBUG: Program ID {program_id} not found in database")
                default_role = 'Program Admin'
                context_info = 'Program: Invalid Program ID'
        except Exception as e:
            print(f"DEBUG: Error processing program_id {program_id}: {e}")
            default_role = 'Program Admin'
            context_info = 'Program: Invalid Program ID'
    
    # Check department_id second
    elif department_id and department_id.strip() and department_id.strip().lower() not in ['none', 'null', '']:
        try:
            dept_doc = schools_col.find_one({'_id': ObjectId(department_id), 'type': 'department'})
            if dept_doc:
                dept_name = dept_doc.get('name', 'Unknown Department')
                default_role = 'Department Admin'
                context_info = f'Department: {dept_name}'
                print(f"DEBUG: Detected Department Admin for: {dept_name}")
            else:
                print(f"DEBUG: Department ID {department_id} not found in database")
                default_role = 'Department Admin'
                context_info = 'Department: Invalid Department ID'
        except Exception as e:
            print(f"DEBUG: Error processing department_id {department_id}: {e}")
            default_role = 'Department Admin'
            context_info = 'Department: Invalid Department ID'
    
    # Check school_id third
    elif school_id and school_id.strip() and school_id.strip().lower() not in ['none', 'null', '']:
        try:
            school_doc = schools_col.find_one({'_id': ObjectId(school_id), 'type': {'$ne': 'department'}})
            if school_doc:
                school_name = school_doc.get('name', 'Unknown School')
                default_role = 'School Admin'
                context_info = f'School: {school_name}'
                print(f"DEBUG: Detected School Admin for: {school_name}")
            else:
                print(f"DEBUG: School ID {school_id} not found in database")
                default_role = 'School Admin'
                context_info = 'School: Invalid School ID'
        except Exception as e:
            print(f"DEBUG: Error processing school_id {school_id}: {e}")
            default_role = 'School Admin'
            context_info = 'School: Invalid School ID'
    
    # No context detected
    else:
        print("DEBUG: No context detected - no school_id, program_id, or department_id provided")
        default_role = ''
        context_info = ''
    
    print(f"DEBUG: Final role detection - default_role: '{default_role}', context_info: '{context_info}'")
    
    inst = institutions_col.find_one({'_id': ObjectId(inst_id)})
    main_color = inst.get('theme_color', '#667eea')
    
    # Modal form
    with ui.row().style('position: fixed; top: 0; left: 0; width: 100vw; height: 100vh; z-index: 1000; align-items: center; justify-content: center; backdrop-filter: blur(8px); background: rgba(0,0,0,0.18);'):
        with ui.card().style(f'align-items: center; padding: 2.5rem; background: #fff; border-radius: 18px; box-shadow: 0 4px 32px rgba(0,0,0,0.18); min-width: 450px;'):
            ui.label('Create User Group').style(f'font-size: 1.7rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
            
            # Add instruction about role assignment
            ui.label('The role will be automatically assigned based on where you clicked the "Create User Group" button.').style(
                'font-size: 0.9rem; color: #666; text-align: center; margin-bottom: 1.5rem; max-width: 400px; line-height: 1.4;'
            )
            
            # Debug info for troubleshooting (only show if no role detected)
            if not default_role:
                with ui.card().style('background: #fff3cd; border: 1px solid #ffeaa7; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                    ui.label('🔍 Debug Information').style('font-size: 0.9rem; font-weight: bold; color: #856404; margin-bottom: 0.5rem;')
                    ui.label(f'Received Parameters:').style('font-size: 0.8rem; color: #856404; margin-bottom: 0.3rem;')
                    ui.label(f'• school_id: "{school_id}"').style('font-size: 0.8rem; color: #856404; font-family: monospace;')
                    ui.label(f'• program_id: "{program_id}"').style('font-size: 0.8rem; color: #856404; font-family: monospace;')
                    ui.label(f'• department_id: "{department_id}"').style('font-size: 0.8rem; color: #856404; font-family: monospace;')
                    ui.label(f'• Full URL: {full_query}').style('font-size: 0.8rem; color: #856404; font-family: monospace; margin-top: 0.5rem;')
                    
                    # Add a button to manually refresh and try to get URL from browser
                    ui.button('🔄 Refresh & Retry', on_click=lambda: ui.run_javascript('window.location.reload()')).style(
                        'background: #856404; color: white; padding: 0.5rem 1rem; border-radius: 4px; border: none; margin-top: 0.5rem;'
                    )
            
            # Form fields
            first_name = ui.input('First Name').style('width: 350px; margin-bottom: 1rem;')
            last_name = ui.input('Last Name').style('width: 350px; margin-bottom: 1rem;')
            email = ui.input('Email').props('type=email').style('width: 350px; margin-bottom: 1rem;')

            # Simple role display
            if default_role and context_info:
                # Extract entity name from context_info
                if ': ' in context_info:
                    entity_name = context_info.split(': ', 1)[1]
                else:
                    entity_name = context_info
                
                ui.label(f"Role: {default_role} :- {entity_name}").style('font-size: 1.1rem; font-weight: bold; color: #1976d2; margin-bottom: 1rem; text-align: center;')
            else:
                ui.label('Role: Not detected').style('font-size: 1.1rem; font-weight: bold; color: #dc3545; margin-bottom: 1rem; text-align: center;')

            def submit_user():
                # Validate form fields
                if not first_name.value.strip():
                    ui.notify('First Name is required', color='negative')
                    return
                if not last_name.value.strip():
                    ui.notify('Last Name is required', color='negative')
                    return
                if not email.value.strip():
                    ui.notify('Email is required', color='negative')
                    return
                
                # Validate role detection
                if not default_role:
                    ui.notify('Role could not be detected. Please ensure you clicked the correct button.', color='negative')
                    return
                
                # Check academic year
                selected_academic_year_id = app.storage.user.get('selected_academic_year_id') if app.storage.user else None
                if not selected_academic_year_id:
                    ui.notify('Please select an academic year first', color='negative')
                    return
                
                # Show confirmation with role details
                confirm_message = f"Create user '{first_name.value.strip()} {last_name.value.strip()}' as {default_role}?"
                if context_info and ': ' in context_info:
                    entity_name = context_info.split(': ', 1)[1]
                    confirm_message += f" for {entity_name}?"
                
                # Create the user
                try:
                    result, raw_password = create_user(
                        email=email.value.strip(),
                        role=default_role,
                        institution_id=inst_id,
                        first_name=first_name.value.strip(),
                        last_name=last_name.value.strip(),
                        school_id=school_id if school_id else None,
                        program_id=program_id if program_id else None,
                        department_id=department_id if department_id else None,
                        academic_cycle_id=selected_academic_year_id
                    )
                    
                    # Success notification with role details and password
                    success_msg = f"User created successfully as {default_role}!"
                    if context_info and ': ' in context_info:
                        entity_name = context_info.split(': ', 1)[1]
                        success_msg += f" for {entity_name}"
                    
                    # Show password in a dialog
                    with ui.dialog() as password_dialog:
                        with ui.card().style('padding: 2rem; min-width: 500px; text-align: center;'):
                            ui.label('✅ User Created Successfully!').style(
                                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                            )
                            
                            ui.label(f'User: {email.value}').style('font-size: 1.1rem; margin-bottom: 0.5rem;')
                            ui.label(f'Role: {default_role}').style('font-size: 1.1rem; margin-bottom: 0.5rem;')
                            if context_info and ': ' in context_info:
                                ui.label(f'Entity: {entity_name}').style('font-size: 1.1rem; margin-bottom: 1.5rem;')
                            
                            ui.separator().style('margin: 1rem 0;')
                            
                            ui.label('🔑 Default Password').style('font-size: 1.2rem; font-weight: bold; color: var(--warning-color); margin-bottom: 0.5rem;')
                            ui.label('ISU2025IQAC').style('font-size: 1.5rem; font-weight: bold; color: var(--error-color); font-family: monospace; padding: 0.5rem; background: #f8f9fa; border: 2px solid var(--border); border-radius: 6px; margin-bottom: 1rem;')
                            
                            ui.label('⚠️ User must change password on first login').style('font-size: 0.9rem; color: var(--warning-color); margin-bottom: 1.5rem;')
                            
                            with ui.row().style('gap: 1rem; justify-content: center;'):
                                ui.button('OK', on_click=password_dialog.close).style(
                                    f'background: {main_color}; color: white; padding: 0.75rem 2rem; border-radius: 8px; border: none; font-weight: 600;'
                                )
                                ui.button('Go to Hierarchy', on_click=lambda: ui.navigate.to(f'/institution_admin/{inst_id}/hierarchy')).style(
                                    f'background: #28a745; color: white; padding: 0.75rem 2rem; border-radius: 8px; border: none; font-weight: 600;'
                                )
                    
                    password_dialog.open()
                    ui.notify(success_msg, color='positive')
                    # Don't navigate immediately - let user see the password first
                    # ui.navigate.to(f'/institution_admin/{inst_id}/hierarchy')
                    
                except Exception as e:
                    ui.notify(f'Error creating user: {str(e)}', color='negative')
                
            # Form status indicator
            if default_role and context_info:
                ui.label('✅ Form is ready to submit').style('color: #4caf50; font-size: 0.9rem; margin-top: 1rem; text-align: center;')
            else:
                ui.label('❌ Cannot submit - role not detected').style('color: #f44336; font-size: 0.9rem; margin-top: 1rem; text-align: center;')
                
                # Additional help for debugging
                ui.label('💡 Try refreshing the page or going back to the hierarchy page').style('color: #856404; font-size: 0.8rem; margin-top: 0.5rem; text-align: center;')
                ui.label('Make sure you clicked the button from the correct entity (School/Program/Department)').style('color: #856404; font-size: 0.8rem; margin-top: 0.3rem; text-align: center;')
            
            ui.button('Create User', on_click=submit_user).style(f'background: {main_color}; color: #fff; width: 350px; margin-top: 1rem; font-size: 1.1rem;')
            ui.button('Cancel', on_click=lambda: ui.navigate.to(f'/institution_admin/{inst_id}/hierarchy')).style(f'background: #6c757d; color: #fff; width: 350px; margin-top: 0.5rem; font-size: 1.1rem;')

# Sidebar for Institution Admin
def institution_admin_sidebar(inst_id, content_func):
    """Create institution admin sidebar with themed styling"""
    global current_user
    current_user = app.storage.user.get('user')
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Initialize app.storage.user if needed
    if not hasattr(app.storage, 'user'):
        app.storage.user = {}
    
    from bson import ObjectId
    inst = institutions_col.find_one({'_id': ObjectId(inst_id)})
    if not inst:
        ui.label('Institution not found').style('font-size: 1.2rem; color: #c00; margin-top: 2rem;')
        return
    
    # Set theme colors
    main_color = inst.get('theme_color', '#667eea')
    set_theme_colors(main_color)
    light_color = lighten_color(main_color, 0.8)
    dark_color = darken_color(main_color, 0.1)
    
    logo_url = inst.get('logo') or 'https://ui-avatars.com/api/?name=' + (inst.get('name') or 'Institution')
    
    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
        # Enhanced Sidebar with Proper Styling
        with ui.column().classes('sidebar').style(f'min-width: 280px; background: var(--surface); border-right: 2px solid var(--border); height: 100vh; overflow-y: auto; padding: 0;'):
            # Institution header with border and user info
            with ui.card().style(f'background: linear-gradient(135deg, {main_color} 0%, {dark_color} 100%); color: white; margin: 1rem; padding: 1.5rem; border-radius: 12px; border: none;'):
                with ui.row().style('align-items: center; margin-bottom: 1rem;'):
                    ui.image(logo_url).style('width: 60px; height: 60px; border-radius: 10px; margin-right: 1rem; border: 2px solid rgba(255,255,255,0.3);')
                    with ui.column():
                        ui.label(inst.get('name', 'Institution')).style('font-size: 1.2rem; font-weight: bold; color: white; margin-bottom: 0.2rem;')
                        ui.label('Admin Portal').style('font-size: 0.9rem; color: rgba(255,255,255,0.8);')
                
                # User Info Section below the logo
                if current_user:
                    ui.separator().style('background: rgba(255,255,255,0.3); margin: 1rem 0;')
                    ui.label(f"👤 {current_user.get('email','')}").style('font-size: 0.9rem; color: rgba(255,255,255,0.9); margin-bottom: 0.3rem; text-align: left; font-weight: 500;')
                    ui.label(f"Role: {current_user.get('role','')}").style('font-size: 0.85rem; color: rgba(255,255,255,0.8); text-align: left; font-weight: 400;')
            
            # Navigation menu with bordered sections
            with ui.column().style('padding: 0 1rem; gap: 1.5rem;'):
                # Overview Section
                with ui.card().style('background: #f8f9fa; border: 2px solid #e9ecef; border-radius: 10px; padding: 1rem;'):
                    ui.label('📊 OVERVIEW').style('font-size: 0.85rem; font-weight: 700; color: rgb(154, 44, 84); margin-bottom: 1rem; text-align: left; letter-spacing: 1px; text-transform: uppercase;')
                    overview_items = [
                        ('🏠', 'Dashboard', f'/institution_admin/{inst_id}'),
                        ('🏛️', 'Institution Details', f'/institution_admin/{inst_id}/details'),
                        ('🏢', 'Institutional Hierarchy', f'/institution_admin/{inst_id}/hierarchy'),
                        ('📅', 'Academic Years', f'/institution_admin/{inst_id}/academic_years'),
                        ('📝', 'Create Criteria', f'/institution_admin/{inst_id}/criterias'),
                        ('👤', 'Extended Profiles', f'/institution_admin/{inst_id}/extended_profiles'),
                        ('📊', 'Spreadsheet', f'/institution_admin/{inst_id}/spreadsheets'),
                        ('📥', 'Submissions', f'/institution_admin/{inst_id}/submissions'),
                        ('👥', 'Manage Users', f'/institution_admin/{inst_id}/users'),
                        ('📋', 'Audit Logs', f'/institution_admin/{inst_id}/audit_logs'),
                    ]
                    
                    for icon, label, url in overview_items:
                        ui.button(f'{icon} {label}', on_click=lambda u=url: ui.navigate.to(u)).style(
                            f'width: 100%; justify-content: flex-start; margin-bottom: 0.5rem; text-align: left; '
                            f'background: white; color: var(--text-primary); border: 1px solid #e9ecef; '
                            f'padding: 0.75rem 1rem; border-radius: 8px; '
                            f'transition: all 0.3s ease; font-weight: 500;'
                        ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "rgb(154, 44, 84)"; event.target.style.color = "white"; event.target.style.borderColor = "rgb(154, 44, 84)";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "white"; event.target.style.color = "var(--text-primary)"; event.target.style.borderColor = "#e9ecef";'))
                    
                    # Logout button
                    ui.separator().style('margin: 1rem 0; background: #e9ecef;')
                    ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                        f'width: 100%; justify-content: center; background: #dc3545; color: white; border: 1px solid #dc3545; '
                        f'padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 600;'
                    ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "#c82333"; event.target.style.borderColor = "#c82333";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "#dc3545"; event.target.style.borderColor = "#dc3545";'))

        
        # Main content area
        with ui.column().classes('main-content').style('flex: 1; padding: 2rem; overflow-y: auto;'):
            content_func(inst, main_color)

@ui.page('/institution_admin/{inst_id}')
def institution_admin_dashboard(inst_id: str):
    """Institution admin dashboard"""
    add_beautiful_global_styles()
    
    def content(inst, main_color):
        ui.label('Institution Admin Dashboard').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Quick stats cards
        with ui.row().style('width: 100%; gap: 1rem; margin-bottom: 2rem;'):
            # Academic Years count
            years_count = academic_years_col.count_documents({'institution_id': inst_id})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('📅').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(years_count)).style('font-size: 2rem; font-weight: bold; color: var(--primary-color);')
                ui.label('Academic Years').style('color: var(--text-secondary);')
            
            # Schools count
            schools_count = schools_col.count_documents({'institution_id': inst_id})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('🏫').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(schools_count)).style('font-size: 2rem; font-weight: bold; color: var(--success-color);')
                ui.label('Schools').style('color: var(--text-secondary);')
            
            # Programs count
            programs_count = programs_col.count_documents({'institution_id': inst_id})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('🎓').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(programs_count)).style('font-size: 2rem; font-weight: bold; color: var(--info-color);')
                ui.label('Programs').style('color: var(--text-secondary);')
            
            # Criterias count
            criterias_count = criterias_col.count_documents({'institution_id': inst_id})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('📊').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(criterias_count)).style('font-size: 2rem; font-weight: bold; color: var(--warning-color);')
                ui.label('Criterias').style('color: var(--text-secondary);')
            
            # Extended Profiles count
            profiles_count = extended_profiles_col.count_documents({'institution_id': inst_id})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('📝').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(profiles_count)).style('font-size: 2rem; font-weight: bold; color: var(--primary-color);')
                ui.label('Extended Profiles').style('color: var(--text-secondary);')
        
        # Recent activity
        with ui.card().classes('beautiful-card').style('width: 100%; padding: 1.5rem;'):
            ui.label('Recent Activity').style('font-size: 1.5rem; font-weight: bold; margin-bottom: 1rem;')
            
            recent_logs = list(audit_logs_col.find(
                {'institution_id': inst_id}
            ).sort('timestamp', -1).limit(5))
            
            if recent_logs:
                for log in recent_logs:
                    with ui.row().style('align-items: center; padding: 0.5rem 0; border-bottom: 1px solid var(--border);'):
                        ui.label('•').style('color: var(--primary-color); margin-right: 0.5rem;')
                        ui.label(log.get('action', 'Unknown Action')).style('font-weight: 500;')
                        ui.label(log.get('details', '')).style('color: var(--text-secondary); margin-left: 1rem;')
                        if log.get('timestamp'):
                            ui.label(log['timestamp'].strftime('%Y-%m-%d %H:%M')).style(
                                'margin-left: auto; color: var(--text-secondary); font-size: 0.9rem;'
                            )
            else:
                ui.label('No recent activity').style('color: var(--text-secondary); font-style: italic;')
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/criterias')
def institution_admin_criterias(inst_id: str):
    """Enhanced criterias management page"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    def content(inst, main_color):
        ui.label('Criterias Management').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
        )
        
        # Academic year selector
        from nicegui import app
        unlocked_years = list(academic_years_col.find({'institution_id': inst_id, 'is_locked': False}))
        year_options = [y['name'] for y in unlocked_years]
        name_to_id = {y['name']: str(y['_id']) for y in unlocked_years}
        
        # Initialize session if needed
        if not hasattr(app.storage, 'user'):
            app.storage.user = {}
        
        selected_year_id = app.storage.user.get('selected_academic_year_id') if hasattr(app.storage, 'user') else None
        
        def on_year_change(e):
            if hasattr(app.storage, 'user'):
                app.storage.user['selected_academic_year_id'] = name_to_id.get(e.value)
            # Automatically refresh the page content
            ui.run_javascript('window.location.reload()')
        
        with ui.row().style('width: 100%; align-items: center; justify-content: flex-end; margin-bottom: 1rem;'):
            if year_options:
                selected_year_name = None
                for name, id_val in name_to_id.items():
                    if id_val == selected_year_id:
                        selected_year_name = name
                        break
                
                ui.select(
                    options=year_options,
                    value=selected_year_name,
                    on_change=on_year_change,
                    label='Academic Year'
                ).style('min-width: 200px;').classes('beautiful-input')
        
        # Create criteria button
        def open_create_dialog():
            if not selected_year_id:
                ui.notify('Select an academic year first', color='negative')
                return
            
            with ui.dialog() as dialog:
                with ui.card().style(
                    'padding: 2rem; min-width: 800px; max-width: 1000px; max-height: 90vh; overflow-y: auto;'
                ).classes('beautiful-card'):
                    ui.label('Create New Criteria').style(
                        f'font-size: 1.8rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem; text-align: center;'
                    )
                    
                    with ui.column().style('width: 100%; gap: 1.5rem;'):
                        # Basic Information
                        ui.label('Basic Information').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem;'
                        )
                        
                        criteria_name = ui.input('Criteria Name').classes('beautiful-input').style('width: 100%;')
                        criteria_description = ui.textarea('Description').classes('beautiful-input').style('width: 100%; min-height: 100px;')
                        
                        # Criteria Type Selection
                        ui.label('Criteria Type').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem; margin-top: 1rem;'
                        )
                        
                        scope_type = ui.select(
                            options=['Program-based', 'Department-based'],
                            value='Program-based',
                            label='Choose whether this criteria applies to programs or departments'
                        ).classes('beautiful-input').style('width: 100%; margin-bottom: 1rem;')
                        
                        ui.label('Choose whether this criteria applies to programs or departments.').style(
                            'color: var(--text-secondary); font-size: 0.9rem; margin-top: -0.5rem; margin-bottom: 1rem;'
                        )
                        
                        # Deadline
                        deadline_input = ui.input('Deadline').props('type=date').classes('beautiful-input').style('width: 100%; margin-bottom: 1rem;')
                        
                        # Headers/Fields
                        ui.label('Spreadsheet Headers').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem;'
                        )
                        
                        headers_input = ui.textarea(
                            'Headers (one per line)',
                            placeholder='ID\nName\nDescription\nScore\nRemarks'
                        ).classes('beautiful-input').style('width: 100%; min-height: 120px;')
                        
                        # Table Preview
                        ui.label('Criteria Preview').style(
                            f'font-size: 1.1rem; font-weight: bold; color: {main_color}; margin: 1.5rem 0 0.5rem 0;'
                        )
                        
                        preview_container = ui.column().style('width: 100%;')
                        
                        def update_preview():
                            if not headers_input.value:
                                preview_container.clear()
                                with preview_container:
                                    ui.label('Enter headers above to see table preview').style('color: #666; font-style: italic; text-align: center;')
                                return
                                
                            preview_container.clear()
                            headers = [h.strip() for h in headers_input.value.split('\n') if h.strip()]
                            
                            if headers:
                                with preview_container:
                                    # Show scope type indicator
                                    ui.label(f'Scope: {scope_type.value}').style(f'color: {main_color}; font-weight: bold; margin-bottom: 0.5rem;')
                                    
                                    with ui.card().style('width: 100%; background: white; border: 2px solid rgba(154, 44, 84, 0.2); padding: 1rem; border-radius: 8px;'):
                                        # Create table with proper styling
                                        with ui.element('table').style('width: 100%; border-collapse: collapse;'):
                                            # Header row
                                            with ui.element('thead'):
                                                with ui.element('tr'):
                                                    for header in headers:
                                                        with ui.element('th').style('background: rgb(154, 44, 84); color: white; padding: 12px; text-align: left; border: 1px solid #ddd; font-weight: bold;'):
                                                            ui.label(header)
                                            
                                            # Sample data rows
                                            with ui.element('tbody'):
                                                sample_data = [
                                                    ['1', 'Sample Item 1', 'Description 1', '85', 'Good'],
                                                    ['2', 'Sample Item 2', 'Description 2', '92', 'Excellent'],
                                                    ['3', 'Sample Item 3', 'Description 3', '78', 'Satisfactory']
                                                ]
                                                
                                                for i, row_data in enumerate(sample_data):
                                                    with ui.element('tr'):
                                                        for j, header in enumerate(headers):
                                                            sample_value = row_data[j] if j < len(row_data) else f'Sample {i+1}-{j+1}'
                                                            with ui.element('td').style('padding: 10px; border: 1px solid #ddd; background: #f9f9f9;'):
                                                                ui.label(sample_value).style('font-size: 0.9rem;')
                                    
                                    ui.label(f'Table will have {len(headers)} columns as shown above').style(
                                        'font-size: 0.8rem; color: #666; margin-top: 0.5rem; text-align: center;'
                                    )
                        
                        # Real-time update on input change
                        headers_input.on_value_change(lambda: update_preview())
                        scope_type.on_value_change(lambda: update_preview())
                        criteria_name.on_value_change(lambda: update_preview())
                        
                        # Initialize preview
                        update_preview()
                        
                        # Options
                        with ui.row().style('width: 100%; gap: 2rem; margin-top: 1rem;'):
                            needs_docs = ui.checkbox('Needs Supporting Documents', value=True)
                        
                        # Create function
                        def create_criteria():
                            global current_user
                            if not criteria_name.value:
                                ui.notify('Criteria name is required', color='negative')
                                return
                            
                            if not headers_input.value:
                                ui.notify('At least one header is required', color='negative')
                                return
                            
                            # Parse headers
                            headers = [h.strip() for h in headers_input.value.split('\n') if h.strip()]
                            
                            # Determine scope type and department_id
                            scope_val = 'program_based' if scope_type.value == 'Program-based' else 'department_based'
                            dept_id = None  # For now, we'll implement department selection later
                            
                            # Parse deadline
                            deadline_val = None
                            if deadline_input.value:
                                try:
                                    deadline_val = datetime.strptime(deadline_input.value, '%Y-%m-%d')
                                except:
                                    pass
                            
                            criteria_doc = {
                                'name': criteria_name.value,
                                'description': criteria_description.value,
                                'deadline': deadline_val,
                                'institution_id': inst_id,
                                'academic_cycle_id': selected_year_id,
                                'scope_type': scope_val,
                                'department_id': dept_id,
                                'headers': headers,
                                'needs_supporting_docs': needs_docs.value,
                                'created_at': datetime.now(timezone.utc),
                                'updated_at': datetime.now(timezone.utc),
                                'created_by': current_user.get('email', 'admin') if current_user else 'admin'
                            }
                            
                            try:
                                result = criterias_col.insert_one(criteria_doc)
                                
                                log_audit_action(
                                    action='Created Criteria',
                                    details=f'Criteria "{criteria_name.value}" created with scope: {scope_val}',
                                    institution_id=inst_id,
                                    entity_type='criteria',
                                    entity_id=str(result.inserted_id)
                                )
                                
                                ui.notify(f'Criteria "{criteria_name.value}" created successfully!', color='positive')
                                dialog.close()
                                # Force page refresh instead of navigation to avoid flicker
                                ui.run_javascript('window.location.reload()')
                            
                            except Exception as e:
                                ui.notify(f'Error creating criteria: {str(e)}', color='negative')
                        
                        # Action buttons
                        with ui.row().style('width: 100%; justify-content: space-between; margin-top: 2rem;'):
                            ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                            ui.button('Create Criteria', on_click=create_criteria).classes('btn-primary')
                
                dialog.open()
        
        ui.button('+ Create New Criteria', on_click=open_create_dialog).classes('btn-primary').style('margin-bottom: 2rem;')
        
        # Display existing criterias
        if selected_year_id:
            existing_criterias = list(criterias_col.find({
                'institution_id': inst_id,
                'academic_cycle_id': selected_year_id
            }))
            
            if existing_criterias:
                ui.label(f'Existing Criterias ({len(existing_criterias)} found)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                
                for criteria in existing_criterias:
                    with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1rem; padding: 1.5rem;'):
                        with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                            with ui.column().style('flex: 1;'):
                                criteria_name = criteria.get('name', 'Unnamed Criteria')
                                scope_type = criteria.get('scope_type', 'program_based')
                                scope_label = 'Program-based' if scope_type == 'program_based' else 'Department-based'
                                
                                ui.label(f"📊 {criteria_name}").style(
                                    'font-size: 1.2rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                                )
                                ui.label(f"Type: {scope_label}").style(
                                    'font-size: 0.9rem; color: var(--primary-color); margin-bottom: 0.3rem;'
                                )
                                
                                headers = criteria.get('headers', [])
                                if headers:
                                    ui.label(f"Headers: {', '.join(headers[:3])}{'...' if len(headers) > 3 else ''}").style(
                                        'font-size: 0.8rem; color: var(--text-secondary);'
                                    )
                                
                                if criteria.get('deadline'):
                                    deadline_str = criteria['deadline'].strftime('%Y-%m-%d')
                                    ui.label(f"⏰ Deadline: {deadline_str}").style(
                                        'font-size: 0.9rem; color: var(--warning-color); font-weight: bold;'
                                    )
                                
                                created_at = criteria.get('created_at')
                                if created_at:
                                    ui.label(f"Created: {created_at.strftime('%Y-%m-%d %H:%M')}").style(
                                        'font-size: 0.8rem; color: var(--text-secondary);'
                                    )
                            
                            with ui.row().style('gap: 0.5rem;'):
                                ui.button(
                                    '📈 View Data',
                                    on_click=lambda c_id=str(criteria['_id']): ui.navigate.to(
                                        f'/institution_admin/{inst_id}/spreadsheets?criteria_id={c_id}'
                                    )
                                ).classes('btn-success')
                                
                                # Edit button
                                def edit_criteria(criteria_id=str(criteria['_id']), crit=criteria):
                                    with ui.dialog() as edit_dialog:
                                        with ui.card().style('padding: 2rem; min-width: 600px;').classes('beautiful-card'):
                                            ui.label('Edit Criteria').style(
                                                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                                            )
                                            
                                            edit_name = ui.input('Criteria Name', value=crit.get('name', '')).style('width: 100%; margin-bottom: 1rem;')
                                            edit_description = ui.textarea('Description', value=crit.get('description', '')).style('width: 100%; margin-bottom: 1rem;')
                                            
                                            edit_headers = ui.textarea(
                                                'Headers (one per line)',
                                                value='\n'.join(crit.get('headers', []))
                                            ).style('width: 100%; margin-bottom: 1rem;')
                                            
                                            edit_needs_docs = ui.checkbox('Needs Supporting Documents', value=crit.get('needs_supporting_docs', False))
                                            
                                            def save_changes():
                                                if not edit_name.value:
                                                    ui.notify('Criteria name is required', color='negative')
                                                    return
                                                
                                                headers = [h.strip() for h in edit_headers.value.split('\n') if h.strip()]
                                                
                                                criterias_col.update_one(
                                                    {'_id': ObjectId(criteria_id)},
                                                    {'$set': {
                                                        'name': edit_name.value,
                                                        'description': edit_description.value,
                                                        'headers': headers,
                                                        'needs_supporting_docs': edit_needs_docs.value,
                                                        'updated_at': datetime.datetime.now(datetime.timezone.utc)
                                                    }}
                                                )
                                                
                                                log_audit_action(
                                                    action='Updated Criteria',
                                                    details=f'Criteria "{edit_name.value}" was updated',
                                                    institution_id=inst_id,
                                                    entity_type='criteria',
                                                    entity_id=criteria_id
                                                )
                                                
                                                ui.notify('Criteria updated successfully!', color='positive')
                                                edit_dialog.close()
                                                ui.run_javascript('window.location.reload()')
                                            
                                            with ui.row().style('margin-top: 1rem; gap: 1rem;'):
                                                ui.button('Cancel', on_click=edit_dialog.close).classes('btn-secondary')
                                                ui.button('Save Changes', on_click=save_changes).classes('btn-primary')
                                    
                                    edit_dialog.open()
                                
                                ui.button('✏️ Edit', on_click=edit_criteria).classes('btn-warning')
                                
                                def delete_criteria(criteria_id=str(criteria['_id']), criteria_name=criteria.get('name', '')):
                                    with ui.dialog() as delete_dialog:
                                        with ui.card().style('padding: 2rem;'):
                                            ui.label('Confirm Deletion').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                            ui.label(f'Are you sure you want to delete criteria "{criteria_name}"?').style('margin-bottom: 1rem;')
                                            ui.label('This action cannot be undone.').style('color: red; font-weight: bold; margin-bottom: 1rem;')
                                            
                                            def confirm_delete():
                                                log_audit_action(
                                                    action='Deleted Criteria',
                                                    details=f'Criteria "{criteria_name}" deleted',
                                                    institution_id=inst_id,
                                                    entity_type='criteria',
                                                    entity_id=criteria_id
                                                )
                                                criterias_col.delete_one({'_id': ObjectId(criteria_id)})
                                                ui.notify(f'Criteria "{criteria_name}" deleted', color='positive')
                                                delete_dialog.close()
                                                ui.run_javascript('window.location.reload()')
                                            
                                            with ui.row().style('gap: 1rem;'):
                                                ui.button('Cancel', on_click=delete_dialog.close).classes('btn-secondary')
                                                ui.button('Delete', on_click=confirm_delete).classes('btn-danger')
                                    
                                    delete_dialog.open()
                                
                                ui.button('🗑️ Delete', on_click=delete_criteria).classes('btn-danger')
            else:
                ui.label('No criterias found for the selected academic year.').style(
                    'color: var(--text-secondary); font-style: italic; margin-top: 2rem;'
                )
        else:
            ui.label('Please select an academic year to view criterias.').style(
                'color: var(--text-secondary); font-style: italic; margin-top: 2rem;'
            )
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/details')
def institution_admin_details(inst_id: str):
    """Institution details management page with complete information"""
    add_beautiful_global_styles()
    
    def content(inst, main_color):
        ui.label('Institution Details').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Get complete institution document
        complete_inst = institutions_col.find_one({'_id': ObjectId(inst_id)})
        if not complete_inst:
            ui.label('Institution not found').style('color: var(--error-color); font-size: 1.2rem;')
            return
        
        # Basic Information Card
        with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1.5rem; padding: 2rem;'):
            ui.label('📋 Basic Information').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
            
            with ui.row().style('width: 100%; gap: 3rem;'):
                with ui.column().style('flex: 2;'):
                    # Institution details in a nice grid
                    details = [
                        ('Institution Name', complete_inst.get('name', 'N/A')),
                        ('Website URL', complete_inst.get('website_url', 'N/A')),
                        ('Email', complete_inst.get('email', 'N/A')),
                        ('Phone', complete_inst.get('phone', 'N/A')),
                        ('Address', complete_inst.get('address', 'N/A')),
                        ('City', complete_inst.get('city', 'N/A')),
                        ('State', complete_inst.get('state', 'N/A')),
                        ('Country', complete_inst.get('country', 'N/A')),
                        ('Postal Code', complete_inst.get('postal_code', 'N/A')),
                        ('Institution Type', complete_inst.get('institution_type', 'N/A')),
                        ('Establishment Year', complete_inst.get('establishment_year', 'N/A')),
                        ('Total Students', complete_inst.get('total_students', 'N/A')),
                        ('Total Faculty', complete_inst.get('total_faculty', 'N/A')),
                    ]
                    
                    with ui.element('div').style('display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;'):
                        for label, value in details:
                            with ui.card().style('padding: 1rem; border: 1px solid var(--border); background: #f8f9fa;'):
                                ui.label(label).style('font-weight: bold; color: var(--text-secondary); font-size: 0.9rem; margin-bottom: 0.5rem;')
                                ui.label(str(value)).style('font-size: 1rem; color: var(--text-primary);')
                
                with ui.column().style('flex: 1; align-items: center;'):
                    # Logo and theme
                    if complete_inst.get('logo'):
                        ui.image(complete_inst['logo']).style('width: 150px; height: 150px; border-radius: 12px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); margin-bottom: 1rem;')
                    else:
                        with ui.card().style('width: 150px; height: 150px; display: flex; align-items: center; justify-content: center; background: #f0f0f0; border: 2px dashed #ccc;'):
                            ui.label('No Logo').style('color: #999; font-style: italic;')
                    
                    # Theme color
                    theme_color = complete_inst.get('theme_color', '#667eea')
                    ui.label('Theme Color').style('font-weight: bold; margin-bottom: 0.5rem;')
                    ui.html(f'''
                        <div style="
                            width: 80px; 
                            height: 40px; 
                            background: {theme_color}; 
                            border-radius: 8px; 
                            border: 1px solid #ccc;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            color: white;
                            font-size: 0.8rem;
                            font-weight: bold;
                            margin-bottom: 0.5rem;
                        ">{theme_color}</div>
                    ''')
                    
                    # Edit button
                    ui.button('✏️ Edit Details', on_click=lambda: open_edit_dialog()).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; margin-top: 1rem;')
        
        # Academic Structure Card
        with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1.5rem; padding: 2rem;'):
            ui.label('🎓 Academic Structure').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
            
            # Get schools and programs count
            schools = list(schools_col.find({'institution_id': inst_id}))
            programs = list(programs_col.find({'institution_id': inst_id}))
            
            with ui.row().style('width: 100%; gap: 2rem;'):
                with ui.card().style('flex: 1; padding: 1.5rem; text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;'):
                    ui.label(str(len(schools))).style('font-size: 3rem; font-weight: bold; margin-bottom: 0.5rem;')
                    ui.label('Schools/Departments').style('font-size: 1.1rem;')
                
                with ui.card().style('flex: 1; padding: 1.5rem; text-align: center; background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white;'):
                    ui.label(str(len(programs))).style('font-size: 3rem; font-weight: bold; margin-bottom: 0.5rem;')
                    ui.label('Programs').style('font-size: 1.1rem;')
        
        # System Information Card
        with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1.5rem; padding: 2rem;'):
            ui.label('⚙️ System Information').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
            
            with ui.row().style('width: 100%; gap: 2rem;'):
                with ui.column().style('flex: 1;'):
                    system_details = [
                        ('Institution ID', str(complete_inst.get('_id', 'N/A'))),
                        ('Created Date', complete_inst.get('created_at', datetime.now()).strftime('%Y-%m-%d %H:%M:%S') if complete_inst.get('created_at') else 'N/A'),
                        ('Last Updated', complete_inst.get('updated_at', datetime.now()).strftime('%Y-%m-%d %H:%M:%S') if complete_inst.get('updated_at') else 'N/A'),
                        ('Status', complete_inst.get('status', 'Active')),
                        ('NAAC Accreditation', complete_inst.get('naac_grade', 'Not Available')),
                        ('NBA Accreditation', complete_inst.get('nba_accreditation', 'Not Available')),
                    ]
                    
                    for label, value in system_details:
                        with ui.row().style('width: 100%; align-items: center; padding: 0.75rem; border-bottom: 1px solid var(--border);'):
                            ui.label(label).style('font-weight: bold; color: var(--text-secondary); min-width: 150px;')
                            ui.label(str(value)).style('color: var(--text-primary);')
        
        # Additional Data Card (if any extra fields exist)
        extra_fields = {k: v for k, v in complete_inst.items() 
                       if k not in ['_id', 'name', 'website_url', 'email', 'phone', 'address', 'city', 'state', 'country', 
                                   'postal_code', 'institution_type', 'establishment_year', 'total_students', 'total_faculty',
                                   'logo', 'theme_color', 'created_at', 'updated_at', 'status', 'naac_grade', 'nba_accreditation']}
        
        if extra_fields:
            with ui.card().classes('beautiful-card').style('width: 100%; padding: 2rem;'):
                ui.label('📝 Additional Information').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
                
                for key, value in extra_fields.items():
                    with ui.row().style('width: 100%; align-items: center; padding: 0.75rem; border-bottom: 1px solid var(--border);'):
                        ui.label(str(key).replace('_', ' ').title()).style('font-weight: bold; color: var(--text-secondary); min-width: 200px;')
                        ui.label(str(value)[:200] + ('...' if len(str(value)) > 200 else '')).style('color: var(--text-primary);')
        
        def open_edit_dialog():
            with ui.dialog() as dialog:
                with ui.card().style('padding: 2rem; min-width: 600px; max-height: 80vh; overflow-y: auto;'):
                    ui.label('Edit Institution Details').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
                    
                    # Create input fields for editable data
                    inputs = {}
                    editable_fields = [
                        ('name', 'Institution Name'),
                        ('website_url', 'Website URL'),
                        ('email', 'Email'),
                        ('phone', 'Phone'),
                        ('address', 'Address'),
                        ('city', 'City'),
                        ('state', 'State'),
                        ('country', 'Country'),
                        ('postal_code', 'Postal Code'),
                        ('institution_type', 'Institution Type'),
                        ('establishment_year', 'Establishment Year'),
                        ('total_students', 'Total Students'),
                        ('total_faculty', 'Total Faculty'),
                    ]
                    
                    for field, label in editable_fields:
                        inputs[field] = ui.input(label, value=str(complete_inst.get(field, ''))).style('width: 100%; margin-bottom: 1rem;')
                    
                    def save_changes():
                        try:
                            update_data = {}
                            for field in inputs:
                                value = inputs[field].value.strip()
                                if field in ['establishment_year', 'total_students', 'total_faculty']:
                                    try:
                                        update_data[field] = int(value) if value else None
                                    except ValueError:
                                        update_data[field] = value
                                else:
                                    update_data[field] = value
                            
                            update_data['updated_at'] = datetime.now()
                            
                            institutions_col.update_one(
                                {'_id': ObjectId(inst_id)},
                                {'$set': update_data}
                            )
                            
                            ui.notify('Institution details updated successfully!', color='positive')
                            dialog.close()
                            ui.run_javascript('window.location.reload()')
                        except Exception as e:
                            ui.notify(f'Error updating details: {str(e)}', color='negative')
                    
                    with ui.row().style('gap: 1rem; margin-top: 1.5rem;'):
                        ui.button('Cancel', on_click=dialog.close).style('background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                        ui.button('Save Changes', on_click=save_changes).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
            
            dialog.open()
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/hierarchy')
def institution_admin_hierarchy(inst_id: str):
    """Schools and programs management page"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    def content(inst, main_color):
        ui.label('Institution Hierarchy').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Academic year selector
        from nicegui import app
        unlocked_years = list(academic_years_col.find({'institution_id': inst_id, 'is_locked': False}))
        year_options = [y['name'] for y in unlocked_years]
        name_to_id = {y['name']: str(y['_id']) for y in unlocked_years}
        
        # Initialize session if needed
        if not hasattr(app.storage, 'user'):
            app.storage.user = {}
        
        selected_year_id = app.storage.user.get('selected_academic_year_id') if hasattr(app.storage, 'user') else None
        
        def on_year_change(e):
            if hasattr(app.storage, 'user'):
                app.storage.user['selected_academic_year_id'] = name_to_id.get(e.value)
            # Automatically refresh the page content
            ui.run_javascript('window.location.reload()')
        
        with ui.row().style('width: 100%; align-items: center; justify-content: flex-end; margin-bottom: 1rem;'):
            if year_options:
                selected_year_name = None
                for name, id_val in name_to_id.items():
                    if id_val == selected_year_id:
                        selected_year_name = name
                        break
                
                ui.select(
                    options=year_options,
                    value=selected_year_name,
                    on_change=on_year_change,
                    label='Academic Year'
                ).style('min-width: 200px;').classes('beautiful-input')
        
        # Institution header with add options
        with ui.card().classes('beautiful-card').style(f'width: 100%; margin-bottom: 2rem; padding: 1.5rem; background: {main_color}10; border-left: 4px solid {main_color};'):
            with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                with ui.column():
                    ui.label(f"🏛️ {inst.get('name', 'Institution')}").style(f'font-size: 1.3rem; font-weight: bold; color: {main_color};')
                    ui.label('Institution Level').style('font-size: 0.9rem; color: var(--text-secondary);')
                
                # Add buttons
                with ui.row().style('gap: 0.5rem;'):
                    def add_school():
                        if not selected_year_id:
                            ui.notify('Please select an academic year first', color='warning')
                            return
                        
                        with ui.dialog() as dialog:
                            with ui.card().style('padding: 2rem; min-width: 400px;').classes('beautiful-card'):
                                ui.label('Add New School').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;')
                                
                                school_name = ui.input('School Name', placeholder='e.g., School of Engineering').style('width: 100%; margin-bottom: 1rem;')
                                school_abbr = ui.input('Abbreviation (Optional)', placeholder='e.g., SOE').style('width: 100%; margin-bottom: 1rem;')
                                
                                def create_school():
                                    if not school_name.value:
                                        ui.notify('School name is required', color='negative')
                                        return
                                    
                                    school_doc = {
                                        'name': school_name.value,
                                        'abbreviation': school_abbr.value,
                                        'institution_id': inst_id,
                                        'academic_cycle_id': selected_year_id,
                                        'created_at': datetime.now(timezone.utc),
                                        'created_by': current_user.get('email', 'admin') if current_user else 'admin'
                                    }
                                    
                                    result = schools_col.insert_one(school_doc)
                                    
                                    log_audit_action(
                                        action='Created School',
                                        details=f'School "{school_name.value}" created',
                                        institution_id=inst_id,
                                        entity_type='school',
                                        entity_id=str(result.inserted_id)
                                    )
                                    
                                    ui.notify(f'School "{school_name.value}" created successfully!', color='positive')
                                    dialog.close()
                                    ui.run_javascript('window.location.reload()')
                                
                                with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                    ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                                    ui.button('Create School', on_click=create_school).classes('btn-primary')
                        
                        dialog.open()
                    
                    def add_department():
                        if not selected_year_id:
                            ui.notify('Please select an academic year first', color='warning')
                            return
                        
                        with ui.dialog() as dialog:
                            with ui.card().style('padding: 2rem; min-width: 400px;').classes('beautiful-card'):
                                ui.label('Add New Department').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;')
                                
                                dept_name = ui.input('Department Name', placeholder='e.g., Human Resources').style('width: 100%; margin-bottom: 1rem;')
                                dept_abbr = ui.input('Abbreviation (Optional)', placeholder='e.g., HR').style('width: 100%; margin-bottom: 1rem;')
                                
                                def create_department():
                                    if not dept_name.value:
                                        ui.notify('Department name is required', color='negative')
                                        return
                                    
                                    dept_doc = {
                                        'name': dept_name.value,
                                        'abbreviation': dept_abbr.value,
                                        'institution_id': inst_id,
                                        'academic_cycle_id': selected_year_id,
                                        'type': 'department',
                                        'created_at': datetime.now(timezone.utc),
                                        'created_by': current_user.get('email', 'admin') if current_user else 'admin'
                                    }
                                    
                                    result = schools_col.insert_one(dept_doc)
                                    
                                    log_audit_action(
                                        action='Created Department',
                                        details=f'Department "{dept_name.value}" created',
                                        institution_id=inst_id,
                                        entity_type='department',
                                        entity_id=str(result.inserted_id)
                                    )
                                    
                                    ui.notify(f'Department "{dept_name.value}" created successfully!', color='positive')
                                    dialog.close()
                                    ui.run_javascript('window.location.reload()')
                                
                                with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                    ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                                    ui.button('Create Department', on_click=create_department).classes('btn-primary')
                        
                        dialog.open()
                    
                    ui.button('+ Add School', on_click=add_school).classes('btn-primary')
                    ui.button('+ Add Department', on_click=add_department).classes('btn-success')
        
        if selected_year_id:
            # Get schools and departments
            schools = list(schools_col.find({'institution_id': inst_id, 'academic_cycle_id': selected_year_id, 'type': {'$ne': 'department'}}))
            programs = list(programs_col.find({'institution_id': inst_id, 'academic_cycle_id': selected_year_id}))
            departments = list(schools_col.find({'institution_id': inst_id, 'type': 'department', 'academic_cycle_id': selected_year_id}))
            
            # Display Schools
            for school in schools:
                with ui.card().classes('beautiful-card drop-zone').style(f'width: 100%; margin-bottom: 1rem; padding: 1.5rem; border-left: 4px solid {main_color};').classes(f'school-{school["_id"]}'):
                    with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                        with ui.column():
                            school_name = school.get('name', 'Unnamed School')
                            school_abbr = school.get('abbreviation', '')
                            ui.label(f"🏫 {school_name} {f'({school_abbr})' if school_abbr else ''}").style(
                                f'font-size: 1.2rem; font-weight: bold; color: {main_color};'
                            )
                            ui.label('School').style('font-size: 0.9rem; color: var(--text-secondary);')
                        
                        with ui.row().style('gap: 0.5rem;'):
                            # Add Program button
                            def add_program_to_school(school_id=str(school['_id']), school_name=school_name):
                                with ui.dialog() as dialog:
                                    with ui.card().style('padding: 2rem; min-width: 500px;').classes('beautiful-card'):
                                        ui.label(f'Add Program to {school_name}').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;')
                                        
                                        program_name = ui.input('Program Name', placeholder='e.g., Computer Science Engineering').style('width: 100%; margin-bottom: 1rem;')
                                        program_abbr = ui.input('Abbreviation (Optional)', placeholder='e.g., CSE').style('width: 100%; margin-bottom: 1rem;')
                                        
                                        program_type = ui.select(
                                            options=['Undergraduate', 'Postgraduate', 'Diploma', 'Certificate'],
                                            value='Undergraduate',
                                            label='Program Type'
                                        ).style('width: 100%; margin-bottom: 1rem;')
                                        
                                        program_duration = ui.select(
                                            options=['1 Year', '2 Years', '3 Years', '4 Years', '5 Years', '6 Months'],
                                            value='4 Years',
                                            label='Duration'
                                        ).style('width: 100%; margin-bottom: 1rem;')
                                        
                                        # Preview of generated code (prominent display)
                                        with ui.card().style('background: #f8f9fa; border: 2px solid #e9ecef; padding: 1rem; margin-bottom: 1rem;'):
                                            ui.label('Program Code (Auto-Generated):').style('font-size: 0.9rem; color: #666; font-weight: bold; margin-bottom: 0.5rem;')
                                            preview_code = ui.label('Enter program name to see code preview').style('font-size: 1.2rem; color: #2e7d32; font-weight: bold; font-family: monospace;')
                                        
                                        def update_code_preview():
                                            if program_name.value and selected_year_id:
                                                try:
                                                    year_doc = academic_years_col.find_one({'_id': ObjectId(selected_year_id)})
                                                    year_name = year_doc.get('name', '') if year_doc else ''
                                                    type_short = program_type.value.split(' ')[0] if program_type.value else 'UG'
                                                    
                                                    code = generate_program_code(inst_id, year_name, type_short, program_name.value)
                                                    preview_code.text = f'{code}'
                                                    preview_code.style('font-size: 1.2rem; color: #2e7d32; font-weight: bold; font-family: monospace;')
                                                except:
                                                    preview_code.text = 'Code will be generated'
                                                    preview_code.style('font-size: 1.1rem; color: #666; font-style: italic; font-family: monospace;')
                                            else:
                                                preview_code.text = 'Enter program name to see code preview'
                                                preview_code.style('font-size: 1.1rem; color: #666; font-style: italic; font-family: monospace;')
                                        
                                        # Real-time update on input change
                                        program_name.on_value_change(lambda: update_code_preview())
                                        program_type.on_value_change(lambda: update_code_preview())
                                        
                                        # Initial preview update
                                        update_code_preview()
                                        
                                        def create_program():
                                            if not program_name.value:
                                                ui.notify('Program name is required', color='negative')
                                                return
                                            
                                            # Generate the program code
                                            try:
                                                year_doc = academic_years_col.find_one({'_id': ObjectId(selected_year_id)})
                                                year_name = year_doc.get('name', '') if year_doc else ''
                                                type_short = program_type.value.split(' ')[0] if program_type.value else 'UG'
                                                
                                                program_code = generate_program_code(inst_id, year_name, type_short, program_name.value)
                                            except Exception as e:
                                                ui.notify(f'Error generating program code: {str(e)}', color='negative')
                                                return
                                            
                                            program_doc = {
                                                'name': program_name.value,
                                                'code': program_code,
                                                'abbreviation': program_abbr.value,
                                                'type': type_short,
                                                'type_full': program_type.value,
                                                'duration': program_duration.value,
                                                'school_id': school_id,
                                                'institution_id': inst_id,
                                                'academic_cycle_id': selected_year_id,
                                                'created_at': datetime.now(timezone.utc),
                                                'created_by': current_user.get('email', 'admin') if current_user else 'admin'
                                            }
                                            
                                            result = programs_col.insert_one(program_doc)
                                            
                                            log_audit_action(
                                                action='Created Program',
                                                details=f'Program "{program_name.value}" ({program_duration.value}) created with code {program_code} in school "{school_name}"',
                                                institution_id=inst_id,
                                                entity_type='program',
                                                entity_id=str(result.inserted_id)
                                            )
                                            
                                            ui.notify(f'Program "{program_name.value}" created with code: {program_code}', color='positive')
                                            dialog.close()
                                            ui.run_javascript('window.location.reload()')
                                        
                                        with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                            ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                                            ui.button('Create Program', on_click=create_program).classes('btn-primary')
                                
                                dialog.open()
                            
                            ui.button('+ Add Program', on_click=add_program_to_school).classes('btn-success')
                            ui.button('Create User Group', on_click=lambda s_id=str(school['_id']): ui.navigate.to(f'/institution_admin/{inst_id}/create_user_group?school_id={urllib.parse.quote(s_id)}')).classes('btn-warning')
                            
                            # Edit school button
                            def edit_school(school_id=str(school['_id']), current_data=school):
                                with ui.dialog() as edit_dialog:
                                    with ui.card().style('padding: 2rem; min-width: 400px;'):
                                        ui.label('Edit School').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                        
                                        name_input = ui.input('School Name', value=current_data.get('name', '')).style('width: 100%; margin-bottom: 1rem;')
                                        abbr_input = ui.input('Abbreviation', value=current_data.get('abbreviation', '')).style('width: 100%; margin-bottom: 1rem;')
                                        desc_input = ui.textarea('Description', value=current_data.get('description', '')).style('width: 100%; margin-bottom: 1rem;')
                                        
                                        def save_changes():
                                            if not name_input.value.strip():
                                                ui.notify('School name is required', color='negative')
                                                return
                                            
                                            try:
                                                schools_col.update_one(
                                                    {'_id': ObjectId(school_id)},
                                                    {'$set': {
                                                        'name': name_input.value.strip(),
                                                        'abbreviation': abbr_input.value.strip(),
                                                        'description': desc_input.value.strip(),
                                                        'updated_at': datetime.now()
                                                    }}
                                                )
                                                
                                                log_audit_action(
                                                    action='Updated School',
                                                    details=f'School "{name_input.value.strip()}" was updated',
                                                    institution_id=inst_id,
                                                    entity_type='school',
                                                    entity_id=school_id
                                                )
                                                
                                                ui.notify('School updated successfully!', color='positive')
                                                edit_dialog.close()
                                                ui.run_javascript('window.location.reload()')
                                            
                                            except Exception as e:
                                                ui.notify(f'Error updating school: {str(e)}', color='negative')
                                        
                                        with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                            ui.button('Cancel', on_click=edit_dialog.close).classes('btn-secondary')
                                            ui.button('Save Changes', on_click=save_changes).style(f'background: {main_color}; color: white;')
                                
                                edit_dialog.open()
                            
                            ui.button('✏️', on_click=edit_school).style('background: #17a2b8; color: #fff;')
                            
                            # Delete school button
                            def delete_school(school_id=str(school['_id']), school_name=school_name):
                                # Check if school has programs
                                school_programs = list(programs_col.find({'school_id': school_id}))
                                if school_programs:
                                    ui.notify(f'Cannot delete school "{school_name}" - it has {len(school_programs)} program(s). Delete programs first.', color='warning')
                                    return
                                
                                with ui.dialog() as delete_dialog:
                                    with ui.card().style('padding: 2rem;'):
                                        ui.label('Confirm Deletion').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                        ui.label(f'Are you sure you want to delete school "{school_name}"?').style('margin-bottom: 1rem;')
                                        ui.label('This action cannot be undone.').style('color: red; font-weight: bold; margin-bottom: 1rem;')
                                        
                                        def confirm_delete():
                                            log_audit_action(
                                                action='Deleted School',
                                                details=f'School "{school_name}" was deleted',
                                                institution_id=inst_id,
                                                entity_type='school',
                                                entity_id=school_id
                                            )
                                            
                                            schools_col.delete_one({'_id': ObjectId(school_id)})
                                            ui.notify(f'School "{school_name}" deleted successfully!', color='positive')
                                            delete_dialog.close()
                                            ui.run_javascript('window.location.reload()')
                                        
                                        with ui.row().style('gap: 1rem;'):
                                            ui.button('Cancel', on_click=delete_dialog.close).classes('btn-secondary')
                                            ui.button('Delete', on_click=confirm_delete).classes('btn-danger')
                                
                                delete_dialog.open()
                            
                            ui.button('🗑️', on_click=delete_school).classes('btn-danger')
                    
                    # Display programs under this school
                    school_programs = [p for p in programs if p.get('school_id') == str(school['_id'])]
                    if school_programs:
                        ui.label(f'Programs ({len(school_programs)}):').style('font-weight: bold; margin-top: 1rem; margin-bottom: 0.5rem;')
                        for program in school_programs:
                            program_card_id = f'program-{program["_id"]}'
                            with ui.card().style('width: 100%; margin-bottom: 0.5rem; padding: 1rem; background: #f8f9fa; border-left: 3px solid var(--success-color); cursor: move;').classes(f'draggable-program {program_card_id}'):
                                with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                                    with ui.column():
                                        program_name = program.get('name', 'Unnamed Program')
                                        program_abbr = program.get('abbreviation', '')
                                        program_type = program.get('type', 'Unknown')
                                        program_duration = program.get('duration', 'Unknown')
                                        program_code = program.get('code', 'N/A')
                                        
                                        ui.label(f"🎓 {program_name} {f'({program_abbr})' if program_abbr else ''}").style(
                                            'font-size: 1rem; font-weight: bold; color: var(--success-color);'
                                        )
                                        ui.label(f'Code: {program_code}').style('font-size: 0.85rem; color: #666; font-family: monospace;')
                                        ui.label(f'{program_type} • {program_duration}').style('font-size: 0.8rem; color: var(--text-secondary);')
                                    
                                    with ui.row().style('gap: 0.5rem;'):
                                        ui.button('Create User Group', on_click=lambda p_id=str(program['_id']): ui.navigate.to(f'/institution_admin/{inst_id}/create_user_group?program_id={urllib.parse.quote(p_id)}')).style('background: #1976d2; color: #fff; font-size: 0.8rem; padding: 0.25rem 0.5rem;')
                                        
                                        # Edit program button
                                        def edit_program(program_id=str(program['_id']), current_data=program):
                                            with ui.dialog() as edit_dialog:
                                                with ui.card().style('padding: 2rem; min-width: 400px;'):
                                                    ui.label('Edit Program').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                                    
                                                    name_input = ui.input('Program Name', value=current_data.get('name', '')).style('width: 100%; margin-bottom: 1rem;')
                                                    abbr_input = ui.input('Abbreviation', value=current_data.get('abbreviation', '')).style('width: 100%; margin-bottom: 1rem;')
                                                    type_input = ui.input('Type', value=current_data.get('type', '')).style('width: 100%; margin-bottom: 1rem;')
                                                    duration_input = ui.input('Duration', value=current_data.get('duration', '')).style('width: 100%; margin-bottom: 1rem;')
                                                    
                                                    # Current code preview
                                                    with ui.card().style('background: #f8f9fa; border: 2px solid #e9ecef; padding: 1rem; margin-bottom: 1rem;'):
                                                        ui.label('Current Code:').style('font-size: 0.8rem; color: #666; font-weight: bold; margin-bottom: 0.5rem;')
                                                        current_code_label = ui.label(current_data.get('code', 'N/A')).style('font-size: 1.1rem; color: #2e7d32; font-weight: bold; font-family: monospace;')
                                                        ui.label('Note: Code will be regenerated if program name or type changes').style('font-size: 0.75rem; color: #666; margin-top: 0.5rem;')
                                                    
                                                    def save_changes():
                                                        if not name_input.value.strip():
                                                            ui.notify('Program name is required', color='negative')
                                                            return
                                                        
                                                        # Generate new code if name changed
                                                        new_code = current_data.get('code', '')
                                                        if name_input.value.strip() != current_data.get('name', ''):
                                                            try:
                                                                year_doc = academic_years_col.find_one({'_id': ObjectId(current_data['academic_year_id'])})
                                                                if year_doc:
                                                                    new_code = generate_program_code(inst_id, year_doc['name'], type_input.value.strip())
                                                            except:
                                                                pass
                                                        
                                                        try:
                                                            programs_col.update_one(
                                                                {'_id': ObjectId(program_id)},
                                                                {'$set': {
                                                                    'name': name_input.value.strip(),
                                                                    'abbreviation': abbr_input.value.strip(),
                                                                    'type': type_input.value.strip(),
                                                                    'duration': duration_input.value.strip(),
                                                                    'code': new_code,
                                                                    'updated_at': datetime.now()
                                                                }}
                                                            )
                                                            
                                                            log_audit_action(
                                                                action='Updated Program',
                                                                details=f'Program "{name_input.value.strip()}" was updated',
                                                                institution_id=inst_id,
                                                                entity_type='program',
                                                                entity_id=program_id
                                                            )
                                                            
                                                            ui.notify('Program updated successfully!', color='positive')
                                                            edit_dialog.close()
                                                            ui.run_javascript('window.location.reload()')
                                                        
                                                        except Exception as e:
                                                            ui.notify(f'Error updating program: {str(e)}', color='negative')
                                                    
                                                    with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                                        ui.button('Cancel', on_click=edit_dialog.close).classes('btn-secondary')
                                                        ui.button('Save Changes', on_click=save_changes).style(f'background: {main_color}; color: white;')
                                            
                                            edit_dialog.open()
                                        
                                        ui.button('✏️', on_click=edit_program).style('background: #17a2b8; color: #fff; font-size: 0.8rem; padding: 0.25rem 0.5rem;')
                                        
                                        # Delete program button
                                        def delete_program(program_id=str(program['_id']), program_name=program_name):
                                            with ui.dialog() as delete_dialog:
                                                with ui.card().style('padding: 2rem;'):
                                                    ui.label('Confirm Deletion').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                                    ui.label(f'Are you sure you want to delete program "{program_name}"?').style('margin-bottom: 1rem;')
                                                    ui.label('This action cannot be undone.').style('color: red; font-weight: bold; margin-bottom: 1rem;')
                                                    
                                                    def confirm_delete():
                                                        log_audit_action(
                                                            action='Deleted Program',
                                                            details=f'Program "{program_name}" with code "{program.get("code", "N/A")}" was deleted',
                                                            institution_id=inst_id,
                                                            entity_type='program',
                                                            entity_id=program_id
                                                        )
                                                        
                                                        programs_col.delete_one({'_id': ObjectId(program_id)})
                                                        ui.notify(f'Program "{program_name}" deleted successfully!', color='positive')
                                                        delete_dialog.close()
                                                        ui.run_javascript('window.location.reload()')
                                                    
                                                    with ui.row().style('gap: 1rem;'):
                                                        ui.button('Cancel', on_click=delete_dialog.close).classes('btn-secondary')
                                                        ui.button('Delete', on_click=confirm_delete).classes('btn-danger')
                                            
                                            delete_dialog.open()
                                        
                                        ui.button('🗑️', on_click=delete_program).style('background: #dc3545; color: #fff; font-size: 0.8rem; padding: 0.25rem 0.5rem;').style('font-size: 0.8rem; padding: 0.3rem 0.6rem;')
            
            # Display Departments
            if departments:
                ui.label('Departments').style(f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-top: 2rem; margin-bottom: 1rem;')
                
                for dept in departments:
                    with ui.card().classes('beautiful-card').style(f'width: 100%; margin-bottom: 1rem; padding: 1.5rem; border-left: 4px solid var(--info-color);'):
                        with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                            with ui.column():
                                dept_name = dept.get('name', 'Unnamed Department')
                                dept_abbr = dept.get('abbreviation', '')
                                ui.label(f"🏢 {dept_name} {f'({dept_abbr})' if dept_abbr else ''}").style(
                                    'font-size: 1.2rem; font-weight: bold; color: var(--info-color);'
                                )
                                ui.label('Department').style('font-size: 0.9rem; color: var(--text-secondary);')
                            
                            with ui.row().style('gap: 0.5rem;'):
                                ui.button('Create User Group', on_click=lambda d_id=str(dept['_id']): ui.navigate.to(f'/institution_admin/{inst_id}/create_user_group?department_id={urllib.parse.quote(d_id)}')).classes('btn-warning')
                                
                                # Edit department button
                                def edit_department(dept_id=str(dept['_id']), current_data=dept):
                                    with ui.dialog() as edit_dialog:
                                        with ui.card().style('padding: 2rem; min-width: 400px;'):
                                            ui.label('Edit Department').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                            
                                            name_input = ui.input('Department Name', value=current_data.get('name', '')).style('width: 100%; margin-bottom: 1rem;')
                                            abbr_input = ui.input('Abbreviation', value=current_data.get('abbreviation', '')).style('width: 100%; margin-bottom: 1rem;')
                                            desc_input = ui.textarea('Description', value=current_data.get('description', '')).style('width: 100%; margin-bottom: 1rem;')
                                            
                                            def save_changes():
                                                if not name_input.value.strip():
                                                    ui.notify('Department name is required', color='negative')
                                                    return
                                                
                                                try:
                                                    schools_col.update_one(
                                                        {'_id': ObjectId(dept_id)},
                                                        {'$set': {
                                                            'name': name_input.value.strip(),
                                                            'abbreviation': abbr_input.value.strip(),
                                                            'description': desc_input.value.strip(),
                                                            'updated_at': datetime.now()
                                                        }}
                                                    )
                                                    
                                                    log_audit_action(
                                                        action='Updated Department',
                                                        details=f'Department "{name_input.value.strip()}" was updated',
                                                        institution_id=inst_id,
                                                        entity_type='department',
                                                        entity_id=dept_id
                                                    )
                                                    
                                                    ui.notify('Department updated successfully!', color='positive')
                                                    edit_dialog.close()
                                                    ui.run_javascript('window.location.reload()')
                                                
                                                except Exception as e:
                                                    ui.notify(f'Error updating department: {str(e)}', color='negative')
                                            
                                            with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                                                ui.button('Cancel', on_click=edit_dialog.close).classes('btn-secondary')
                                                ui.button('Save Changes', on_click=save_changes).style(f'background: {main_color}; color: white;')
                                    
                                    edit_dialog.open()
                                
                                ui.button('✏️', on_click=edit_department).style('background: #17a2b8; color: #fff;')
                                
                                # Delete department button
                            def delete_department(dept_id=str(dept['_id']), dept_name=dept_name):
                                with ui.dialog() as delete_dialog:
                                    with ui.card().style('padding: 2rem;'):
                                        ui.label('Confirm Deletion').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                        ui.label(f'Are you sure you want to delete department "{dept_name}"?').style('margin-bottom: 1rem;')
                                        ui.label('This action cannot be undone.').style('color: red; font-weight: bold; margin-bottom: 1rem;')
                                        
                                        def confirm_delete():
                                            log_audit_action(
                                                action='Deleted Department',
                                                details=f'Department "{dept_name}" was deleted',
                                                institution_id=inst_id,
                                                entity_type='department',
                                                entity_id=dept_id
                                            )
                                            
                                            schools_col.delete_one({'_id': ObjectId(dept_id)})
                                            ui.notify(f'Department "{dept_name}" deleted successfully!', color='positive')
                                            delete_dialog.close()
                                            ui.run_javascript('window.location.reload()')
                                        
                                        with ui.row().style('gap: 1rem;'):
                                            ui.button('Cancel', on_click=delete_dialog.close).classes('btn-secondary')
                                            ui.button('Delete', on_click=confirm_delete).classes('btn-danger')
                                
                                delete_dialog.open()
                            
                            ui.button('🗑️', on_click=delete_department).classes('btn-danger')
            
            if not schools and not departments:
                ui.label('No schools or departments found for the selected academic year.').style(
                    'color: var(--text-secondary); font-style: italic; margin-top: 2rem; text-align: center;'
                )
        
        else:
            ui.label('Please select an academic year to view the institution hierarchy.').style(
                'color: var(--text-secondary); font-style: italic; margin-top: 2rem; text-align: center;'
            )
        
        # Add drag and drop JavaScript
        ui.add_head_html(f'''
        <script>
        // Initialize drag and drop for programs
        function initializeDragAndDrop() {{
            const programCards = document.querySelectorAll('.draggable-program');
            const dropZones = document.querySelectorAll('.drop-zone');
            
            // Add drag event listeners to program cards
            programCards.forEach(card => {{
                card.draggable = true;
                
                card.addEventListener('dragstart', function(e) {{
                    e.dataTransfer.setData('text/plain', card.className.split(' ').find(cls => cls.startsWith('program-')));
                    card.style.opacity = '0.5';
                    card.style.transform = 'scale(0.95)';
                }});
                
                card.addEventListener('dragend', function(e) {{
                    card.style.opacity = '1';
                    card.style.transform = 'scale(1)';
                }});
            }});
            
            // Add drop event listeners to school zones
            dropZones.forEach(zone => {{
                zone.addEventListener('dragover', function(e) {{
                    e.preventDefault();
                    zone.style.background = '#e3f2fd';
                    zone.style.borderLeft = '6px solid #2196f3';
                }});
                
                zone.addEventListener('dragleave', function(e) {{
                    zone.style.background = '#fff';
                    zone.style.borderLeft = '4px solid {main_color}';
                }});
                
                zone.addEventListener('drop', function(e) {{
                    e.preventDefault();
                    zone.style.background = '#fff';
                    zone.style.borderLeft = '4px solid {main_color}';
                    
                    const programClass = e.dataTransfer.getData('text/plain');
                    const programId = programClass.replace('program-', '');
                    const schoolClass = zone.className.split(' ').find(cls => cls.startsWith('school-'));
                    const schoolId = schoolClass.replace('school-', '');
                    
                    // Don't move if dropped on same school
                    const programCard = document.querySelector('.' + programClass);
                    const currentSchoolCard = programCard.closest('.drop-zone');
                    if (currentSchoolCard === zone) {{
                        return;
                    }}
                    
                    // Make API call to move program
                    fetch('/api/move_program', {{
                        method: 'POST',
                        headers: {{
                            'Content-Type': 'application/json',
                        }},
                        body: JSON.stringify({{
                            program_id: programId,
                            new_school_id: schoolId,
                            institution_id: '{inst_id}'
                        }})
                    }})
                    .then(response => response.json())
                    .then(data => {{
                        if (data.success) {{
                            alert('✓ ' + data.message);
                            setTimeout(() => {{
                                window.location.reload();
                            }}, 500);
                        }} else {{
                            alert('✗ ' + (data.message || 'Failed to move program'));
                        }}
                    }})
                    .catch(error => {{
                        console.error('Error:', error);
                        alert('✗ Error moving program');
                    }});
                }});
            }});
        }}
        
        // Initialize when DOM is ready
        if (document.readyState === 'loading') {{
            document.addEventListener('DOMContentLoaded', initializeDragAndDrop);
        }} else {{
            initializeDragAndDrop();
        }}
        </script>
        ''')
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/extended_profiles')
def institution_admin_extended_profiles(inst_id: str):
    """Extended profiles management page"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    def content(inst, main_color):
        ui.label('Extended Profiles').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Academic year selector
        from nicegui import app
        unlocked_years = list(academic_years_col.find({'institution_id': inst_id, 'is_locked': False}))
        year_options = [y['name'] for y in unlocked_years]
        name_to_id = {y['name']: str(y['_id']) for y in unlocked_years}
        
        # Initialize session if needed
        if not hasattr(app.storage, 'user'):
            app.storage.user = {}
        
        selected_year_id = app.storage.user.get('selected_academic_year_id') if hasattr(app.storage, 'user') else None
        
        def on_year_change(e):
            if hasattr(app.storage, 'user'):
                app.storage.user['selected_academic_year_id'] = name_to_id.get(e.value)
        
        with ui.row().style('width: 100%; align-items: center; justify-content: flex-end; margin-bottom: 1rem;'):
            if year_options:
                selected_year_name = None
                for name, id_val in name_to_id.items():
                    if id_val == selected_year_id:
                        selected_year_name = name
                        break
                
                ui.select(
                    options=year_options,
                    value=selected_year_name,
                    on_change=on_year_change,
                    label='Academic Year'
                ).style('min-width: 200px;').classes('beautiful-input')
        
        # Create Extended Profile button
        def open_create_dialog():
            if not selected_year_id:
                ui.notify('Select an academic year first', color='negative')
                return
            
            with ui.dialog() as dialog:
                with ui.card().style(
                    'padding: 2rem; min-width: 800px; max-width: 1000px; max-height: 90vh; overflow-y: auto;'
                ).classes('beautiful-card'):
                    ui.label('Create Extended Profile').style(
                        f'font-size: 1.8rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem; text-align: center;'
                    )
                    
                    with ui.column().style('width: 100%; gap: 1.5rem;'):
                        # Basic Information
                        ui.label('Basic Information').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem;'
                        )
                        
                        profile_name = ui.input('Profile Name').classes('beautiful-input').style('width: 100%;')
                        profile_description = ui.textarea('Description').classes('beautiful-input').style('width: 100%; min-height: 100px;')
                        
                        # Extended Profile Type
                        ui.label('Extended Profile Type').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem; margin-top: 1rem;'
                        )
                        
                        scope_type = ui.select(
                            options=['Program-based', 'Department-based'],
                            value='Program-based',
                            label='Choose whether this Extended profile applies to programs or departments'
                        ).classes('beautiful-input').style('width: 100%; margin-bottom: 1rem;')
                        
                        ui.label('Choose whether this Extended profile applies to programs or departments.').style(
                            'color: var(--text-secondary); font-size: 0.9rem; margin-top: -0.5rem; margin-bottom: 1rem;'
                        )
                        
                        # Headers/Fields
                        ui.label('Profile Fields').style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem; margin-top: 1rem;'
                        )
                        
                        headers_input = ui.textarea(
                            'Field Names (one per line)',
                            placeholder='Name\nPosition\nDepartment\nEmail\nPhone'
                        ).classes('beautiful-input').style('width: 100%; min-height: 120px;')
                        
                        # Table Preview
                        ui.label('Profile Preview').style(
                            f'font-size: 1.1rem; font-weight: bold; color: {main_color}; margin: 1.5rem 0 0.5rem 0;'
                        )
                        
                        preview_container = ui.column().style('width: 100%;')
                        
                        def update_preview():
                            if not headers_input.value:
                                preview_container.clear()
                                with preview_container:
                                    ui.label('Enter field names above to see table preview').style('color: #666; font-style: italic; text-align: center;')
                                return
                                
                            preview_container.clear()
                            headers = [h.strip() for h in headers_input.value.split('\n') if h.strip()]
                            
                            if headers:
                                with preview_container:
                                    ui.label('Table Preview:').style('font-weight: bold; margin-bottom: 0.5rem;')
                                    with ui.card().style('width: 100%; background: white; border: 2px solid rgba(154, 44, 84, 0.2); padding: 1rem; border-radius: 8px;'):
                                        # Create table with proper styling
                                        with ui.element('table').style('width: 100%; border-collapse: collapse;'):
                                            # Header row
                                            with ui.element('thead'):
                                                with ui.element('tr'):
                                                    for header in headers:
                                                        with ui.element('th').style('background: rgb(154, 44, 84); color: white; padding: 12px; text-align: left; border: 1px solid #ddd; font-weight: bold;'):
                                                            ui.label(header)
                                            
                                            # Sample data rows
                                            with ui.element('tbody'):
                                                sample_data = [
                                                    ['John Doe', 'Professor', 'Computer Science', 'john.doe@university.edu', '+1-555-0123'],
                                                    ['Jane Smith', 'Associate Professor', 'Mathematics', 'jane.smith@university.edu', '+1-555-0124'],
                                                    ['Dr. Johnson', 'Assistant Professor', 'Physics', 'johnson@university.edu', '+1-555-0125']
                                                ]
                                                
                                                for i, row_data in enumerate(sample_data):
                                                    with ui.element('tr'):
                                                        for j, header in enumerate(headers):
                                                            sample_value = row_data[j] if j < len(row_data) else f'Sample {i+1}-{j+1}'
                                                            with ui.element('td').style('padding: 10px; border: 1px solid #ddd; background: #f9f9f9;'):
                                                                ui.label(sample_value).style('font-size: 0.9rem;')
                                    
                                    ui.label(f'Table will have {len(headers)} columns as shown above').style(
                                        'font-size: 0.8rem; color: #666; margin-top: 0.5rem; text-align: center;'
                                    )
                        
                        # Update preview when headers change
                        headers_input.on_value_change(lambda: update_preview())
                        scope_type.on_value_change(lambda: update_preview())
                        profile_name.on_value_change(lambda: update_preview())
                        
                        # Initialize preview
                        update_preview()
                        
                        # Options
                        with ui.row().style('width: 100%; gap: 2rem; margin-top: 1rem;'):
                            needs_docs = ui.checkbox('Requires Supporting Documents', value=False)
                        
                        # Create function
                        def create_profile():
                            global current_user
                            if not profile_name.value:
                                ui.notify('Profile name is required', color='negative')
                                return
                            
                            if not headers_input.value:
                                ui.notify('At least one field is required', color='negative')
                                return
                            
                            # Parse headers
                            headers = [h.strip() for h in headers_input.value.split('\n') if h.strip()]
                            
                            profile_doc = {
                                'name': profile_name.value,
                                'description': profile_description.value,
                                'scope_type': 'program_based' if scope_type.value == 'Program-based' else 'department_based',
                                'institution_id': inst_id,
                                'academic_cycle_id': selected_year_id,
                                'headers': headers,
                                'needs_supporting_docs': needs_docs.value,
                                'created_at': datetime.now(timezone.utc),
                                'updated_at': datetime.now(timezone.utc),
                                'created_by': current_user.get('email', 'admin') if current_user else 'admin'
                            }
                            
                            try:
                                result = extended_profiles_col.insert_one(profile_doc)
                                
                                log_audit_action(
                                    action='Created Extended Profile',
                                    details=f'Extended Profile "{profile_name.value}" created with {len(headers)} fields',
                                    institution_id=inst_id,
                                    entity_type='extended_profile',
                                    entity_id=str(result.inserted_id)
                                )
                                
                                ui.notify(f'Extended Profile "{profile_name.value}" created successfully!', color='positive')
                                dialog.close()
                                ui.run_javascript('window.location.reload()')
                            
                            except Exception as e:
                                ui.notify(f'Error creating profile: {str(e)}', color='negative')
                        
                        # Action buttons
                        with ui.row().style('width: 100%; justify-content: space-between; margin-top: 2rem;'):
                            ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                            ui.button('Create Profile', on_click=create_profile).classes('btn-primary')
                
                dialog.open()
        
        ui.button('+ Create Extended Profile', on_click=open_create_dialog).classes('btn-primary').style('margin-bottom: 2rem;')
        
        # Display existing profiles
        if selected_year_id:
            existing_profiles = list(extended_profiles_col.find({
                'institution_id': inst_id,
                'academic_cycle_id': selected_year_id
            }))
            
            if existing_profiles:
                ui.label(f'Existing Extended Profiles ({len(existing_profiles)} found)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                
                for profile in existing_profiles:
                    with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1rem; padding: 1.5rem;'):
                        with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                            with ui.column().style('flex: 1;'):
                                profile_name = profile.get('name', 'Unnamed Profile')
                                
                                ui.label(f"📋 {profile_name}").style(
                                    'font-size: 1.2rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                                )
                                
                                if profile.get('description'):
                                    ui.label(f"Description: {profile['description']}").style(
                                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.3rem;'
                                    )
                                
                                headers = profile.get('headers', [])
                                if headers:
                                    ui.label(f"Fields: {', '.join(headers[:3])}{'...' if len(headers) > 3 else ''} ({len(headers)} total)").style(
                                        'font-size: 0.8rem; color: var(--text-secondary);'
                                    )
                                
                                needs_docs = 'Yes' if profile.get('needs_supporting_docs', False) else 'No'
                                ui.label(f"Supporting Documents: {needs_docs}").style(
                                    'font-size: 0.8rem; color: var(--text-secondary);'
                                )
                                
                                created_at = profile.get('created_at')
                                if created_at:
                                    ui.label(f"Created: {created_at.strftime('%Y-%m-%d %H:%M')}").style(
                                        'font-size: 0.8rem; color: var(--text-secondary);'
                                    )
                            
                            with ui.row().style('gap: 0.5rem;'):
                                ui.button(
                                    '📈 View Data',
                                    on_click=lambda p_id=str(profile['_id']): ui.navigate.to(
                                        f'/institution_admin/{inst_id}/spreadsheets?profile_id={p_id}'
                                    )
                                ).classes('btn-success')
                                
                                # Edit button
                                def edit_profile(profile_id=str(profile['_id']), prof=profile):
                                    with ui.dialog() as edit_dialog:
                                        with ui.card().style('padding: 2rem; min-width: 600px;').classes('beautiful-card'):
                                            ui.label('Edit Extended Profile').style(
                                                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                                            )
                                            
                                            edit_name = ui.input('Profile Name', value=prof.get('name', '')).style('width: 100%; margin-bottom: 1rem;')
                                            edit_description = ui.textarea('Description', value=prof.get('description', '')).style('width: 100%; margin-bottom: 1rem;')
                                            
                                            edit_headers = ui.textarea(
                                                'Field Names (one per line)',
                                                value='\n'.join(prof.get('headers', []))
                                            ).style('width: 100%; margin-bottom: 1rem;')
                                            
                                            edit_needs_docs = ui.checkbox('Requires Supporting Documents', value=prof.get('needs_supporting_docs', False))
                                            
                                            def save_changes():
                                                if not edit_name.value:
                                                    ui.notify('Profile name is required', color='negative')
                                                    return
                                                
                                                headers = [h.strip() for h in edit_headers.value.split('\n') if h.strip()]
                                                
                                                extended_profiles_col.update_one(
                                                    {'_id': ObjectId(profile_id)},
                                                    {'$set': {
                                                        'name': edit_name.value,
                                                        'description': edit_description.value,
                                                        'headers': headers,
                                                        'needs_supporting_docs': edit_needs_docs.value,
                                                        'updated_at': datetime.datetime.now(datetime.timezone.utc)
                                                    }}
                                                )
                                                
                                                log_audit_action(
                                                    action='Updated Extended Profile',
                                                    details=f'Extended Profile "{edit_name.value}" was updated',
                                                    institution_id=inst_id,
                                                    entity_type='extended_profile',
                                                    entity_id=profile_id
                                                )
                                                
                                                ui.notify('Extended Profile updated successfully!', color='positive')
                                                edit_dialog.close()
                                                ui.run_javascript('window.location.reload()')
                                            
                                            with ui.row().style('margin-top: 1rem; gap: 1rem;'):
                                                ui.button('Cancel', on_click=edit_dialog.close).classes('btn-secondary')
                                                ui.button('Save Changes', on_click=save_changes).classes('btn-primary')
                                    
                                    edit_dialog.open()
                                
                                ui.button('✏️ Edit', on_click=edit_profile).classes('btn-warning')
                                
                                def delete_profile(profile_id=str(profile['_id']), profile_name=profile.get('name', '')):
                                    with ui.dialog() as delete_dialog:
                                        with ui.card().style('padding: 2rem;'):
                                            ui.label('Confirm Deletion').style('font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem;')
                                            ui.label(f'Are you sure you want to delete extended profile "{profile_name}"?').style('color: red; font-weight: bold; margin-bottom: 1rem;')
                                            
                                            def confirm_delete():
                                                log_audit_action(
                                                    action='Deleted Extended Profile',
                                                    details=f'Extended Profile "{profile_name}" deleted',
                                                    institution_id=inst_id,
                                                    entity_type='extended_profile',
                                                    entity_id=profile_id
                                                )
                                                extended_profiles_col.delete_one({'_id': ObjectId(profile_id)})
                                                ui.notify(f'Extended Profile "{profile_name}" deleted', color='positive')
                                                delete_dialog.close()
                                                ui.run_javascript('window.location.reload()')
                                            
                                            with ui.row().style('gap: 1rem;'):
                                                ui.button('Cancel', on_click=delete_dialog.close).classes('btn-secondary')
                                                ui.button('Delete', on_click=confirm_delete).classes('btn-danger')
                                    
                                    delete_dialog.open()
                                
                                ui.button('🗑️ Delete', on_click=delete_profile).classes('btn-danger')
            else:
                ui.label('No extended profiles found for the selected academic year.').style(
                    'color: var(--text-secondary); font-style: italic; margin-top: 2rem;'
                )
        else:
            ui.label('Please select an academic year to view extended profiles.').style(
                'color: var(--text-secondary); font-style: italic; margin-top: 2rem;'
            )
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/spreadsheets')
def institution_admin_spreadsheets(inst_id: str):
    """Spreadsheets management page with Excel-like interface"""
    add_beautiful_global_styles()
    
    def content(inst, main_color):
        ui.label('Spreadsheets').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Academic Year Selector
        ui.label('📅 Select Academic Year').style(
            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
        )
        
        # Get available academic years
        academic_years = list(academic_years_col.find({'institution_id': inst_id}).sort('created_at', -1))
        if not academic_years:
            ui.label('No academic years found. Please create an academic year first.').style(
                'color: var(--warning-color); margin-bottom: 1rem;'
            )
            return
        
        selected_year_selector = ui.select(
            label='Academic Year',
            options=[year['name'] for year in academic_years],
            value=academic_years[0]['name'] if academic_years else None
        ).style('width: 300px; margin-bottom: 2rem;')
        
        def on_year_change():
            ui.run_javascript('window.location.reload()')
        
        selected_year_selector.on('change', on_year_change)
        
        # Get selected academic year ID
        selected_year = next((year for year in academic_years if year['name'] == selected_year_selector.value), None)
        selected_year_id = str(selected_year['_id']) if selected_year else None
        
        # Store in app storage for other functions
        if hasattr(app.storage, 'user'):
            app.storage.user['selected_academic_year_id'] = selected_year_id
        
        if not selected_year_id:
            ui.label('Please select an academic year to view spreadsheets.').style(
                'color: var(--warning-color); font-size: 1.2rem; text-align: center; padding: 2rem;'
            )
            return
        
        # Debug: Show current academic year selection
        year_doc = academic_years_col.find_one({'_id': ObjectId(selected_year_id)})
        if year_doc:
            ui.label(f'📅 Current Academic Year: {year_doc.get("name", "Unknown")}').style(
                'color: var(--text-secondary); font-size: 0.9rem; margin-bottom: 1rem;'
            )
        
        # Get criterias and extended profiles
        criterias = list(criterias_col.find({'institution_id': inst_id, 'academic_year_id': selected_year_id}))
        profiles = list(extended_profiles_col.find({'institution_id': inst_id, 'academic_year_id': selected_year_id}))
        
        # Debug: Also check for profiles without academic_year_id (legacy data)
        legacy_profiles = list(extended_profiles_col.find({
            'institution_id': inst_id, 
            'academic_year_id': {'$exists': False}
        }))
        legacy_criterias = list(criterias_col.find({
            'institution_id': inst_id, 
            'academic_year_id': {'$exists': False}
        }))
        
        # Debug information
        ui.label(f'🔍 Debug: Found {len(criterias)} criterias, {len(profiles)} profiles for selected year').style(
            'color: var(--text-secondary); font-size: 0.8rem; font-style: italic;'
        )
        ui.label(f'🔍 Debug: Found {len(legacy_criterias)} legacy criterias, {len(legacy_profiles)} legacy profiles without year').style(
            'color: var(--text-secondary); font-size: 0.8rem; font-style: italic;'
        )
        
        # Include legacy data if current year has no data
        if not criterias and legacy_criterias:
            criterias = legacy_criterias
            ui.label('📌 Showing legacy criterias (no academic year assigned)').style(
                'color: var(--warning-color); font-size: 0.9rem; margin-bottom: 0.5rem;'
            )
        
        if not profiles and legacy_profiles:
            profiles = legacy_profiles
            ui.label('📌 Showing legacy extended profiles (no academic year assigned)').style(
                'color: var(--warning-color); font-size: 0.9rem; margin-bottom: 0.5rem;'
            )
        legacy_profiles = list(extended_profiles_col.find({
            'institution_id': inst_id, 
            '$or': [
                {'academic_year_id': {'$exists': False}},
                {'academic_year_id': None}
            ]
        }))
        
        legacy_criterias = list(criterias_col.find({
            'institution_id': inst_id, 
            '$or': [
                {'academic_year_id': {'$exists': False}},
                {'academic_year_id': None}
            ]
        }))
        
        # Debug info
        total_profiles = len(profiles) + len(legacy_profiles)
        total_criterias = len(criterias) + len(legacy_criterias)
        
        if total_profiles > 0 or total_criterias > 0:
            debug_text = f'Found: {len(profiles)} profiles, {len(criterias)} criterias for current year'
            if legacy_profiles or legacy_criterias:
                debug_text += f' + {len(legacy_profiles)} legacy profiles, {len(legacy_criterias)} legacy criterias'
            ui.label(debug_text).style('color: var(--text-secondary); font-size: 0.8rem; margin-bottom: 1rem;')
        
        # Include legacy data if no current year data exists
        if not profiles and legacy_profiles:
            profiles = legacy_profiles
        if not criterias and legacy_criterias:
            criterias = legacy_criterias
        
        all_sheets = []
        
        # Add criterias as sheets (always show with default headers even if empty)
        for criteria in criterias:
            headers = criteria.get('headers', [])
            # If no headers defined, show default headers based on criteria structure
            if not headers:
                headers = ['Criteria Item', 'Value', 'Status', 'Comments']
            
            all_sheets.append({
                'type': 'criteria',
                'name': criteria.get('name', 'Unnamed Criteria'),
                'headers': headers,
                'data': criteria,
                'icon': '📊',
                'has_custom_headers': bool(criteria.get('headers'))
            })
        
        # Add extended profiles as sheets (always show with default headers even if empty)
        for profile in profiles:
            headers = profile.get('headers', [])
            # If no headers defined, show default headers based on profile structure  
            if not headers:
                headers = ['Profile Field', 'Description', 'Type', 'Required']
            
            all_sheets.append({
                'type': 'profile',
                'name': profile.get('name', 'Unnamed Profile'),
                'headers': headers,
                'data': profile,
                'icon': '📝',
                'has_custom_headers': bool(profile.get('headers'))
            })
        
        if not all_sheets:
            with ui.card().classes('beautiful-card').style('width: 100%; padding: 3rem; text-align: center;'):
                ui.label('No spreadsheets available').style(f'font-size: 1.5rem; color: {main_color}; margin-bottom: 1rem;')
                ui.label('Create criterias and extended profiles to see them as spreadsheet tabs here.').style('color: var(--text-secondary);')
                
                with ui.row().style('gap: 1rem; margin-top: 2rem; justify-content: center;'):
                    ui.button('Create Criteria', on_click=lambda: ui.navigate.to(f'/institution_admin/{inst_id}/criterias')).classes('btn-primary')
                    ui.button('Create Extended Profile', on_click=lambda: ui.navigate.to(f'/institution_admin/{inst_id}/extended_profiles')).classes('btn-secondary')
            return
        
        # Excel-like interface with JavaScript-based tab switching
        with ui.card().classes('beautiful-card').style('width: 100%; padding: 0; overflow: hidden;'):
            # Sheet tabs (like Excel)
            with ui.row().style('background: #f8f9fa; border-bottom: 1px solid var(--border); padding: 0; overflow-x: auto;'):
                for i, sheet in enumerate(all_sheets):
                    tab_style = f'''
                        background: {"white" if i == 0 else "#f8f9fa"}; 
                        border: 1px solid var(--border); 
                        border-bottom: {"none" if i == 0 else "1px solid var(--border)"}; 
                        padding: 0.75rem 1.5rem; 
                        margin-right: 2px; 
                        cursor: pointer;
                        border-radius: 8px 8px 0 0;
                        transition: all 0.3s ease;
                    '''
                    
                    with ui.element('div').style(tab_style).classes(f'sheet-tab-{i}').props(f'onclick="switchSheet({i})"'):
                        ui.label(f"{sheet['icon']} {sheet['name']}").style(
                            f'font-weight: {"bold" if i == 0 else "normal"}; color: {main_color if i == 0 else "var(--text-secondary)"};'
                        )
            
            # Add JavaScript for tab switching
            ui.add_head_html('''
                <script>
                function switchSheet(index) {
                    // Hide all sheet contents
                    document.querySelectorAll('[class*="sheet-content-"]').forEach(content => {
                        content.style.display = 'none';
                    });
                    
                    // Show selected sheet content
                    const selectedContent = document.querySelector('.sheet-content-' + index);
                    if (selectedContent) {
                        selectedContent.style.display = 'block';
                    }
                    
                    // Update tab styles
                    document.querySelectorAll('[class*="sheet-tab-"]').forEach((tab, i) => {
                        const label = tab.querySelector('div');
                        if (i === index) {
                            tab.style.background = 'white';
                            tab.style.borderBottom = 'none';
                            if (label) label.style.fontWeight = 'bold';
                        } else {
                            tab.style.background = '#f8f9fa';
                            tab.style.borderBottom = '1px solid var(--border)';
                            if (label) label.style.fontWeight = 'normal';
                        }
                    });
                }
                </script>
            ''')
            
            # Sheet contents - Show only one at a time
            with ui.column().style('padding: 2rem; background: white; min-height: 500px;'):
                for i, sheet in enumerate(all_sheets):
                    with ui.column().classes(f'sheet-content-{i}').style(f'display: {"block" if i == 0 else "none"}; width: 100%;'):
                        # Sheet header
                        with ui.row().style('width: 100%; align-items: center; justify-content: space-between; margin-bottom: 1.5rem;'):
                            ui.label(f"{sheet['icon']} {sheet['name']}").style(f'font-size: 1.5rem; font-weight: bold; color: {main_color};')
                            
                            sheet_type = sheet['data'].get('scope_type', 'program_based') if sheet['type'] == 'criteria' else 'N/A'
                            if sheet_type != 'N/A':
                                scope_label = 'Program-based' if sheet_type == 'program_based' else 'Department-based'
                                ui.label(f"Scope: {scope_label}").style('color: var(--text-secondary); font-size: 0.9rem;')
                        
                        # Excel-like table
                        headers = sheet['headers']
                        # Always show table with headers (default or custom)
                        with ui.element('div').style('overflow-x: auto; border: 1px solid var(--border); border-radius: 8px;'):
                            with ui.element('table').style('width: 100%; border-collapse: collapse; background: white;'):
                                # Header row
                                with ui.element('thead'):
                                    with ui.element('tr'):
                                        # Row number column
                                        with ui.element('th').style('background: #f8f9fa; padding: 12px; text-align: center; border: 1px solid var(--border); font-weight: bold; width: 50px;'):
                                            ui.label('#')
                                        
                                        # Data columns
                                        for header in headers:
                                            header_style = f'background: {main_color}; color: white; padding: 12px; text-align: left; border: 1px solid var(--border); font-weight: bold; min-width: 150px;'
                                            # Add indicator for default headers
                                            if not sheet.get('has_custom_headers', True):
                                                header_style += ' opacity: 0.8;'
                                            
                                            with ui.element('th').style(header_style):
                                                header_text = header
                                                if not sheet.get('has_custom_headers', True):
                                                    header_text += ' (default)'
                                                ui.label(header_text)
                                
                                # Sample data rows (showing empty structure)
                                with ui.element('tbody'):
                                    for row_num in range(1, 11):  # Show 10 sample rows
                                        with ui.element('tr'):
                                            # Row number
                                            with ui.element('td').style('background: #f8f9fa; padding: 10px; text-align: center; border: 1px solid var(--border); font-weight: bold;'):
                                                ui.label(str(row_num))
                                            
                                            # Empty data cells (editable in future)
                                            for col_idx, header in enumerate(headers):
                                                with ui.element('td').style('padding: 5px; border: 1px solid var(--border); min-height: 40px; background: white;'):
                                                    # Use input for editable cells
                                                    cell_input = ui.input('').style('width: 100%; border: none; background: transparent; padding: 5px;').props('dense outlined')
                                                    cell_input.props('placeholder=Click to edit')
                        
                        # Status indicator for default headers
                        if not sheet.get('has_custom_headers', True):
                            with ui.card().style('background: #fff3cd; border: 1px solid #ffeaa7; padding: 1rem; margin: 1rem 0; border-radius: 8px;'):
                                ui.label('⚠️ This sheet is using default headers. Create custom headers to define your data structure.').style('color: #856404; font-size: 0.9rem;')
                        
                        # Action buttons
                        with ui.row().style('margin-top: 1.5rem; gap: 1rem;'):
                            ui.button('📥 Import Data', on_click=lambda: ui.notify('Import functionality coming soon', color='info')).classes('btn-primary')
                            ui.button('📤 Export Excel', on_click=lambda s=sheet: export_to_excel(s, inst_id)).classes('btn-secondary')
                            
                            if sheet.get('has_custom_headers', True):
                                ui.button('✏️ Edit Headers', on_click=lambda s=sheet: edit_sheet_headers(s)).style('background: var(--warning-color); color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                            else:
                                ui.button('➕ Create Headers', on_click=lambda s=sheet: edit_sheet_headers(s)).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
        
        def export_to_excel(sheet, institution_id):
            """Export sheet data to Excel file"""
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                import io
                import base64
                from datetime import datetime
                
                # Create workbook and worksheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet['name'][:31]  # Excel sheet name limit
                
                # Get institution info for filename
                inst = institutions_col.find_one({'_id': ObjectId(institution_id)})
                inst_name = inst.get('name', 'Institution') if inst else 'Institution'
                
                # Add title and metadata
                ws['A1'] = f"{inst_name} - {sheet['name']}"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A2'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A2'].font = Font(size=10)
                
                # Add headers starting from row 4
                headers = sheet['headers']
                header_row = 4
                
                # Add row number column
                ws.cell(row=header_row, column=1, value='#')
                ws.cell(row=header_row, column=1).font = Font(bold=True)
                ws.cell(row=header_row, column=1).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # Add data headers
                for col_idx, header in enumerate(headers, start=2):
                    ws.cell(row=header_row, column=col_idx, value=header)
                    ws.cell(row=header_row, column=col_idx).font = Font(bold=True, color='FFFFFF')
                    ws.cell(row=header_row, column=col_idx).fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
                
                # Add sample data rows (10 empty rows for data entry)
                for row_num in range(1, 11):
                    data_row = header_row + row_num
                    
                    # Row number
                    ws.cell(row=data_row, column=1, value=row_num)
                    ws.cell(row=data_row, column=1).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    
                    # Empty data cells
                    for col_idx in range(2, len(headers) + 2):
                        ws.cell(row=data_row, column=col_idx, value='')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Create download
                filename = f"{inst_name}_{sheet['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filename = filename.replace(' ', '_').replace('/', '_')  # Clean filename
                
                # Encode for download
                excel_data = excel_buffer.getvalue()
                b64_data = base64.b64encode(excel_data).decode()
                
                # Trigger download using JavaScript
                download_js = f'''
                const link = document.createElement('a');
                link.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}';
                link.download = '{filename}';
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
                '''
                
                ui.run_javascript(download_js)
                ui.notify(f'Excel file "{filename}" downloaded successfully!', color='positive')
                
            except Exception as e:
                ui.notify(f'Error exporting to Excel: {str(e)}', color='negative')
                print(f"Excel export error: {str(e)}")  # Debug log
        
        def edit_sheet_headers(sheet):
            with ui.dialog() as dialog:
                with ui.card().style('padding: 2rem; min-width: 500px;'):
                    ui.label(f'Edit Headers - {sheet["name"]}').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
                    
                    current_headers = '\n'.join(sheet['headers'])
                    headers_input = ui.textarea('Headers (one per line)', value=current_headers).style('width: 100%; min-height: 150px; margin-bottom: 1rem;')
                    
                    def save_headers():
                        new_headers = [h.strip() for h in headers_input.value.split('\n') if h.strip()]
                        
                        if not new_headers:
                            ui.notify('At least one header is required', color='negative')
                            return
                        
                        try:
                            collection = criterias_col if sheet['type'] == 'criteria' else extended_profiles_col
                            collection.update_one(
                                {'_id': sheet['data']['_id']},
                                {'$set': {'headers': new_headers, 'updated_at': datetime.now()}}
                            )
                            
                            ui.notify('Headers updated successfully!', color='positive')
                            dialog.close()
                            ui.run_javascript('window.location.reload()')
                        except Exception as e:
                            ui.notify(f'Error updating headers: {str(e)}', color='negative')
                    
                    with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                        ui.button('Cancel', on_click=dialog.close).style('background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                        ui.button('Save Headers', on_click=save_headers).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
            
            dialog.open()
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/users')
def institution_admin_users(inst_id: str):
    """Manage Users page - Show all users except Platform Owner"""
    add_beautiful_global_styles()
    
    def content(inst, main_color):
        ui.label('Manage Users').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Get all users except Platform Owner (show users from all institutions)
        users = list(users_col.find({
            'role': {'$ne': 'Platform Owner'}  # Exclude Platform Owner
        }))
        
        # Action buttons
        with ui.row().style('width: 100%; margin-bottom: 1.5rem; gap: 1rem;'):
            ui.button('➕ Add New User', on_click=lambda: open_add_user_dialog()).classes('btn-primary')
            ui.button('📤 Export Users', on_click=lambda: ui.notify('Export functionality coming soon', color='info')).classes('btn-secondary')
        
        with ui.card().classes('beautiful-card').style('width: 100%; padding: 0; overflow: hidden;'):
            ui.label(f'Institution Users ({len(users)} found)').style(
                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin: 1.5rem; margin-bottom: 1rem;'
            )
            
            if users:
                # Users table
                with ui.element('div').style('overflow-x: auto;'):
                    with ui.element('table').style('width: 100%; border-collapse: collapse;'):
                        # Header
                        with ui.element('thead'):
                            with ui.element('tr'):
                                headers = ['User', 'Email', 'Role', 'Department', 'Last Login', 'Status', 'Actions']
                                for header in headers:
                                    with ui.element('th').style(f'background: {main_color}; color: white; padding: 12px; text-align: left; border: 1px solid var(--border); font-weight: bold;'):
                                        ui.label(header)
                        
                        # Rows
                        with ui.element('tbody'):
                            for i, user in enumerate(users):
                                row_bg = '#f8f9fa' if i % 2 == 0 else 'white'
                                
                                with ui.element('tr').style(f'background: {row_bg};'):
                                    # User info
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
                                        with ui.column().style('gap: 0.25rem;'):
                                            ui.label(f"👤 {full_name or 'Unnamed User'}").style('font-weight: bold;')
                                            if user.get('phone'):
                                                ui.label(f"📞 {user['phone']}").style('color: var(--text-secondary); font-size: 0.8rem;')
                                    
                                    # Email
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        ui.label(user.get('email', 'No email')).style('font-family: monospace; font-size: 0.9rem;')
                                    
                                    # Role
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        role = user.get('role', 'User')
                                        role_colors = {
                                            'Institution Admin': '#dc3545',
                                            'IQAC Coordinator': '#28a745',
                                            'Department Coordinator': '#ffc107',
                                            'Faculty': '#007bff',
                                            'User': '#6c757d'
                                        }
                                        role_color = role_colors.get(role, '#6c757d')
                                        ui.label(role).style(f'background: {role_color}; color: white; padding: 4px 8px; border-radius: 4px; font-size: 0.8rem; font-weight: bold;')
                                    
                                    # Department
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        ui.label(user.get('department', 'N/A')).style('color: var(--text-secondary);')
                                    
                                    # Last Login
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border); white-space: nowrap;'):
                                        last_login = user.get('last_login')
                                        if last_login:
                                            if isinstance(last_login, str):
                                                ui.label(last_login)
                                            else:
                                                ui.label(last_login.strftime('%Y-%m-%d %H:%M'))
                                        else:
                                            ui.label('Never').style('color: var(--text-secondary); font-style: italic;')
                                    
                                    # Status
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        status = user.get('status', 'active')
                                        status_color = '#28a745' if status == 'active' else '#dc3545'
                                        ui.label(status.title()).style(f'color: {status_color}; font-weight: bold;')
                                    
                                    # Actions
                                    with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                        with ui.row().style('gap: 0.5rem;'):
                                            ui.button('✏️', on_click=lambda u=user: edit_user(u)).style('background: var(--warning-color); color: white; padding: 0.5rem; border-radius: 4px; border: none; font-size: 0.8rem;')
                                            ui.button('🗑️', on_click=lambda u=user: delete_user(u)).style('background: var(--error-color); color: white; padding: 0.5rem; border-radius: 4px; border: none; font-size: 0.8rem;')
            else:
                with ui.element('div').style('padding: 3rem; text-align: center;'):
                    ui.label('No users found').style(f'font-size: 1.5rem; color: {main_color}; margin-bottom: 1rem;')
                    ui.label('Add users to manage institution access and roles.').style('color: var(--text-secondary);')
        
        def open_add_user_dialog():
            with ui.dialog() as dialog:
                with ui.card().style('padding: 2rem; min-width: 500px;'):
                    ui.label('Add New User').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
                    
                    first_name = ui.input('First Name').style('width: 100%; margin-bottom: 1rem;')
                    last_name = ui.input('Last Name').style('width: 100%; margin-bottom: 1rem;')
                    email = ui.input('Email', placeholder='user@example.com').style('width: 100%; margin-bottom: 1rem;')
                    phone = ui.input('Phone (Optional)').style('width: 100%; margin-bottom: 1rem;')
                    
                    role = ui.select(
                        ['Institution Admin', 'IQAC Coordinator', 'Department Coordinator', 'Faculty', 'User'],
                        value='User',
                        label='Role'
                    ).style('width: 100%; margin-bottom: 1rem;')
                    
                    department = ui.input('Department (Optional)').style('width: 100%; margin-bottom: 1rem;')
                    password = ui.input('Password', password=True).style('width: 100%; margin-bottom: 1rem;')
                    
                    def add_user():
                        if not all([first_name.value.strip(), last_name.value.strip(), email.value.strip(), password.value.strip()]):
                            ui.notify('Please fill in all required fields', color='negative')
                            return
                        
                        # Check if email already exists
                        if users_col.find_one({'email': email.value.strip()}):
                            ui.notify('Email already exists', color='negative')
                            return
                        
                        try:
                            new_user = {
                                'first_name': first_name.value.strip(),
                                'last_name': last_name.value.strip(),
                                'email': email.value.strip(),
                                'phone': phone.value.strip() if phone.value.strip() else None,
                                'role': role.value,
                                'department': department.value.strip() if department.value.strip() else None,
                                'password': password.value.strip(),  # In real app, hash this
                                'institution_id': inst_id,
                                'status': 'active',
                                'created_at': datetime.now(),
                                'last_login': None
                            }
                            
                            users_col.insert_one(new_user)
                            
                            ui.notify('User added successfully!', color='positive')
                            dialog.close()
                            ui.run_javascript('window.location.reload()')
                        except Exception as e:
                            ui.notify(f'Error adding user: {str(e)}', color='negative')
                    
                    with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                        ui.button('Cancel', on_click=dialog.close).style('background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                        ui.button('Add User', on_click=add_user).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
            
            dialog.open()
        
        def edit_user(user):
            with ui.dialog() as dialog:
                with ui.card().style('padding: 2rem; min-width: 500px;'):
                    ui.label(f'Edit User - {user.get("email", "Unknown")}').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;')
                    
                    first_name = ui.input('First Name', value=user.get('first_name', '')).style('width: 100%; margin-bottom: 1rem;')
                    last_name = ui.input('Last Name', value=user.get('last_name', '')).style('width: 100%; margin-bottom: 1rem;')
                    email = ui.input('Email', value=user.get('email', '')).style('width: 100%; margin-bottom: 1rem;')
                    phone = ui.input('Phone', value=user.get('phone', '') or '').style('width: 100%; margin-bottom: 1rem;')
                    
                    role = ui.select(
                        ['Institution Admin', 'IQAC Coordinator', 'Department Coordinator', 'Faculty', 'User'],
                        value=user.get('role', 'User'),
                        label='Role'
                    ).style('width: 100%; margin-bottom: 1rem;')
                    
                    department = ui.input('Department', value=user.get('department', '') or '').style('width: 100%; margin-bottom: 1rem;')
                    
                    status = ui.select(
                        ['active', 'inactive'],
                        value=user.get('status', 'active'),
                        label='Status'
                    ).style('width: 100%; margin-bottom: 1rem;')
                    
                    def save_user():
                        if not all([first_name.value.strip(), last_name.value.strip(), email.value.strip()]):
                            ui.notify('Please fill in all required fields', color='negative')
                            return
                        
                        try:
                            update_data = {
                                'first_name': first_name.value.strip(),
                                'last_name': last_name.value.strip(),
                                'email': email.value.strip(),
                                'phone': phone.value.strip() if phone.value.strip() else None,
                                'role': role.value,
                                'department': department.value.strip() if department.value.strip() else None,
                                'status': status.value,
                                'updated_at': datetime.datetime.now()
                            }
                            
                            users_col.update_one(
                                {'_id': user['_id']},
                                {'$set': update_data}
                            )
                            
                            ui.notify('User updated successfully!', color='positive')
                            dialog.close()
                            ui.run_javascript('window.location.reload()')
                        except Exception as e:
                            ui.notify(f'Error updating user: {str(e)}', color='negative')
                    
                    with ui.row().style('gap: 1rem; margin-top: 1rem;'):
                        ui.button('Cancel', on_click=dialog.close).style('background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                        ui.button('Save Changes', on_click=save_user).style(f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
            
            dialog.open()
        
        def delete_user(user):
            with ui.dialog() as dialog:
                with ui.card().style('padding: 2rem; min-width: 400px;'):
                    ui.label('Confirm Deletion').style(f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;')
                    ui.label(f'Are you sure you want to delete user "{user.get("email", "Unknown")}"?').style('margin-bottom: 1.5rem;')
                    ui.label('This action cannot be undone.').style('color: var(--error-color); font-weight: bold; margin-bottom: 1.5rem;')
                    
                    def confirm_delete():
                        try:
                            users_col.delete_one({'_id': user['_id']})
                            ui.notify('User deleted successfully!', color='positive')
                            dialog.close()
                            ui.run_javascript('window.location.reload()')
                        except Exception as e:
                            ui.notify(f'Error deleting user: {str(e)}', color='negative')
                    
                    with ui.row().style('gap: 1rem; justify-content: flex-end;'):
                        ui.button('Cancel', on_click=dialog.close).style('background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
                        ui.button('Delete User', on_click=confirm_delete).style('background: var(--error-color); color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none;')
            
            dialog.open()
    
    institution_admin_sidebar(inst_id, content)

# Institution Admin Submissions Page
@ui.page('/institution_admin/{inst_id}/submissions')
def institution_admin_submissions(inst_id: str):
    """Institution admin page to view all submissions"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    def content(inst, main_color):
        ui.label('📥 All Submissions').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label(f'Institution: {inst.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        # Get all submissions for this institution
        from bson import ObjectId
        submissions = list(criteria_submissions_col.find({
            'institution_id': ObjectId(inst_id)
        }).sort('submitted_at', -1))
        
        if submissions:
            ui.label(f'Total Submissions: {len(submissions)}').style(
                'font-size: 1.1rem; color: var(--text-primary); margin-bottom: 1.5rem; text-align: center;'
            )
            
            for submission in submissions:
                with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 1rem; border-radius: 10px;'):
                    with ui.row().style('justify-content: space-between; align-items: center; margin-bottom: 1rem;'):
                        ui.label(f"📋 {submission.get('status', 'Unknown').title()}").style(
                            f'font-size: 1.2rem; font-weight: bold; color: {main_color};'
                        )
                        
                        status_color = 'success' if submission.get('status') == 'submitted' else 'warning'
                        ui.label(submission.get('status', 'Unknown').title()).style(
                            f'background: var(--{status_color}-color); color: white; padding: 0.3rem 0.8rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600;'
                        )
                    
                    # Get criteria details
                    criteria = criterias_col.find_one({'_id': submission.get('criteria_id')})
                    criteria_name = criteria.get('name', 'Unknown Criteria') if criteria else 'Unknown Criteria'
                    
                    ui.label(f"Criteria: {criteria_name}").style(
                        'font-size: 1rem; font-weight: 600; color: var(--text-primary); margin-bottom: 0.5rem;'
                    )
                    
                    ui.label(f"Submitted by: {submission.get('submitted_by', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    
                    ui.label(f"Submitted at: {submission.get('submitted_at', 'Unknown').strftime('%Y-%m-%d %H:%M') if hasattr(submission.get('submitted_at'), 'strftime') else submission.get('submitted_at', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    
                    ui.label(f"Data rows: {len(submission.get('data', []))}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 1rem;'
                    )
                    
                    # Show data preview
                    data = submission.get('data', [])
                    if data:
                        with ui.expansion('📊 View Data', icon='visibility').style('margin-top: 1rem;'):
                            for i, row in enumerate(data[:5]):  # Show first 5 rows
                                with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 0.8rem; margin-bottom: 0.5rem; border-radius: 6px;'):
                                    ui.label(f"Row {i+1}").style(
                                        'font-weight: bold; color: var(--text-primary); margin-bottom: 0.3rem;'
                                    )
                                    for header, value in row.items():
                                        ui.label(f"{header}: {value}").style(
                                            'font-size: 0.8rem; color: var(--text-secondary); margin-bottom: 0.1rem;'
                                        )
                            
                            if len(data) > 5:
                                ui.label(f"... and {len(data) - 5} more rows").style(
                                    'font-style: italic; color: var(--text-secondary); text-align: center; margin-top: 0.5rem;'
                                )
        else:
            ui.label('No submissions found for this institution.').style(
                'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
            )
    
    institution_admin_sidebar(inst_id, content)

# Program Admin Submissions Page
@ui.page('/program_admin/{program_id}/submissions')
def program_admin_submissions(program_id: str):
    """Program admin page to view their own submissions"""
    add_beautiful_global_styles()
    
    # Get program and institution data
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])}) if program else None
    def content():
        # Sidebar and navigation
        with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
            with ui.column().style('width: 200px; background: var(--surface); border-right: 2px solid var(--border); padding: 0; height: 100vh; overflow-y: auto;'):
                with ui.card().style('background: linear-gradient(135deg, var(--primary-color), var(--primary-dark)); color: white; padding: 1.5rem; margin: 1rem; border: none; border-radius: 12px; box-shadow: var(--shadow);'):
                    ui.element('div').style('width: 60px; height: 60px; background: rgba(255,255,255,0.2); border-radius: 50%; margin: 0 auto 1rem; display: flex; align-items: center; justify-content: center; font-size: 1.5rem; font-weight: bold;').inner_html = '🎓'
                    ui.label(institution.get('name', 'Institution')).style('font-size: 1.2rem; font-weight: bold; text-align: center; margin-bottom: 0.5rem;')
                    ui.separator().style('margin: 1rem 0; background: rgba(255,255,255,0.3);')
                    current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
                    ui.label(current_user_email).style('font-size: 0.9rem; text-align: center; opacity: 0.9; margin-bottom: 0.5rem;')
                    ui.label('Program Admin').style('font-size: 0.9rem; text-align: center; opacity: 0.8;')
                with ui.column().style('padding: 1rem; gap: 0.5rem;'):
                    with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 1rem; margin-bottom: 1rem;'):
                        ui.label('📊 OVERVIEW').style('font-size: 0.85rem; font-weight: 700; color: rgb(154, 44, 84); margin-bottom: 1rem; text-align: left; letter-spacing: 1px; text-transform: uppercase;')
                        overview_items = [
                            ('🏠', 'Dashboard', f'/program_admin/{program_id}'),
                            ('📊', 'Criteria Management', f'/program_admin/{program_id}/criterias'),
                            ('👤', 'Extended Profiles', f'/program_admin/{program_id}/profiles'),
                            ('📥', 'My Submissions', f'/program_admin/{program_id}/submissions'),
                        ]
                        for icon, label, url in overview_items:
                            ui.button(f'{icon} {label}', on_click=lambda u=url: ui.navigate.to(u)).style(
                                f'width: 100%; justify-content: flex-start; margin-bottom: 0.5rem; text-align: left; '
                                f'background: white; color: var(--text-primary); border: 1px solid #e9ecef; '
                                f'padding: 0.75rem 1rem; border-radius: 8px; '
                                f'transition: all 0.3s ease; font-weight: 500;'
                            ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "rgb(154, 44, 84)"; event.target.style.color = "white"; event.target.style.borderColor = "rgb(154, 44, 84)";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "white"; event.target.style.color = "var(--text-primary)"; event.target.style.borderColor = "#e9ecef";'))
                        ui.separator().style('margin: 1rem 0; background: #e9ecef;')
                        ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                            'width: 100%; justify-content: center; background: #dc3545; color: white; border: 1px solid #dc3545; padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 600;'
                        ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "#c82333"; event.target.style.borderColor = "#c82333";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "#dc3545"; event.target.style.borderColor = "#dc3545";'))
            # Main content
            with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
                ui.label('📥 My Submissions').style(
                    f'font-size: 2rem; font-weight: bold; color: rgb(154, 44, 84); margin-bottom: 1rem; text-align: center;'
                )
                # Get current user
                current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
                # Get submissions for this program and user (show both drafts and submitted)
        submissions = list(criteria_submissions_col.find({
            'program_id': ObjectId(program_id),
            'submitted_by': current_user_email
        }).sort('submitted_at', -1))
        if submissions:
            ui.label(f'Total Submissions: {len(submissions)}').style(
                'font-size: 1.1rem; color: var(--text-primary); margin-bottom: 1.5rem; text-align: center;'
            )
            for submission in submissions:
                with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 1rem; border-radius: 10px;'):
                    with ui.row().style('justify-content: space-between; align-items: center; margin-bottom: 1rem;'):
                        ui.label(f"📋 {submission.get('status', 'Unknown').title()}").style(
                                            'font-size: 1.2rem; font-weight: bold; color: rgb(154, 44, 84);'
                        )
                        status_color = 'success' if submission.get('status') == 'submitted' else 'warning'
                        ui.label(submission.get('status', 'Unknown').title()).style(
                            f'background: var(--{status_color}-color); color: white; padding: 0.3rem 0.8rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600;'
                        )
                    # Get criteria details
                    criteria = criterias_col.find_one({'_id': submission.get('criteria_id')})
                    criteria_name = criteria.get('name', 'Unknown Criteria') if criteria else 'Unknown Criteria'
                    ui.label(f"Criteria: {criteria_name}").style(
                        'font-size: 1rem; font-weight: 600; color: var(--text-primary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Submitted by: {submission.get('submitted_by', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Submitted at: {submission.get('submitted_at', 'Unknown').strftime('%Y-%m-%d %H:%M') if hasattr(submission.get('submitted_at'), 'strftime') else submission.get('submitted_at', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Data rows: {len(submission.get('data', []))}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 1rem;'
                    )
                    # Show data preview
                    data = submission.get('data', [])
                    if data:
                        with ui.expansion('📊 View Data', icon='visibility').style('margin-top: 1rem;'):
                            for i, row in enumerate(data[:5]):  # Show first 5 rows
                                with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 0.8rem; margin-bottom: 0.5rem; border-radius: 6px;'):
                                    ui.label(f"Row {i+1}").style(
                                        'font-weight: bold; color: var(--text-primary); margin-bottom: 0.3rem;'
                                    )
                                    for header, value in row.items():
                                        ui.label(f"{header}: {value}").style(
                                            'font-size: 0.8rem; color: var(--text-secondary); margin-bottom: 0.1rem;'
                                        )
                            if len(data) > 5:
                                ui.label(f"... and {len(data) - 5} more rows").style(
                                    'font-style: italic; color: var(--text-secondary); text-align: center; margin-top: 0.5rem;'
                                )
        else:
            ui.label('No submissions found for this program.').style(
                'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
            )
    content()

# Department Admin Submissions Page
@ui.page('/department_admin/{department_id}/submissions')
def department_admin_submissions(department_id: str):
    """Department admin page to view their own submissions"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        def content():
            from bson import ObjectId
            # Sidebar and navigation
    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
        with ui.column().style('width: 280px; background: var(--surface); border-right: 2px solid var(--border); padding: 0; height: 100vh; overflow-y: auto;'):
            with ui.card().style('background: linear-gradient(135deg, var(--primary-color), var(--primary-dark)); color: white; padding: 1.5rem; margin: 1rem; border: none; border-radius: 12px; box-shadow: var(--shadow);'):
                ui.element('div').style('width: 60px; height: 60px; background: rgba(255,255,255,0.2); border-radius: 50%; margin: 0 auto 1rem; display: flex; align-items: center; justify-content: center; font-size: 1.5rem; font-weight: bold;').inner_html = '🎓'
                ui.label(institution.get('name', 'Institution')).style('font-size: 1.2rem; font-weight: bold; text-align: center; margin-bottom: 0.5rem;')
                ui.separator().style('margin: 1rem 0; background: rgba(255,255,255,0.3);')
                current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
                ui.label(current_user_email).style('font-size: 0.9rem; text-align: center; opacity: 0.9; margin-bottom: 0.5rem;')
                ui.label('Program Admin').style('font-size: 0.9rem; text-align: center; opacity: 0.8;')
            with ui.column().style('padding: 1rem; gap: 0.5rem;'):
                with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 1rem; margin-bottom: 1rem;'):
                    ui.label('📊 OVERVIEW').style('font-size: 0.85rem; font-weight: 700; color: rgb(154, 44, 84); margin-bottom: 1rem; text-align: left; letter-spacing: 1px; text-transform: uppercase;')
                    overview_items = [
                        ('🏠', 'Dashboard', f'/program_admin/{program_id}'),
                        ('📊', 'Criteria Management', f'/program_admin/{program_id}/criterias'),
                        ('👤', 'Extended Profiles', f'/program_admin/{program_id}/profiles'),
                        ('📥', 'My Submissions', f'/program_admin/{program_id}/submissions'),
                    ]
                    for icon, label, url in overview_items:
                        ui.button(f'{icon} {label}', on_click=lambda u=url: ui.navigate.to(u)).style(
                            f'width: 100%; justify-content: flex-start; margin-bottom: 0.5rem; text-align: left; '
                            f'background: white; color: var(--text-primary); border: 1px solid #e9ecef; '
                            f'padding: 0.75rem 1rem; border-radius: 8px; '
                            f'transition: all 0.3s ease; font-weight: 500;'
                        ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "rgb(154, 44, 84)"; event.target.style.color = "white"; event.target.style.borderColor = "rgb(154, 44, 84)";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "white"; event.target.style.color = "var(--text-primary)"; event.target.style.borderColor = "#e9ecef";'))
                    ui.separator().style('margin: 1rem 0; background: #e9ecef;')
                    ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                                'width: 100%; justify-content: center; background: #dc3545; color: white; border: 1px solid #dc3545; padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 600;'
                    ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "#c82333"; event.target.style.borderColor = "#c82333";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "#dc3545"; event.target.style.borderColor = "#dc3545";'))
        # Main content
        with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
            ui.label('📥 My Submissions').style(
                f'font-size: 2rem; font-weight: bold; color: rgb(154, 44, 84); margin-bottom: 1rem; text-align: center;'
            )
            # Get current user
            current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
            # Get submissions for this program and user
        submissions = list(criteria_submissions_col.find({
                'program_id': ObjectId(program_id),
            'submitted_by': current_user_email
        }).sort('submitted_at', -1))
        if submissions:
            ui.label(f'Total Submissions: {len(submissions)}').style(
                'font-size: 1.1rem; color: var(--text-primary); margin-bottom: 1.5rem; text-align: center;'
            )
            for submission in submissions:
                with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 1rem; border-radius: 10px;'):
                    with ui.row().style('justify-content: space-between; align-items: center; margin-bottom: 1rem;'):
                        ui.label(f"📋 {submission.get('status', 'Unknown').title()}").style(
                                'font-size: 1.2rem; font-weight: bold; color: rgb(154, 44, 84);'
                        )
                        status_color = 'success' if submission.get('status') == 'submitted' else 'warning'
                        ui.label(submission.get('status', 'Unknown').title()).style(
                            f'background: var(--{status_color}-color); color: white; padding: 0.3rem 0.8rem; border-radius: 20px; font-size: 0.8rem; font-weight: 600;'
                        )
                    # Get criteria details
                    criteria = criterias_col.find_one({'_id': submission.get('criteria_id')})
                    criteria_name = criteria.get('name', 'Unknown Criteria') if criteria else 'Unknown Criteria'
                    ui.label(f"Criteria: {criteria_name}").style(
                        'font-size: 1rem; font-weight: 600; color: var(--text-primary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Submitted by: {submission.get('submitted_by', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Submitted at: {submission.get('submitted_at', 'Unknown').strftime('%Y-%m-%d %H:%M') if hasattr(submission.get('submitted_at'), 'strftime') else submission.get('submitted_at', 'Unknown')}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                    )
                    ui.label(f"Data rows: {len(submission.get('data', []))}").style(
                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 1rem;'
                    )
                    # Show data preview
                    data = submission.get('data', [])
                    if data:
                        with ui.expansion('📊 View Data', icon='visibility').style('margin-top: 1rem;'):
                            for i, row in enumerate(data[:5]):  # Show first 5 rows
                                with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 0.8rem; margin-bottom: 0.5rem; border-radius: 6px;'):
                                    ui.label(f"Row {i+1}").style(
                                        'font-weight: bold; color: var(--text-primary); margin-bottom: 0.3rem;'
                                    )
                                    for header, value in row.items():
                                        ui.label(f"{header}: {value}").style(
                                            'font-size: 0.8rem; color: var(--text-secondary); margin-bottom: 0.1rem;'
                                        )
                            if len(data) > 5:
                                ui.label(f"... and {len(data) - 5} more rows").style(
                                    'font-style: italic; color: var(--text-secondary); text-align: center; margin-top: 0.5rem;'
                                )
        else:
                ui.label('No submissions found for this program.').style(
                'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
            )
        
        # Main content
        with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
            content()

@ui.page('/institution_admin/{inst_id}/academic_years')
def institution_admin_academic_years(inst_id: str):
    """Academic years management page with copy functionality"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    def content(inst, main_color):
        ui.label('Academic Years Management').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Create new academic year button
        def open_create_dialog():
            with ui.dialog() as dialog:
                with ui.card().style(
                    'padding: 2rem; min-width: 600px; max-width: 800px; max-height: 90vh; overflow-y: auto;'
                ).classes('beautiful-card'):
                    ui.label('Create New Academic Year').style(
                        f'font-size: 1.8rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem; text-align: center;'
                    )
                    
                    with ui.column().style('width: 100%; gap: 1.5rem;'):
                        year_name = ui.input('Academic Year Name').classes('beautiful-input').style('width: 100%;')
                        year_description = ui.textarea('Description').classes('beautiful-input').style('width: 100%; min-height: 100px;')
                        
                        # Copy from existing year
                        existing_years = list(academic_years_col.find({'institution_id': inst_id}).sort('created_at', -1))
                        if existing_years:
                            ui.label('Copy from Existing Year (Optional)').style(
                                f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-top: 1rem;'
                            )
                            
                            copy_from_year = ui.select(
                                options=['None'] + [y['name'] for y in existing_years],
                                value='None',
                                label='Select year to copy from'
                            ).classes('beautiful-input').style('width: 100%;')
                            
                            # Copy options
                            copy_options = ui.select(
                                label='What to copy?',
                                options=['Copy Institutional Hierarchy', 'Copy Criterias', 'Copy Extended Profiles'],
                                multiple=True,
                                value=[]
                            ).style('width: 100%;')
                        
                        # Create function
                        def create_year():
                            if not year_name.value:
                                ui.notify('Year name is required', color='negative')
                                return
                            
                            year_doc = {
                                'name': year_name.value,
                                'description': year_description.value,
                                'institution_id': inst_id,
                                'created_at': datetime.now(timezone.utc),
                                'updated_at': datetime.now(timezone.utc),
                                'is_locked': False
                            }
                            
                            try:
                                result = academic_years_col.insert_one(year_doc)
                                new_year_id = str(result.inserted_id)
                                
                                # Copy selected items if requested
                                if existing_years and copy_from_year.value != 'None':
                                    source_year = next((y for y in existing_years if y['name'] == copy_from_year.value), None)
                                    if source_year:
                                        copy_counts = {'hierarchy': 0, 'criterias': 0, 'profiles': 0}
                                        
                                        # Copy hierarchy
                                        if 'Copy Institutional Hierarchy' in copy_options.value:
                                            source_schools = list(schools_col.find({'institution_id': inst_id, 'academic_year_id': source_year['_id']}))
                                            for school in source_schools:
                                                school_copy = school.copy()
                                                del school_copy['_id']
                                                school_copy['academic_year_id'] = ObjectId(new_year_id)
                                                schools_col.insert_one(school_copy)
                                                copy_counts['hierarchy'] += 1
                                        
                                        # Copy criterias
                                        if 'Copy Criterias' in copy_options.value:
                                            source_criterias = list(criterias_col.find({'institution_id': inst_id, 'academic_year_id': source_year['_id']}))
                                            for criteria in source_criterias:
                                                criteria_copy = criteria.copy()
                                                del criteria_copy['_id']
                                                criteria_copy['academic_year_id'] = ObjectId(new_year_id)
                                                criterias_col.insert_one(criteria_copy)
                                                copy_counts['criterias'] += 1
                                        
                                        # Copy extended profiles
                                        if 'Copy Extended Profiles' in copy_options.value:
                                            source_profiles = list(extended_profiles_col.find({'institution_id': inst_id, 'academic_year_id': source_year['_id']}))
                                            for profile in source_profiles:
                                                profile_copy = profile.copy()
                                                del profile_copy['_id']
                                                profile_copy['academic_year_id'] = ObjectId(new_year_id)
                                                extended_profiles_col.insert_one(profile_copy)
                                                copy_counts['profiles'] += 1
                                        
                                        copy_summary = f"Copied: {copy_counts['hierarchy']} hierarchy items, {copy_counts['criterias']} criterias, {copy_counts['profiles']} profiles"
                                        ui.notify(f'Academic year created successfully! {copy_summary}', color='positive')
                                    else:
                                        ui.notify('Academic year created successfully!', color='positive')
                                else:
                                    ui.notify('Academic year created successfully!', color='positive')
                                
                                dialog.close()
                                ui.run_javascript('window.location.reload()')
                            
                            except Exception as e:
                                ui.notify(f'Error creating academic year: {str(e)}', color='negative')
                        
                        # Action buttons
                        with ui.row().style('width: 100%; justify-content: space-between; margin-top: 2rem;'):
                            ui.button('Cancel', on_click=dialog.close).classes('btn-secondary')
                            ui.button('Create Year', on_click=create_year).classes('btn-primary')
                
                dialog.open()
        
        ui.button('+ Create Academic Year', on_click=open_create_dialog).classes('btn-primary').style('margin-bottom: 2rem;')
        
        # Display existing academic years
        existing_years = list(academic_years_col.find({'institution_id': inst_id}).sort('created_at', -1))
        
        if existing_years:
            ui.label(f'Existing Academic Years ({len(existing_years)} found)').style(
                f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            
            for year in existing_years:
                with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1rem; padding: 1.5rem;'):
                    with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                        with ui.column().style('flex: 1;'):
                            ui.label(f"📅 {year.get('name', 'Unnamed Year')}").style(
                                'font-size: 1.2rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                            )
                            
                            if year.get('description'):
                                ui.label(f"Description: {year['description']}").style(
                                    'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.3rem;'
                                )
                            
                            status = '🔒 Locked' if year.get('is_locked', False) else '🔓 Active'
                            ui.label(f"Status: {status}").style(
                                'font-size: 0.9rem; color: var(--success-color) if not year.get("is_locked") else var(--warning-color);'
                            )
                            
                            created_at = year.get('created_at')
                            if created_at:
                                ui.label(f"Created: {created_at.strftime('%Y-%m-%d %H:%M')}").style(
                                    'font-size: 0.8rem; color: var(--text-secondary);'
                                )
                        
                        with ui.row().style('gap: 0.5rem;'):
                            # Lock/Unlock button
                            if year.get('is_locked', False):
                                ui.button('🔓 Unlock', on_click=lambda y=year: toggle_lock(y)).style(
                                    f'background: var(--success-color); color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                                )
                            else:
                                ui.button('🔒 Lock', on_click=lambda y=year: toggle_lock(y)).style(
                                    f'background: var(--warning-color); color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                                )
                            
                            # Delete button
                            ui.button('🗑️ Delete', on_click=lambda y=year: delete_year(y)).style(
                                'background: var(--error-color); color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                            )
        
        else:
            ui.label('No academic years found. Create your first one!').style(
                'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
            )
        
        def toggle_lock(year):
            try:
                new_lock_status = not year.get('is_locked', False)
                academic_years_col.update_one(
                    {'_id': year['_id']},
                    {'$set': {'is_locked': new_lock_status, 'updated_at': datetime.datetime.now(datetime.timezone.utc)}}
                )
                
                action = 'locked' if new_lock_status else 'unlocked'
                ui.notify(f'Academic year "{year["name"]}" {action} successfully!', color='positive')
                ui.run_javascript('window.location.reload()')
            
            except Exception as e:
                ui.notify(f'Error updating academic year: {str(e)}', color='negative')
        
        def delete_year(year):
            try:
                # Check if year has data
                has_data = (
                    schools_col.count_documents({'academic_year_id': year['_id']}) > 0 or
                    criterias_col.count_documents({'academic_year_id': year['_id']}) > 0 or
                    extended_profiles_col.count_documents({'academic_year_id': year['_id']}) > 0
                )
                
                if has_data:
                    ui.notify('Cannot delete academic year with existing data. Lock it instead.', color='warning')
                    return
                
                academic_years_col.delete_one({'_id': year['_id']})
                ui.notify(f'Academic year "{year["name"]}" deleted successfully!', color='positive')
                ui.run_javascript('window.location.reload()')
            
            except Exception as e:
                ui.notify(f'Error deleting academic year: {str(e)}', color='negative')
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/institution_admin/{inst_id}/audit_logs')
def institution_admin_audit_logs(inst_id: str):
    """Audit logs management page with filters and pagination"""
    add_beautiful_global_styles()
    
    def content(inst, main_color):
        ui.label('Audit Logs').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem;'
        )
        
        # Filters container
        with ui.card().classes('beautiful-card').style('width: 100%; margin-bottom: 1rem; padding: 1.5rem;'):
            ui.label('Filters & Search').style(f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;')
            
            with ui.row().style('width: 100%; gap: 1rem; align-items: end;'):
                # Search input
                search_input = ui.input('Search logs...', placeholder='Search by action, user, or details').style('flex: 2; min-width: 200px;')
                
                # Action filter
                action_filter = ui.select(
                    ['All Actions', 'CREATE', 'UPDATE', 'DELETE', 'LOGIN', 'LOGOUT'],
                    value='All Actions',
                    label='Action Type'
                ).style('flex: 1; min-width: 150px;')
                
                # Date range
                date_from = ui.input('Start Date').props('type=date').classes('beautiful-input').style('flex: 1; min-width: 150px;')
                date_to = ui.input('End Date').props('type=date').classes('beautiful-input').style('flex: 1; min-width: 150px;')
                
                # Filter button
                filter_btn = ui.button('🔍 Filter', on_click=lambda: load_logs()).classes('btn-primary')
                clear_btn = ui.button('🔄 Clear', on_click=lambda: clear_filters()).style('background: #6c757d; color: white; padding: 0.75rem 1rem; border-radius: 8px; border: none;')
        
        # Pagination state
        current_page = 1
        page_size = 10
        
        # Logs container
        logs_container = ui.column().style('width: 100%;')
        
        # Pagination container
        pagination_container = ui.row().style('width: 100%; justify-content: center; margin-top: 1rem; gap: 1rem;')
        
        def clear_filters():
            search_input.value = ''
            action_filter.value = 'All Actions'
            date_from.value = None
            date_to.value = None
            load_logs()
        
        def load_logs(page=1):
            nonlocal current_page
            current_page = page
            
            # Build filter query
            filter_query = {'institution_id': inst_id}
            
            # Search filter
            if search_input.value.strip():
                search_term = search_input.value.strip()
                filter_query['$or'] = [
                    {'action': {'$regex': search_term, '$options': 'i'}},
                    {'user_email': {'$regex': search_term, '$options': 'i'}},
                    {'details': {'$regex': search_term, '$options': 'i'}}
                ]
            
            # Action filter
            if action_filter.value != 'All Actions':
                filter_query['action'] = action_filter.value
            
            # Date filters
            date_filter = {}
            if date_from.value:
                # Handle both string and date object types
                if isinstance(date_from.value, str):
                    try:
                        from_date = datetime.datetime.strptime(date_from.value, '%Y-%m-%d').date()
                        date_filter['$gte'] = datetime.datetime.combine(from_date, datetime.time.min)
                    except ValueError:
                        pass  # Skip invalid date format
                elif hasattr(date_from.value, 'year'):  # It's a date object
                    date_filter['$gte'] = datetime.datetime.combine(date_from.value, datetime.time.min)
            
            if date_to.value:
                # Handle both string and date object types
                if isinstance(date_to.value, str):
                    try:
                        to_date = datetime.datetime.strptime(date_to.value, '%Y-%m-%d').date()
                        date_filter['$lte'] = datetime.datetime.combine(to_date, datetime.time.max)
                    except ValueError:
                        pass  # Skip invalid date format
                elif hasattr(date_to.value, 'year'):  # It's a date object
                    date_filter['$lte'] = datetime.datetime.combine(date_to.value, datetime.time.max)
            
            if date_filter:
                filter_query['timestamp'] = date_filter
            
            # Get total count for pagination
            total_logs = audit_logs_col.count_documents(filter_query)
            total_pages = max(1, (total_logs + page_size - 1) // page_size)
            
            # Get paginated logs
            skip = (current_page - 1) * page_size
            logs = list(audit_logs_col.find(filter_query)
                       .sort('timestamp', -1)
                       .skip(skip)
                       .limit(page_size))
            
            # Clear containers
            logs_container.clear()
            pagination_container.clear()
            
            # Display logs
            with logs_container:
                if logs:
                    # Stats
                    with ui.row().style('width: 100%; margin-bottom: 1rem; align-items: center; justify-content: space-between;'):
                        ui.label(f'Showing {len(logs)} of {total_logs} logs (Page {current_page} of {total_pages})').style('color: var(--text-secondary);')
                        ui.label(f'Total: {total_logs} audit entries').style(f'color: {main_color}; font-weight: bold;')
                    
                    # Logs table
                    with ui.card().classes('beautiful-card').style('width: 100%; padding: 0; overflow: hidden;'):
                        with ui.element('div').style('overflow-x: auto;'):
                            with ui.element('table').style('width: 100%; border-collapse: collapse;'):
                                # Header
                                with ui.element('thead'):
                                    with ui.element('tr'):
                                        headers = ['Timestamp', 'Action', 'User', 'Details', 'IP Address']
                                        for header in headers:
                                            with ui.element('th').style(f'background: {main_color}; color: white; padding: 12px; text-align: left; border: 1px solid var(--border); font-weight: bold;'):
                                                ui.label(header)
                                
                                # Rows
                                with ui.element('tbody'):
                                    for i, log in enumerate(logs):
                                        row_bg = '#f8f9fa' if i % 2 == 0 else 'white'
                                        
                                        with ui.element('tr').style(f'background: {row_bg};'):
                                            # Timestamp
                                            with ui.element('td').style('padding: 12px; border: 1px solid var(--border); white-space: nowrap;'):
                                                timestamp = log.get('timestamp', datetime.now(timezone.utc))
                                                if isinstance(timestamp, str):
                                                    ui.label(timestamp)
                                                else:
                                                    ui.label(timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                                            
                                            # Action
                                            with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                                action = log.get('action', 'Unknown')
                                                action_color = {
                                                    'CREATE': '#28a745',
                                                    'UPDATE': '#ffc107',
                                                    'DELETE': '#dc3545',
                                                    'LOGIN': '#007bff',
                                                    'LOGOUT': '#6c757d'
                                                }.get(action, '#6c757d')
                                                
                                                ui.label(action).style(f'background: {action_color}; color: white; padding: 4px 8px; border-radius: 4px; font-size: 0.8rem; font-weight: bold;')
                                            
                                            # User
                                            with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                                ui.label(log.get('user_email', 'System'))
                                            
                                            # Details
                                            with ui.element('td').style('padding: 12px; border: 1px solid var(--border); max-width: 300px;'):
                                                details = log.get('details', '')
                                                if len(details) > 100:
                                                    ui.label(details[:100] + '...').style('color: var(--text-secondary);')
                                                else:
                                                    ui.label(details).style('color: var(--text-secondary);')
                                            
                                            # IP Address
                                            with ui.element('td').style('padding: 12px; border: 1px solid var(--border);'):
                                                ui.label(log.get('ip_address', 'N/A')).style('font-family: monospace; font-size: 0.9rem;')
                else:
                    with ui.card().classes('beautiful-card').style('width: 100%; padding: 3rem; text-align: center;'):
                        ui.label('No audit logs found').style(f'font-size: 1.5rem; color: {main_color}; margin-bottom: 1rem;')
                        ui.label('Logs will appear here as users perform actions in the system.').style('color: var(--text-secondary);')
            
            # Pagination
            with pagination_container:
                if total_pages > 1:
                    # Previous button
                    prev_disabled = current_page <= 1
                    ui.button(
                        '← Previous',
                        on_click=lambda: load_logs(current_page - 1),
                        color='primary' if not prev_disabled else 'grey'
                    ).style(f'{"pointer-events: none; opacity: 0.5;" if prev_disabled else ""}')
                    
                    # Page numbers
                    start_page = max(1, current_page - 2)
                    end_page = min(total_pages, current_page + 2)
                    
                    for page_num in range(start_page, end_page + 1):
                        is_current = page_num == current_page
                        ui.button(
                            str(page_num),
                            on_click=lambda p=page_num: load_logs(p),
                            color='primary' if is_current else 'grey'
                        ).style(f'min-width: 40px; {"font-weight: bold;" if is_current else ""}')
                    
                    # Next button
                    next_disabled = current_page >= total_pages
                    ui.button(
                        'Next →',
                        on_click=lambda: load_logs(current_page + 1),
                        color='primary' if not next_disabled else 'grey'
                    ).style(f'{"pointer-events: none; opacity: 0.5;" if next_disabled else ""}')
                    
                    # Page info
                    ui.label(f'Page {current_page} of {total_pages}').style('color: var(--text-secondary); margin-left: 1rem; align-self: center;')
        
        # Initial load
        load_logs()
    
    institution_admin_sidebar(inst_id, content)

@ui.page('/')
def login_page():
    # Add beautiful global styles
    add_beautiful_global_styles()
    
    # Ensure the page has no margins/padding and login is perfectly centered
    ui.add_head_html('''
        <style>
            body, html {
                margin: 0 !important;
                padding: 0 !important;
                overflow: hidden !important;
                height: 100vh !important;
                width: 100vw !important;
                background: white !important;
            }
            .nicegui-content {
                margin: 0 !important;
                padding: 0 !important;
                height: 100vh !important;
                width: 100vw !important;
            }
            .login-input .q-field__control {
                padding: 0.75rem !important;
                border: 2px solid rgba(154, 44, 84, 0.2) !important;
                border-radius: 8px !important;
                font-size: 1rem !important;
                background: white !important;
            }
            .login-input .q-field__control:focus-within {
                border-color: rgb(154, 44, 84) !important;
                box-shadow: 0 0 0 3px rgba(154, 44, 84, 0.1) !important;
            }
            .login-input .q-field__label {
                color: rgba(154, 44, 84, 0.7) !important;
                font-weight: 500 !important;
            }
            .login-input input {
                padding: 0.5rem !important;
                font-size: 1rem !important;
            }
        </style>
    ''')
    
    def show_forgot_password():
        with ui.dialog() as dialog:
            with ui.card().style('min-width: 400px; padding: 2rem; background: white; border-radius: 12px;'):
                ui.label('Reset Password').style('font-size: 1.5rem; font-weight: bold; margin-bottom: 1.5rem; color: rgb(154, 44, 84); text-align: center;')
                
                ui.label('Enter your email address to receive a password reset link.').style('color: #666; text-align: center; margin-bottom: 1rem;')
                forgot_email = ui.input('Email Address', placeholder='Enter your email').props('type=email').style('width: 100%; margin-bottom: 1rem;')
                
                with ui.row().style('width: 100%; gap: 1rem; margin-top: 1rem;'):
                    ui.button('Cancel', on_click=dialog.close).style('flex: 1; background: #6c757d; color: white; padding: 0.75rem; border-radius: 8px; border: none;')
                    ui.button('Send Reset Link', on_click=lambda: (ui.notify(f'Reset link sent to {forgot_email.value}', color='positive'), dialog.close())).style('flex: 1; background: rgb(154, 44, 84); color: white; padding: 0.75rem; border-radius: 8px; border: none;')
        dialog.open()

    def do_login():
        global current_user
        
        print(f"DEBUG: Login attempt for email: {email.value}")
        
        if not email.value or not password.value:
            ui.notify('Please enter both email and password', color='warning')
            return
        
        try:
            # Test database connection
            print(f"DEBUG: Testing database connection...")
            db_test = db.command('ping')
            print(f"DEBUG: Database connection successful: {db_test}")
            
            # Check if users collection exists and has data
            users_count = users_col.count_documents({})
            print(f"DEBUG: Users collection has {users_count} documents")
            
            user = users_col.find_one({'email': email.value})
            print(f"DEBUG: User lookup result: {user is not None}")
            
            if not user:
                ui.notify('Invalid email address', color='negative')
                # Log failed login attempt
                log_audit_action(
                    action='Failed Login Attempt',
                    details=f'Invalid email: {email.value}',
                    user_email=email.value,
                    institution_id=None
                )
                return
            
            entered_hash = hash_password(password.value, user['salt'])
            print(f"DEBUG: Password hash comparison - entered: {entered_hash[:20]}..., stored: {user['password_hash'][:20]}...")
            
            if entered_hash != user['password_hash']:
                ui.notify('Incorrect password', color='negative')
                print(f"DEBUG: Password mismatch for user: {email.value}")
                # Log failed login attempt
                log_audit_action(
                    action='Failed Login Attempt',
                    details=f'Incorrect password for email: {email.value}',
                    user_email=email.value,
                    institution_id=user.get('institution_id')
                )
                return
            
            print(f"DEBUG: Password verification successful for user: {email.value}")
            
            # Save user info in global session
            current_user = {
                'email': user['email'],
                'role': user.get('role', ''),
                'institution_id': str(user.get('institution_id', '')),
                'school_id': str(user.get('school_id', '')),
                'program_id': str(user.get('program_id', '')),
                'department_id': str(user.get('department_id', '')),
                'must_change_password': user.get('must_change_password', False),
                'name': user.get('name', user['email'].split('@')[0])
            }
            
            # Store in app.storage for persistence
            if not hasattr(app.storage, 'user'):
                app.storage.user = {}
            app.storage.user['user'] = current_user
            
        except Exception as e:
            print(f"DEBUG: Database error during login: {str(e)}")
            ui.notify(f'Database connection error: {str(e)}', color='negative')
            return
        
        # Log successful login
        log_audit_action(
            action='Successful Login',
            details=f'User logged in with role: {user.get("role", "Unknown")}',
            user_email=user['email'],
            institution_id=user.get('institution_id')
        )
        
        # Clear any stored academic year selection on login to avoid stale/invalid values
        if hasattr(app.storage, 'user'):
            app.storage.user['selected_academic_year_id'] = None
            
        ui.notify(f'Welcome back!', color='positive')
        
        if user.get('must_change_password', False):
            ui.notify('Redirecting to change password page...', color='info')
            ui.navigate.to('/change_password')
        else:
            # Redirect based on user role
            if user.get('role', '') == 'Institution Admin' and user.get('institution_id'):
                ui.notify(f"Redirecting to Institution Admin page: /institution_admin/{user['institution_id']}", color='info')
                ui.navigate.to(f"/institution_admin/{user['institution_id']}")
            elif user.get('role', '') == 'Program Admin' and user.get('program_id'):
                ui.notify(f"Redirecting to Program Admin page: /program_admin/{user['program_id']}", color='info')
                ui.navigate.to(f"/program_admin/{user['program_id']}")
            elif user.get('role', '') == 'Department Admin' and user.get('department_id'):
                ui.notify(f"Redirecting to Department Admin page: /department_admin/{user['department_id']}", color='info')
                ui.navigate.to(f"/department_admin/{user['department_id']}")
            elif user.get('role', '') == 'School Admin' and user.get('school_id'):
                ui.notify(f"Redirecting to School Admin page: /school_admin/{user['school_id']}", color='info')
                ui.navigate.to(f"/school_admin/{user['school_id']}")
            else:
                ui.notify('Redirecting to dashboard...', color='info')
                ui.navigate.to('/dashboard')

    # Enhanced Login Box with Better UI
    with ui.element('div').style('''
        position: fixed;
        top: 0;
        left: 0;
        width: 100vw;
        height: 100vh;
        display: flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        margin: 0;
        padding: 0;
        box-sizing: border-box;
        overflow: hidden;
    '''):
        with ui.card().style('''
            width: 450px;
            max-width: 90vw;
            padding: 3.5rem;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(154, 44, 84, 0.15);
            border: 2px solid rgba(154, 44, 84, 0.1);
            position: relative;
            z-index: 1;
        '''):
            
            # Header with Logo
            with ui.column().classes('w-full items-center justify-center').style('margin-bottom: 3rem; width: 100%;'):
                # ui.html('<div style="width: 80px; height: 80px; background: linear-gradient(135deg, rgb(154, 44, 84) 0%, rgb(124, 35, 67) 100%); border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-bottom: 1rem;"><span style="color: white; font-size: 2rem; font-weight: bold;">IQ</span></div>')
                ui.label('ITM SKILLS UNIVERSITY').style('font-size: 2.2rem; font-weight: 800; color: rgb(154, 44, 84); margin-bottom: 0.5rem; letter-spacing: -0.5px; text-align: center; width: 100%;')
                ui.label('Institutional Quality Assurance').style('font-size: 1rem; color: #666; font-weight: 500; text-align: center; width: 100%;')

            # Login Form
            with ui.column().classes('w-full justify-center').style('gap: 1.8rem; margin-bottom: 2.5rem; width: 100%;'):
                email = ui.input('Email Address', placeholder='Enter your email address').props('type=email').style('width: 100%; max-width: 100%; text-align: left;').classes('login-input w-full')
                password = ui.input('Password', placeholder='Enter your password').props('type=password').style('width: 100%; max-width: 100%; text-align: left;').classes('login-input w-full')
                password.on('keydown.enter', do_login)  # Login on Enter key

            # Action Buttons
            with ui.column().classes('w-full justify-center').style('gap: 1.2rem; width: 100%;'):
                ui.button('Sign In', on_click=do_login).style('''
                    width: 100%; 
                    padding: 1.2rem; 
                    background: linear-gradient(135deg, rgb(154, 44, 84) 0%, rgb(124, 35, 67) 100%); 
                    color: white; 
                    border: none; 
                    border-radius: 10px; 
                    font-size: 1.1rem; 
                    font-weight: 600;
                    cursor: pointer;
                    box-shadow: 0 4px 15px rgba(154, 44, 84, 0.3);
                    transition: all 0.3s ease;
                ''').classes('w-full')

                ui.button('Forgot Password?', on_click=show_forgot_password).style('''
                    background: none; 
                    color: rgb(154, 44, 84); 
                    border: none; 
                    font-size: 0.95rem; 
                    text-decoration: none; 
                    cursor: pointer; 
                    padding: 0.8rem;
                    width: 100%;
                    border-radius: 8px;
                    transition: all 0.3s ease;
                ''').classes('w-full').on('mouseenter', lambda: ui.run_javascript('event.target.style.background = "rgba(154, 44, 84, 0.1)"')).on('mouseleave', lambda: ui.run_javascript('event.target.style.background = "transparent"'))

@ui.page('/dashboard')
def dashboard():
    """Platform owner dashboard"""
    add_beautiful_global_styles()
    
    global current_user
    
    # Check authentication
    if not check_auth() or current_user.get('role') != 'Platform Owner':
        ui.notify('Access denied', color='negative')
        ui.navigate.to('/')
        return
    
    with ui.column().style('min-height: 100vh; background: var(--background); padding: 2rem;'):
        # Header
        with ui.row().style('width: 100%; align-items: center; justify-content: space-between; margin-bottom: 2rem;'):
            ui.label('Platform Dashboard').style(
                'font-size: 2.5rem; font-weight: bold; color: var(--primary-color);'
            )
            
            def logout():
                global current_user
                current_user = None
                if hasattr(app.storage, 'user'):
                    app.storage.user.clear()
                ui.notify('Logged out successfully', color='positive')
                ui.navigate.to('/')
            
            ui.button('Logout', on_click=logout).classes('btn-secondary')
        
        # Stats cards
        with ui.row().style('width: 100%; gap: 1rem; margin-bottom: 2rem;'):
            # Total institutions
            inst_count = institutions_col.count_documents({})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('🏛️').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(inst_count)).style('font-size: 2rem; font-weight: bold; color: var(--primary-color);')
                ui.label('Institutions').style('color: var(--text-secondary);')
            
            # Total users
            users_count = users_col.count_documents({})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('👥').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(users_count)).style('font-size: 2rem; font-weight: bold; color: var(--success-color);')
                ui.label('Users').style('color: var(--text-secondary);')
            
            # Total criterias
            criterias_count = criterias_col.count_documents({})
            with ui.card().classes('beautiful-card slide-up').style('flex: 1; padding: 1.5rem;'):
                ui.label('📊').style('font-size: 2rem; margin-bottom: 0.5rem;')
                ui.label(str(criterias_count)).style('font-size: 2rem; font-weight: bold; color: var(--info-color);')
                ui.label('Criterias').style('color: var(--text-secondary);')
        
        # Actions
        with ui.row().style('width: 100%; gap: 1rem; margin-bottom: 2rem;'):
            def create_institution():
                ui.navigate.to('/create_institution')
            
            ui.button('+ Create New Institution', on_click=create_institution).classes('btn-primary')
        
        # Institutions list
        institutions = list(institutions_col.find())
        
        if institutions:
            ui.label('Institutions').style(
                'font-size: 1.5rem; font-weight: bold; color: var(--text-primary); margin-bottom: 1rem;'
            )
            
            with ui.column().style('width: 100%; gap: 1rem;'):
                for inst in institutions:
                    with ui.card().classes('beautiful-card').style('padding: 1.5rem;'):
                        with ui.row().style('width: 100%; align-items: center; justify-content: space-between;'):
                            with ui.row().style('align-items: center;'):
                                if inst.get('logo'):
                                    ui.image(inst['logo']).style('width: 60px; height: 60px; border-radius: 8px; margin-right: 1rem;')
                                
                                with ui.column():
                                    ui.label(inst.get('name', 'Unnamed Institution')).style(
                                        'font-size: 1.2rem; font-weight: bold; color: var(--text-primary);'
                                    )
                                    if inst.get('website_url'):
                                        ui.label(inst['website_url']).style('color: var(--text-secondary); font-size: 0.9rem;')
                                    if inst.get('city'):
                                        ui.label(f"📍 {inst['city']}, {inst.get('state', '')}").style(
                                            'color: var(--text-secondary); font-size: 0.9rem;'
                                        )
                            
                            ui.button(
                                'Manage',
                                on_click=lambda inst_id=str(inst['_id']): ui.navigate.to(f'/institution_admin/{inst_id}')
                            ).classes('btn-primary')
        else:
            ui.label('No institutions found. Create your first institution to get started.').style(
                'color: var(--text-secondary); font-style: italic; text-align: center; margin-top: 2rem;'
            )

# API endpoint for moving programs between schools
@app.post('/api/move_program')
async def move_program(request):
    try:
        from bson import ObjectId
        import json
        
        data = await request.json()
        program_id = data.get('program_id')
        new_school_id = data.get('new_school_id')
        institution_id = data.get('institution_id')
        
        if not all([program_id, new_school_id, institution_id]):
            return {'success': False, 'message': 'Missing required parameters'}
        
        # Get program and school details for logging
        program = programs_col.find_one({'_id': ObjectId(program_id)})
        if not program:
            return {'success': False, 'message': 'Program not found'}
            
        old_school = schools_col.find_one({'_id': ObjectId(program.get('school_id', ''))}) if program.get('school_id') else None
        new_school = schools_col.find_one({'_id': ObjectId(new_school_id)})
        
        if not new_school:
            return {'success': False, 'message': 'Target school not found'}
        
        # Update program's school_id
        result = programs_col.update_one(
            {'_id': ObjectId(program_id)},
            {'$set': {'school_id': new_school_id, 'updated_at': datetime.datetime.now(datetime.timezone.utc)}}
        )
        
        if result.modified_count == 0:
            return {'success': False, 'message': 'Failed to update program'}
        
        # Log the move action
        log_audit_action(
            action='Moved Program',
            details=f'Program "{program.get("name", "")}" moved from "{old_school.get("name", "") if old_school else "Unknown"}" to "{new_school.get("name", "")}"',
            institution_id=institution_id,
            entity_type='program',
            entity_id=program_id
        )
        
        return {'success': True, 'message': f'Program moved to {new_school.get("name", "")} successfully'}
        
    except Exception as e:
        print(f"Error moving program: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Failed to move program: {str(e)}'}

# Reusable Program Admin Sidebar
def program_admin_sidebar(program_id: str, active_page: str = 'dashboard'):
    """Reusable sidebar for program admin pages with improved styling"""
    from bson import ObjectId
    
    # Get program details
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    
    with ui.column().classes('fixed left-0 top-0 h-full w-64 bg-gradient-to-b from-gray-800 to-gray-900 text-white p-4 z-50'):
        # Logo and Program Info
        with ui.column().classes('items-center mb-8 pt-4'):
            ui.icon('school', size='2.5rem', color='white')
            if program:
                ui.label(program.get('name', 'Program')).classes('text-lg font-bold text-center text-white mt-2')
                ui.label('Program Admin').classes('text-sm text-gray-300')
            else:
                ui.label('Program Admin').classes('text-lg font-bold text-white')
        
        # Navigation Links
        nav_items = [
            {'name': 'Dashboard', 'icon': 'dashboard', 'page': 'dashboard'},
            {'name': 'Criteria', 'icon': 'checklist', 'page': 'criteria'},
            {'name': 'Extended Profiles', 'icon': 'groups', 'page': 'profiles'},
            {'name': 'Submissions', 'icon': 'send', 'page': 'submissions'},
        ]
        
        with ui.column().classes('w-full mt-4 space-y-1'):
            for item in nav_items:
                is_active = active_page == item['page']
                with ui.button('', 
                             on_click=lambda _, p=item['page']: ui.navigate.to(f'/program_admin/{program_id}/{p}')) \
                     .classes(f'w-full justify-start text-left py-3 px-4 rounded-lg transition-all duration-200 ' + 
                            f'{"bg-indigo-600 text-white shadow-md" if is_active else "text-gray-300 hover:bg-gray-700 hover:text-white"}') \
                     .props('flat'):
                    ui.icon(item['icon']).classes('mr-3')
                    ui.label(item['name']).classes('text-sm font-medium')
        
        # Logout and User Section
        with ui.column().classes('mt-auto mb-6'):
            ui.separator().classes('bg-gray-700 my-4')
            
            # User Info
            if 'current_user' in app.storage.user:
                user = app.storage.user['current_user']
                with ui.row().classes('items-center p-2 rounded-lg hover:bg-gray-700 cursor-pointer'):
                    ui.icon('account_circle', size='lg').classes('text-gray-300')
                    with ui.column().classes('ml-2'):
                        ui.label(f"{user.get('first_name', '')} {user.get('last_name', '')}") \
                            .classes('text-sm font-medium text-white')
                        ui.label('Program Admin').classes('text-xs text-gray-400')
            
            # Logout Button
            ui.button('Sign Out', 
                     on_click=lambda: ui.navigate.to('/logout'), 
                     icon='logout') \
                .classes('w-full mt-4 bg-gray-700 text-white hover:bg-gray-600 justify-start')

# Program Admin Pages
@ui.page('/program_admin/{program_id}')
def program_admin_redirect(program_id: str):
    """Redirect from /program_admin/{program_id} to /program_admin/{program_id}/dashboard"""
    ui.navigate.to(f'/program_admin/{program_id}/dashboard')
    return "Redirecting to dashboard..."

@ui.page('/program_admin/{program_id}/dashboard')
def program_admin_dashboard(program_id: str):
    """Program admin dashboard showing key metrics and overview"""
    from bson import ObjectId
    import datetime  # Using module directly for consistency
    
    # Initialize collections
    global criteria_submissions_col, extended_profile_submissions_col, audit_logs_col
    criteria_submissions_col = db['criteria_submissions']
    extended_profile_submissions_col = db['extended_profile_submissions']
    audit_logs_col = db['audit_logs']
    
    # Check authentication
    if not check_auth():
        ui.navigate.to('/')
        return
    
    # Get program data
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return
    
    # Get institution data
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    
    # Initialize collections if they don't exist
    if 'criteria_submissions_col' not in globals():
        criteria_submissions_col = db['criteria_submissions']
    if 'extended_profile_submissions_col' not in globals():
        extended_profile_submissions_col = db['extended_profile_submissions']
    if 'audit_logs_col' not in globals():
        audit_logs_col = db['audit_logs']
    
    # Get criteria and submissions count
    total_criteria = criterias_col.count_documents({'institution_id': str(program['institution_id'])})
    submitted_criteria = criteria_submissions_col.count_documents({
        'program_id': program_id,
        'status': 'submitted'
    })
    
    # Get extended profiles and submissions count
    total_profiles = extended_profiles_col.count_documents({'institution_id': str(program['institution_id'])})
    submitted_profiles = extended_profile_submissions_col.count_documents({
        'program_id': program_id,
        'status': 'submitted'
    })
    
    # Calculate completion percentages
    criteria_percent = int((submitted_criteria / total_criteria * 100) if total_criteria > 0 else 0)
    profiles_percent = int((submitted_profiles / total_profiles * 100) if total_profiles > 0 else 0)
    
    # Get recent activity
    week_ago = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=7)
    recent_activity = list(audit_logs_col.find({
        'program_id': program_id,
        'timestamp': {'$gte': week_ago}
    }).sort('timestamp', -1).limit(5))
    
    # Apply global styles
    add_beautiful_global_styles()
    
    with ui.column().classes('w-full min-h-screen bg-gray-50'):
        # Sidebar
        program_admin_sidebar(program_id, 'dashboard')
        
        # Main content area with left margin for sidebar
        with ui.column().classes('ml-64 p-8 bg-gray-50 min-h-screen'):
            # Header
            with ui.row().classes('w-full justify-between items-center mb-8'):
                with ui.column():
                    ui.label('Program Dashboard').classes('text-2xl font-bold text-gray-800')
                    breadcrumb = f"{institution.get('name', 'Institution')} > {program.get('name', 'Program')}"
                    ui.label(breadcrumb).classes('text-gray-500 text-sm')
                
                with ui.row().classes('items-center space-x-4'):
                    ui.button('New Submission', 
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                            icon='add').classes('bg-primary-500 text-white hover:bg-primary-600')
            
            # Stats Cards
            with ui.grid(columns=3).classes('w-full gap-6 mb-8'):
                # Criteria Completion
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('checklist', size='2rem', color='indigo-600').classes('p-2 bg-indigo-50 rounded-lg mb-3')
                        ui.label('Criteria Completion').classes('text-sm text-gray-500')
                        with ui.row().classes('items-center justify-between w-full mt-2'):
                            ui.label(f'{submitted_criteria} of {total_criteria}').classes('text-2xl font-bold text-gray-800')
                            ui.label(f'{criteria_percent}%').classes('text-sm font-medium text-indigo-600')
                        ui.linear_progress(value=criteria_percent/100, show_value=False).classes('w-full mt-2 h-2')
                
                # Profiles Completion
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('person', size='2rem', color='teal-600').classes('p-2 bg-teal-50 rounded-lg mb-3')
                        ui.label('Profiles Completion').classes('text-sm text-gray-500')
                        with ui.row().classes('items-center justify-between w-full mt-2'):
                            ui.label(f'{submitted_profiles} of {total_profiles}').classes('text-2xl font-bold text-gray-800')
                            ui.label(f'{profiles_percent}%').classes('text-sm font-medium text-teal-600')
                        ui.linear_progress(value=profiles_percent/100, show_value=False).classes('w-full mt-2 h-2')
                
                # Recent Activity
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('update', size='2rem', color='purple-600').classes('p-2 bg-purple-50 rounded-lg mb-3')
                        ui.label('Recent Activity').classes('text-sm text-gray-500')
                        if recent_activity:
                            for activity in recent_activity:
                                with ui.row().classes('w-full py-2 border-b border-gray-100 last:border-0'):
                                    ui.icon('circle', size='0.5rem', color='gray-400').classes('mt-1.5 mr-2')
                                    with ui.column():
                                        ui.label(activity.get('action', 'Activity')).classes('text-sm font-medium text-gray-800')
                                        timestamp = activity.get('timestamp', datetime.now(datetime.UTC))
                                        if hasattr(timestamp, 'strftime'):
                                            timestamp_str = timestamp.strftime('%b %d, %I:%M %p')
                                        else:
                                            timestamp_str = 'Just now'
                                        ui.label(f"{activity.get('user_name', 'System')} • {timestamp_str}").classes('text-xs text-gray-500')
                        else:
                            ui.label('No recent activity').classes('text-sm text-gray-400 mt-2')
            
            # Quick Actions
            with ui.card().classes('w-full mb-8'):
                with ui.column().classes('w-full p-4'):
                    ui.label('Quick Actions').classes('text-lg font-semibold text-gray-700 mb-4')
                    with ui.row().classes('w-full space-x-4'):
                        ui.button('Submit Criteria', 
                                on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                                icon='checklist').props('flat color=primary')
                        ui.button('Submit Profile', 
                                on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/profiles'),
                                icon='person').props('flat color=teal')
                        ui.button('View Submissions', 
                                on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/submissions'),
                                icon='list_alt').props('flat color=purple')
                
                # Quick Actions Section
                with ui.column().classes('w-80'):
                    with ui.card().classes('w-full border border-gray-100'):
                        with ui.column().classes('w-full'):
                            ui.label('Quick Actions').classes('text-lg font-semibold p-4 border-b')
                            
                            # Quick Action Buttons
                            quick_actions = [
                                {'icon': 'add_circle', 'label': 'Add New Criteria', 'color': 'primary', 'route': f'/program_admin/{program_id}/criteria/new'},
                                {'icon': 'group_add', 'label': 'Add Profile', 'color': 'teal', 'route': f'/program_admin/{program_id}/profiles/new'},
                                {'icon': 'upload_file', 'label': 'Import Data', 'color': 'indigo', 'route': f'/program_admin/{program_id}/import'}
                            ]
                            
                            for action in quick_actions:
                                with ui.button('', on_click=lambda _, r=action['route']: ui.navigate.to(r)) \
                                     .classes(f'w-full justify-start mb-2 text-{action["color"]}-600 hover:bg-{action["color"]}-50') \
                                     .props('flat'):
                                    ui.icon(action['icon']).classes(f'mr-3 text-{action["color"]}-500')
                                    ui.label(action['label']).classes('text-sm font-medium')
                    
                    # Help Card
                    with ui.card().classes('w-full border border-gray-100 mt-4'):
                        with ui.column().classes('w-full'):
                            ui.label('Need Help?').classes('text-lg font-semibold p-4 border-b')
                            ui.label('Having trouble or have questions? Our support team is here to help.').classes('text-sm text-gray-600 px-4 pb-4')
                            ui.button('Contact Support', icon='support_agent', 
                                     on_click=lambda: ui.navigate.to('/support')) \
                                .classes('w-full mt-2 bg-gray-100 text-gray-700 hover:bg-gray-200')

@ui.page('/program_admin/{program_id}/criteria')
def program_admin_criteria(program_id: str):
    """Program admin page to view and fill criteria"""
    from bson import ObjectId
    
    # Check authentication
    if not check_auth():
        ui.navigate.to('/')
        return

    # Get program data
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return

    # Get institution data
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    
    # Get all criteria for the institution with submission status
    criteria_list = list(criterias_col.find({
        'institution_id': str(program['institution_id'])
    }))
    
    # Get submission status for each criteria
    for criteria in criteria_list:
        submission = criteria_submissions_col.find_one({
            'criteria_id': str(criteria['_id']),
            'program_id': program_id,
            'status': 'submitted'
        }) if 'criteria_submissions_col' in globals() else None
        criteria['submission'] = submission
    
    # Apply global styles
    add_beautiful_global_styles()
    
    with ui.column().classes('w-full min-h-screen bg-gray-50'):
        # Sidebar
        program_admin_sidebar(program_id, 'criteria')
        
        # Main content area with left margin for sidebar
        with ui.column().classes('ml-64 p-8 bg-gray-50 min-h-screen'):
            # Header
            with ui.row().classes('w-full justify-between items-center mb-8'):
                with ui.column():
                    ui.label('Criteria Management').classes('text-2xl font-bold text-gray-800')
                    ui.label(f'Manage and submit criteria for {program.get("name", "Program")}') \
                        .classes('text-gray-500 text-sm')
                
                with ui.button('Add New Criteria', 
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/new'),
                            icon='add') \
                        .classes('bg-primary-500 text-white hover:bg-primary-600'):
                    pass
            
            # Stats Cards
            with ui.grid(columns=3).classes('w-full gap-6 mb-8'):
                # Total Criterias
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('checklist', size='2rem', color='indigo-600').classes('p-2 bg-indigo-50 rounded-lg mb-3')
                        ui.label('Total Criterias').classes('text-sm text-gray-500')
                        ui.label(str(len(criteria_list))).classes('text-2xl font-bold text-gray-800 mt-1')
                
                # Filled Criterias
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('task_alt', size='2rem', color='green-600').classes('p-2 bg-green-50 rounded-lg mb-3')
                        ui.label('Completed').classes('text-sm text-gray-500')
                        filled = sum(1 for c in criteria_list if c.get('submission'))
                        ui.label(f'{filled} of {len(criteria_list)}').classes('text-2xl font-bold text-gray-800 mt-1')
                
                # Completion Rate
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('trending_up', size='2rem', color='purple-600').classes('p-2 bg-purple-50 rounded-lg mb-3')
                        ui.label('Completion Rate').classes('text-sm text-gray-500')
                        rate = int((filled / len(criteria_list) * 100) if criteria_list else 0)
                        ui.label(f'{rate}%').classes('text-2xl font-bold text-gray-800 mt-1')
            
            # Criteria List
            if not criteria_list:
                with ui.card().classes('w-full p-8 text-center'):
                    ui.icon('info', size='3rem', color='gray-300').classes('mx-auto mb-4')
                    ui.label('No criteria available').classes('text-xl font-medium text-gray-700 mb-2')
                    ui.label('Please contact your institution admin to add criteria for your program.').classes('text-gray-500')
            else:
                with ui.card().classes('w-full border border-gray-100'):
                    with ui.column().classes('w-full'):
                        # Table Header
                        with ui.row().classes('w-full bg-gray-50 p-4 border-b'):
                            ui.label('Criteria').classes('flex-1 font-medium text-gray-600')
                            ui.label('Status').classes('w-32 font-medium text-gray-600 text-center')
                            ui.label('Last Updated').classes('w-48 font-medium text-gray-600')
                            ui.label('Actions').classes('w-32 font-medium text-gray-600')
                        
                        # Criteria Rows
                        for criteria in criteria_list:
                            submission = criteria.get('submission')
                            status = 'Submitted' if submission else 'Pending'
                            status_color = 'green' if submission else 'gray'
                            last_updated = submission.get('updated_at', '').strftime('%b %d, %Y') if submission else '--'
                            
                            with ui.row().classes('w-full p-4 border-b hover:bg-gray-50 items-center'):
                                # Criteria Name and Description
                                with ui.column().classes('flex-1'):
                                    ui.label(criteria.get('name', 'Unnamed Criteria')).classes('font-medium text-gray-800')
                                    if criteria.get('description'):
                                        ui.label(criteria.get('description')).classes('text-sm text-gray-500 mt-1')
                                
                                # Status
                                with ui.column().classes('w-32'):
                                    ui.label(status).classes(f'px-2 py-1 rounded-full text-xs font-medium text-{status_color}-800 bg-{status_color}-100 text-center')
                                
                                # Last Updated
                                ui.label(last_updated).classes('w-48 text-sm text-gray-600')
                                
                                # Actions
                                with ui.row().classes('w-32 space-x-2'):
                                    if submission:
                                        ui.button('View', on_click=lambda c=criteria: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{c["_id"]}')) \
                                            .props('flat color=primary size=sm').classes('text-xs')
                                        ui.button('Edit', on_click=lambda c=criteria: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{c["_id"]}')) \
                                            .props('outline color=primary size=sm').classes('text-xs')
                                    else:
                                        ui.button('Fill', on_click=lambda c=criteria: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{c["_id"]}')) \
                                            .props('flat color=primary size=sm').classes('text-xs')

@ui.page('/program_admin/{program_id}/profiles')
def program_admin_profiles(program_id: str):
    """Program admin page to view and fill extended profiles with improved UI"""
    from bson import ObjectId
    from datetime import datetime
    
    # Check authentication
    if not check_auth():
        ui.navigate.to('/')
        return

    # Get program data
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return

    # Get institution data
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    
    # Get all extended profiles for the institution with submission status
    profiles_list = list(extended_profiles_col.find({
        'institution_id': str(program['institution_id'])
    }))
    
    # Get submission status for each profile
    for profile in profiles_list:
        submission = extended_profile_submissions_col.find_one({
            'profile_id': str(profile['_id']),
            'program_id': program_id,
            'status': 'submitted'
        }) if 'extended_profile_submissions_col' in globals() else None
        profile['submission'] = submission
    
    # Apply global styles
    add_beautiful_global_styles()
    
    with ui.column().classes('w-full min-h-screen bg-gray-50'):
        # Sidebar
        program_admin_sidebar(program_id, 'profiles')
        
        # Main content area with left margin for sidebar
        with ui.column().classes('ml-64 p-8 bg-gray-50 min-h-screen'):
            # Header
            with ui.row().classes('w-full justify-between items-center mb-8'):
                with ui.column():
                    ui.label('Extended Profiles').classes('text-2xl font-bold text-gray-800')
                    ui.label(f'Manage extended profiles for {program.get("name", "Program")}') \
                        .classes('text-gray-500 text-sm')
                
                with ui.button('Add New Profile', 
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/profiles/new'),
                            icon='person_add') \
                        .classes('bg-teal-500 text-white hover:bg-teal-600'):
                    pass
            
            # Stats Cards
            with ui.grid(columns=3).classes('w-full gap-6 mb-8'):
                # Total Profiles
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('groups', size='2rem', color='teal-600').classes('p-2 bg-teal-50 rounded-lg mb-3')
                        ui.label('Total Profiles').classes('text-sm text-gray-500')
                        ui.label(str(len(profiles_list))).classes('text-2xl font-bold text-gray-800 mt-1')
                
                # Filled Profiles
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('assignment_turned_in', size='2rem', color='green-600').classes('p-2 bg-green-50 rounded-lg mb-3')
                        ui.label('Completed').classes('text-sm text-gray-500')
                        filled = sum(1 for p in profiles_list if p.get('submission'))
                        ui.label(f'{filled} of {len(profiles_list)}').classes('text-2xl font-bold text-gray-800 mt-1')
                
                # Completion Rate
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('trending_up', size='2rem', color='purple-600').classes('p-2 bg-purple-50 rounded-lg mb-3')
                        ui.label('Completion Rate').classes('text-sm text-gray-500')
                        rate = int((filled / len(profiles_list) * 100) if profiles_list else 0)
                        ui.label(f'{rate}%').classes('text-2xl font-bold text-gray-800 mt-1')
            
            # Profiles List
            if not profiles_list:
                with ui.card().classes('w-full p-8 text-center'):
                    ui.icon('person_off', size='3rem', color='gray-300').classes('mx-auto mb-4')
                    ui.label('No profiles available').classes('text-xl font-medium text-gray-700 mb-2')
                    ui.label('Please contact your institution admin to add extended profiles for your program.').classes('text-gray-500')
            else:
                with ui.card().classes('w-full border border-gray-100'):
                    with ui.column().classes('w-full'):
                        # Table Header
                        with ui.row().classes('w-full bg-gray-50 p-4 border-b'):
                            ui.label('Profile').classes('flex-1 font-medium text-gray-600')
                            ui.label('Status').classes('w-32 font-medium text-gray-600 text-center')
                            ui.label('Last Updated').classes('w-48 font-medium text-gray-600')
                            ui.label('Actions').classes('w-32 font-medium text-gray-600')
                        
                        # Profile Rows
                        for profile in profiles_list:
                            submission = profile.get('submission')
                            status = 'Submitted' if submission else 'Pending'
                            status_color = 'green' if submission else 'gray'
                            last_updated = submission.get('updated_at', '').strftime('%b %d, %Y') if submission else '--'
                            with ui.row().classes('w-full p-4 border-b hover:bg-gray-50 items-center'):
                                # Profile Name and Description
                                with ui.column().classes('flex-1'):
                                    ui.label(profile.get('name', 'Unnamed Profile')).classes('font-medium text-gray-800')
                                    if profile.get('description'):
                                        ui.label(profile.get('description')).classes('text-sm text-gray-500 mt-1')
                                # Status
                                with ui.column().classes('w-32'):
                                    ui.label(status).classes(f'px-2 py-1 rounded-full text-xs font-medium text-{status_color}-800 bg-{status_color}-100 text-center')
                                # Last Updated
                                ui.label(last_updated).classes('w-48 text-sm text-gray-600')
                                # Actions
                                with ui.row().classes('w-32 space-x-2'):
                                    if submission:
                                        ui.button('View', on_click=lambda p=profile: ui.navigate.to(f'/program_admin/{program_id}/profiles/{p["_id"]}')) \
                                            .props('flat color=teal size=sm').classes('text-xs')
                                        ui.button('Edit', on_click=lambda p=profile: ui.navigate.to(f'/program_admin/{program_id}/profiles/{p["_id"]}')) \
                                            .props('outline color=teal size=sm').classes('text-xs')
                                    else:
                                        ui.button('Fill', on_click=lambda p=profile: ui.navigate.to(f'/program_admin/{program_id}/profiles/{p["_id"]}')) \
                                            .props('flat color=teal size=sm').classes('text-xs')

def render_editable_draft(draft_data, program_id, criteria_id):
    """Render an editable table for draft data with improved UI/UX"""
    if not draft_data or 'data' not in draft_data or not draft_data['data']:
        ui.notify('No data available in this draft', color='warning')
        return
    
    import pandas as pd
    from copy import deepcopy
    
    # Create a deep copy of the data to avoid modifying the original
    data = deepcopy(draft_data['data'])
    df = pd.DataFrame(data)
    
    # Create a container for the table and buttons
    with ui.column().classes('w-full bg-white p-6 rounded-lg shadow'):
        # Header with draft info and action buttons
        with ui.row().classes('w-full justify-between items-start mb-6'):
            with ui.column():
                ui.label(f'Draft: {draft_data.get("name", "Unnamed Draft")}').classes('text-xl font-semibold')
                if 'updated_at' in draft_data and draft_data['updated_at']:
                    last_updated = draft_data['updated_at'].strftime('%b %d, %Y %I:%M %p')
                    ui.label(f'Last updated: {last_updated}').classes('text-sm text-gray-500')
            
            # Action buttons
            with ui.row().classes('gap-3'):
                save_btn = ui.button('Save Draft', icon='save', color='primary').classes('min-w-[120px]')
                submit_btn = ui.button('Submit', icon='send', color='positive').classes('min-w-[120px]')
        
        # Table container
        table_container = ui.column().classes('w-full overflow-x-auto')
        
        # Function to save draft
        async def save_draft():
            try:
                # Show loading state
                with ui.dialog() as dialog, ui.card():
                    ui.spinner('dots')
                    ui.label('Saving draft...')
                dialog.open()
                
                # Get updated data from the table
                updated_data = df.replace({pd.NA: None}).to_dict('records')
                
                # Update the draft in the database
                result = criteria_submissions_col.update_one(
                    {'_id': draft_data['_id']},
                    {'$set': {
                        'data': updated_data,
                        'updated_at': datetime.utcnow()
                    }}
                )
                
                dialog.close()
                if result.modified_count > 0:
                    ui.notify('Draft saved successfully', color='positive')
                    # Update the UI to show the last updated time
                    ui.navigate.to(f'/program_admin/{program_id}/submissions')
                else:
                    ui.notify('No changes to save', color='info')
                    
            except Exception as e:
                dialog.close()
                ui.notify(f'Error saving draft: {str(e)}', color='negative')
        
        # Function to submit the draft
        async def submit_draft():
            try:
                # Show confirmation dialog
                with ui.dialog() as dialog, ui.card():
                    ui.label('Are you sure you want to submit this draft?')
                    with ui.row():
                        ui.button('Cancel', on_click=dialog.close)
                        ui.button('Submit', on_click=lambda: dialog.submit('submit'), color='positive')
                
                result = await dialog
                if result != 'submit':
                    return
                
                # Show loading state
                with ui.dialog() as loading_dialog, ui.card():
                    ui.spinner('dots')
                    ui.label('Submitting...')
                loading_dialog.open()
                
                # Update status to submitted
                result = criteria_submissions_col.update_one(
                    {'_id': draft_data['_id']},
                    {'$set': {
                        'status': 'submitted',
                        'submitted_at': datetime.utcnow(),
                        'updated_at': datetime.utcnow()
                    }}
                )
                
                loading_dialog.close()
                if result.modified_count > 0:
                    ui.notify('Submission successful!', color='positive')
                    # Refresh the page to show updated status
                    ui.navigate.to(f'/program_admin/{program_id}/submissions')
                else:
                    ui.notify('No changes to submit', color='info')
                    
            except Exception as e:
                if 'loading_dialog' in locals():
                    loading_dialog.close()
                ui.notify(f'Error submitting: {str(e)}', color='negative')
        
        # Set up button click handlers with async/await
        save_btn.on('click', save_draft)
        submit_btn.on('click', submit_draft)
        
        # Function to render the table
        def render_table():
            table_container.clear()
            with table_container:
                # Create a card with shadow and rounded corners for the table
                with ui.card().classes('w-full overflow-hidden border'):
                    # Table header with sticky positioning
                    with ui.element('thead').classes('bg-gray-50 sticky top-0 z-10'):
                        with ui.row().classes('w-full bg-gray-50 border-b'):
                            for col in df.columns:
                                ui.label(str(col)).classes('p-3 text-sm font-medium text-gray-700')
                            ui.label('Actions').classes('p-3 text-sm font-medium text-gray-700')
                    
                    # Table body
                    with ui.element('div').classes('divide-y divide-gray-200'):
                        for idx, row in df.iterrows():
                            with ui.row().classes('w-full hover:bg-gray-50 transition-colors'):
                                for col in df.columns:
                                    # Create editable input for each cell
                                    cell_value = '' if pd.isna(row[col]) else str(row[col])
                                    with ui.input(value=cell_value).classes('flex-1 border-0 focus:ring-1 focus:ring-primary-500 m-1'):
                                        def update_cell(value, r=idx, c=col):
                                            df.at[r, c] = value if value != '' else None
                                        ui.on('change', lambda e, c=col, r=idx: update_cell(e.args, r, c))
                                
                                # Delete button for the row
                                with ui.button(icon='delete', color='red').classes('m-1').props('flat size=sm'):
                                    # Use a list to store the row index and modify it in the callback
                                    row_idx = [idx]
                                    def delete_row():
                                        # Get the current index of the row to delete
                                        current_idx = row_idx[0]
                                        if current_idx in df.index:
                                            df.drop(index=current_idx, inplace=True)
                                            df.reset_index(drop=True, inplace=True)
                                            render_table()
                                    ui.on('click', delete_row)
        
        # Function to add a new row
        def add_row():
            new_row = {col: None for col in df.columns}
            df.loc[len(df)] = new_row
            render_table()
        
        # Initial table render
        render_table()
        
        # Add row button
        with ui.row().classes('w-full justify-end mt-4'):
            ui.button('Add Row', icon='add', on_click=add_row).classes('bg-green-500 text-white hover:bg-green-600')

@ui.page('/program_admin/{program_id}/submissions')
def program_admin_submissions_page(program_id: str):
    """Program admin page to view submissions and drafts with improved UI"""
    from bson import ObjectId
    from datetime import datetime, timedelta
    
    # Get current user from session
    user = app.storage.user.get('current_user', {}) if hasattr(app.storage, 'user') else {}
    
    # Check authentication
    if not check_auth():
        ui.navigate.to('/')
        return
    
    # Get program data
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return
    
    # Get institution data
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    
    # Get all submissions for this program
    criteria_submissions = []
    profile_submissions = []
    
    try:
        # Convert program_id to ObjectId if it's a string
        program_id_obj = ObjectId(program_id) if isinstance(program_id, str) else program_id
        
        # Get criteria submissions including drafts
        criteria_submissions = list(criteria_submissions_col.find({
            '$or': [
                {'program_id': program_id},
                {'program_id': program_id_obj}
            ],
            'status': {'$in': ['draft', 'submitted']}
        }).sort('updated_at', -1))
        
        # Get profile submissions including drafts
        profile_submissions = list(db['extended_profile_submissions'].find({
            '$or': [
                {'program_id': program_id},
                {'program_id': program_id_obj}
            ],
            'status': {'$in': ['draft', 'submitted']}
        }).sort('updated_at', -1))
        
    except Exception as e:
        print(f"Error fetching submissions: {str(e)}")
        ui.notify(f'Error loading submissions: {str(e)}', color='negative')
    
    # Apply global styles
    add_beautiful_global_styles()
    
    with ui.column().classes('w-full min-h-screen bg-gray-50'):
        # Sidebar
        program_admin_sidebar(program_id, 'submissions')
        
        # Main content area with left margin for sidebar
        with ui.column().classes('ml-64 p-8 bg-gray-50 min-h-screen'):
            # Header
            with ui.row().classes('w-full justify-between items-center mb-8'):
                with ui.column():
                    ui.label('My Submissions').classes('text-2xl font-bold text-gray-800')
                    ui.label(f'View and manage your submitted forms for {program.get("name", "Program")}') \
                        .classes('text-gray-500 text-sm')
                
                with ui.button('New Submission', 
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                            icon='add') \
                        .classes('bg-primary-500 text-white hover:bg-primary-600'):
                    pass
            
            # Stats Cards
            with ui.grid(columns=3).classes('w-full gap-6 mb-8'):
                # Total Submissions
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('send', size='2rem', color='indigo-600').classes('p-2 bg-indigo-50 rounded-lg mb-3')
                        ui.label('Total Submissions').classes('text-sm font-medium text-gray-500')
                        ui.label(str(len([s for s in criteria_submissions if s.get('status') == 'submitted']))).classes('text-2xl font-bold text-gray-800')
                
                # Drafts
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('drafts', size='2rem', color='yellow-600').classes('p-2 bg-yellow-50 rounded-lg mb-3')
                        ui.label('Drafts').classes('text-sm font-medium text-gray-500')
                        ui.label(str(len([s for s in criteria_submissions if s.get('status') == 'draft']))).classes('text-2xl font-bold text-gray-800')
                
                # Recent Activity
                with ui.card().classes('shadow-sm hover:shadow-md transition-all border border-gray-100'):
                    with ui.column().classes('p-5'):
                        ui.icon('update', size='2rem', color='green-600').classes('p-2 bg-green-50 rounded-lg mb-3')
                        ui.label('Last Updated').classes('text-sm font-medium text-gray-500')
                        last_updated = max(
                            [s.get('updated_at', datetime.min) for s in criteria_submissions + profile_submissions],
                            default=datetime.utcnow()
                        )
                        ui.label(last_updated.strftime('%b %d, %Y %I:%M %p') if last_updated != datetime.min else 'Never').classes('text-sm font-medium text-gray-800')
            
            # Drafts Section
            with ui.card().classes('w-full mb-8'):
                with ui.column().classes('w-full'):
                    ui.label('Drafts').classes('text-xl font-semibold mb-4')
                    
                    all_drafts = [s for s in criteria_submissions + profile_submissions if s.get('status') == 'draft']
                    
                    if not all_drafts:
                        ui.label('No drafts found').classes('text-gray-500 italic')
                    else:
                        for draft in all_drafts:
                            draft_type = 'Criteria' if 'criteria_id' in draft else 'Extended Profile'
                            draft_name = draft.get('name', 'Unnamed')
                            updated_time = draft.get('updated_at', '').strftime('%b %d, %Y %I:%M %p') if draft.get('updated_at') else 'Unknown'
                            
                            with ui.expansion(f"Draft: {draft_name} ({draft_type}) - {updated_time}", 
                                           icon='draft').classes('w-full mb-2'):
                                with ui.column().classes('w-full p-4'):
                                    # Show draft data
                                    if 'data' in draft and 'table_data' in draft['data']:
                                        table_data = draft['data']['table_data']
                                        if table_data and len(table_data) > 0:
                                            # Create a simple table to show the data
                                            headers = list(table_data[0].keys()) if table_data else []
                                            with ui.table(columns=headers, rows=table_data, row_key='id').classes('w-full mb-4'):
                                                pass
                                            
                                            # Show data summary
                                            ui.label(f'Data rows: {len(table_data)}').classes('text-sm text-gray-600 mb-2')
                                        else:
                                            ui.label('No data in this draft').classes('text-gray-500 italic')
                                    else:
                                        ui.label('No data in this draft').classes('text-gray-500 italic')
                                    
                                    # Action buttons for drafts
                                    with ui.row().classes('w-full gap-2 mt-4'):
                                        if draft_type == 'Criteria':
                                            ui.button('Edit Draft', 
                                                    on_click=lambda d=draft: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{d.get("criteria_id")}'),
                                                    icon='edit',
                                                    color='primary')
                                        else:
                                            ui.button('Edit Draft', 
                                                    on_click=lambda d=draft: ui.navigate.to(f'/program_admin/{program_id}/profiles/{d.get("profile_id")}'),
                                                    icon='edit',
                                                    color='primary')
                                        
                                        ui.button('Submit Draft', 
                                                on_click=lambda d=draft: submit_draft(d, program_id),
                                                icon='send',
                                                color='green')
                                        
                                        ui.button('Delete Draft', 
                                                on_click=lambda d=draft: delete_draft(d, program_id),
                                                icon='delete',
                                                color='red')
            
            # Submissions Section
            with ui.card().classes('w-full'):
                with ui.column().classes('w-full'):
                    ui.label('Submissions').classes('text-xl font-semibold mb-4')
                    
                    all_submissions = [s for s in criteria_submissions + profile_submissions if s.get('status') == 'submitted']
                    
                    if not all_submissions:
                        ui.label('No submissions found').classes('text-gray-500 italic')
                    else:
                        # Create submissions data
                        submissions_data = []
                        for submission in all_submissions:
                            submission_type = 'Criteria' if 'criteria_id' in submission else 'Extended Profile'
                            submission_name = submission.get('name', 'Unnamed')
                            submitted_time = submission.get('submitted_at', '').strftime('%b %d, %Y %I:%M %p') if submission.get('submitted_at') else 'N/A'
                            
                            # Create row data
                            row_data = {
                                'name': submission_name,
                                'type': f'{submission_type} Submission',
                                'status': 'Submitted',
                                'submitted_at': submitted_time,
                                'actions': 'View | Edit | Download'
                            }
                            submissions_data.append(row_data)
                        
                        # Display submissions in a simple table
                        with ui.table(columns=[
                            {'name': 'name', 'label': 'Name', 'field': 'name'},
                            {'name': 'type', 'label': 'Type', 'field': 'type'},
                            {'name': 'status', 'label': 'Status', 'field': 'status'},
                            {'name': 'submitted_at', 'label': 'Submitted At', 'field': 'submitted_at'},
                            {'name': 'actions', 'label': 'Actions', 'field': 'actions'}
                        ], rows=submissions_data).classes('w-full'):
                            pass
                        
                        # Add action buttons below the table
                        for i, submission in enumerate(all_submissions):
                            with ui.row().classes('w-full gap-2 mt-2 p-2 bg-gray-50 rounded'):
                                ui.label(f"{i+1}. {submission.get('name', 'Unnamed')}").classes('flex-1')
                                
                                ui.button('View', 
                                        on_click=lambda s=submission: view_submission(s, program_id),
                                        icon='visibility',
                                        color='primary',
                                        size='sm')
                                
                                ui.button('Edit', 
                                        on_click=lambda s=submission: edit_submitted_submission(s, program_id),
                                        icon='edit',
                                        color='warning',
                                        size='sm')
                                
                                ui.button('Download', 
                                        on_click=lambda s=submission: download_submission(s),
                                        icon='download',
                                        color='teal',
                                        size='sm')
            
            # Empty state for no submissions
            if not any(s.get('status') == 'submitted' for s in criteria_submissions) and \
               not any(s.get('status') == 'draft' for s in criteria_submissions):
                with ui.card().classes('w-full p-8 text-center'):
                    ui.icon('inbox', size='3rem', color='gray-300').classes('mx-auto mb-4')
                    ui.label('No submissions yet').classes('text-xl font-medium text-gray-700 mb-2')
                    ui.label('Get started by creating a new submission').classes('text-gray-500')
                    ui.button('New Submission', 
                             on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                                     color='primary').classes('mt-4')
            
            # Add New Profile Button (if admin)
            if user.get('role') == 'admin':
                with ui.row().classes('mt-4'):
                    ui.button('Add New Profile', on_click=lambda: ui.navigate.to(f'/institutions/{institution["_id"]}/profiles/new')) \
                        .props('flat color=teal icon=person_add')
        # Department Stats Section
        with ui.card().classes('mt-6').style('width: 100%;'):
            ui.label('📊 Department Statistics').style(
                    f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;'
                )
                
            # Get department profiles for the program
            dept_profiles = list(extended_profiles_col.find({
                'institution_id': str(program['institution_id']),
                'department_id': program.get('department_id')
            })) if program.get('department_id') else []
            
            # Get department criterias for the program
            dept_criterias = list(criterias_col.find({
                'institution_id': str(program['institution_id']),
                'scope_type': 'department_based',
                'department_id': program.get('department_id')
            }))
            
            # Get total number of criteria for progress calculation
            total_criteria = len(dept_criterias) if dept_criterias else 0
            total_profiles = len(dept_profiles) if dept_profiles else 0
            
            submitted_criteria = 0
            in_progress_criteria = 0
            
            submitted_profiles = 0
            in_progress_profiles = 0
            
            # Count criteria statuses
            if dept_criterias:
                for criteria in dept_criterias:
                    submission = criteria_submissions_col.find_one({
                        'criteria_id': str(criteria['_id']),
                        'department_id': department_id,
                        'program_id': program_id
                    })
                    if submission:
                        if submission.get('status') == 'submitted':
                            submitted_criteria += 1
                        elif submission.get('status') in ['in_progress', 'draft']:
                            in_progress_criteria += 1
            
            # Count profile statuses
            if dept_profiles:
                for profile in dept_profiles:
                    submission = extended_profile_submissions_col.find_one({
                        'profile_id': str(profile['_id']),
                        'department_id': department_id,
                        'program_id': program_id
                    })
                    
                    if submission:
                        if submission.get('status') == 'submitted':
                            submitted_profiles += 1
                        elif submission.get('status') in ['in_progress', 'draft']:
                            in_progress_profiles += 1
            
            # Display stats in a grid
            with ui.grid(columns=4).classes('w-full gap-4'):
                # Criteria Stats
                with ui.card().classes('p-4 text-center'):
                    ui.label('📋 Total Criteria').classes('text-gray-600 text-sm mb-2')
                    ui.label(str(total_criteria)).classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('✅ Submitted').classes('text-green-600 text-sm mb-2')
                    ui.label(f"{submitted_criteria} of {total_criteria}").classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('⏳ In Progress').classes('text-blue-600 text-sm mb-2')
                    ui.label(str(in_progress_criteria)).classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('📊 Completion').classes('text-purple-600 text-sm mb-2')
                    completion = int((submitted_criteria / total_criteria) * 100) if total_criteria > 0 else 0
                    ui.label(f"{completion}%").classes('text-2xl font-bold')
                    
                    # Progress bar
                    with ui.linear_progress(value=completion/100, show_value=False).classes('w-full h-2 mt-2') as progress:
                        progress.style(f'--q-linear-progress-track-color: #e9ecef; --q-linear-progress-color: {main_color};')
            
            # Add a separator
            ui.separator().classes('my-6')
            
            # Profiles Stats
            with ui.grid(columns=4).classes('w-full gap-4'):
                with ui.card().classes('p-4 text-center'):
                    ui.label('👥 Total Profiles').classes('text-gray-600 text-sm mb-2')
                    ui.label(str(total_profiles)).classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('✅ Submitted').classes('text-green-600 text-sm mb-2')
                    ui.label(f"{submitted_profiles} of {total_profiles}").classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('⏳ In Progress').classes('text-blue-600 text-sm mb-2')
                    ui.label(str(in_progress_profiles)).classes('text-2xl font-bold')
                
                with ui.card().classes('p-4 text-center'):
                    ui.label('📊 Completion').classes('text-purple-600 text-sm mb-2')
                    completion = int((submitted_profiles / total_profiles) * 100) if total_profiles > 0 else 0
                    ui.label(f"{completion}%").classes('text-2xl font-bold')
                    
                    # Progress bar
                    with ui.linear_progress(value=completion/100, show_value=False).classes('w-full h-2 mt-2') as progress:
                        progress.style(f'--q-linear-progress-track-color: #e9ecef; --q-linear-progress-color: {main_color};')
            
            # Extended Profiles Section
            with ui.card().classes('beautiful-card').style('flex: 1; padding: 2rem;'):
                ui.label('👤 Extended Profiles').style(
                    f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;'
                )
                
                # Get extended profiles for this program's department
                dept_profiles = list(extended_profiles_col.find({
                    'institution_id': str(program['institution_id']),
                    'scope_type': 'department_based',
                    'department_id': program.get('department_id')
                }))
                
                if dept_profiles:
                    for profile in dept_profiles:
                        with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                            ui.label(f"📄 {profile.get('name', 'Unnamed')}").style(
                                'font-size: 1.1rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                            )
                            
                            # Fix deadline display for profiles too
                            deadline = profile.get('deadline')
                            if deadline:
                                try:
                                    if isinstance(deadline, str):
                                        deadline_str = deadline
                                    else:
                                        deadline_str = deadline.strftime('%Y-%m-%d')
                                    ui.label(f"Deadline: {deadline_str}").style(
                                        'font-size: 0.9rem; color: var(--warning-color); margin-bottom: 0.5rem;'
                                    )
                                except:
                                    ui.label("Deadline: Invalid date format").style(
                                        'font-size: 0.9rem; color: var(--error-color); margin-bottom: 0.5rem;'
                                    )
                            else:
                                ui.label("Deadline: No deadline set").style(
                                    'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                                )
                            
                            # Fill profile button
                            ui.button('📝 Fill Profile', on_click=lambda p_id=str(profile['_id']): ui.navigate.to(f'/department_admin/{department_id}/fill_profile/{p_id}')).style(
                                f'background: {main_color}; color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                            )
                else:
                    ui.label('No extended profiles available for this department.').style(
                        'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
                    )
    
    # Create Institution Admin-style sidebar for department admin
    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
        # Simple sidebar
        with ui.column().style('min-width: 250px; background: var(--surface); border-right: 2px solid var(--border); height: 100vh; padding: 1.5rem;'):
            ui.label(f'🏢 {program.get("name", "Program")}').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 0.5rem;'
            )
            ui.label('Program Admin').style('font-size: 1rem; color: var(--text-secondary); margin-bottom: 2rem;')
            
            ui.separator().style('margin: 1rem 0; background: var(--border);')
            
            ui.button('📊 Criteria', on_click=lambda: ui.navigate.to(f'/department_admin/{department_id}')).style(
                f'width: 100%; justify-content: flex-start; background: {main_color}; color: white; padding: 0.75rem 1rem; border-radius: 8px; border: none; margin-bottom: 0.5rem; font-weight: 500;'
            )
            
            ui.button('👤 Extended Profiles', on_click=lambda: ui.navigate.to(f'/department_admin/{department_id}')).style(
                f'width: 100%; justify-content: flex-start; background: {main_color}; color: white; padding: 0.75rem 1rem; border-radius: 8px; border: none; margin-bottom: 0.5rem; font-weight: 500;'
            )
            
            ui.separator().style('margin: 1rem 0; background: var(--border);')
            
            def logout():
                global current_user
                current_user = None
                if hasattr(app.storage, 'user'):
                    app.storage.user.clear()
                ui.notify('Logged out successfully', color='positive')
                ui.navigate.to('/')
            
            ui.button('🚪 Logout', on_click=logout).style(
                f'width: 100%; margin-top: auto; background: linear-gradient(135deg, rgb(154, 44, 84) 0%, rgb(124, 35, 67) 100%); color: white; padding: 1rem; border-radius: 10px; border: none; font-weight: 600;'
            )
        
        # Main content
        with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
            # Main content will be rendered here by other UI components
            pass

# Profile View Page
@ui.page('/program_admin/{program_id}/profiles/{profile_id}')
async def program_admin_profile_page(program_id: str, profile_id: str, draft: str = None, request: FastAPIRequest = None):
    """Program admin page to view and edit profile data"""
    global current_user
    
    # Get user from session
    current_user = app.storage.user.get('user')
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Verify user is program admin for this program
    user = users_col.find_one({'email': current_user['email']})
    if not user or user.get('role') != 'Program Admin' or user.get('program_id') != program_id:
        ui.notify('Access denied. You can only access profiles for your assigned program.', color='negative')
        ui.navigate.to('/')
        return
    
    # Get program and profile details
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return
    
    profile = extended_profiles_col.find_one({'_id': ObjectId(profile_id)})
    if not profile:
        ui.notify('Profile not found', color='negative')
        ui.navigate.to(f'/program_admin/{program_id}/criteria')
        return
    
    # Get institution for theming
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    main_color = institution.get('theme_color', '#1a73e8') if institution else '#1a73e8'
    
    # Check for existing draft or submission
    current_data = None
    if draft:
        current_data = criteria_submissions_col.find_one({
            '_id': ObjectId(draft),
            'profile_id': profile_id,
            'status': 'draft'
        })
    else:
        # Check for existing submission
        current_data = criteria_submissions_col.find_one({
            'profile_id': profile_id,
            'program_id': program_id,
            'status': 'draft'
        })
    
    # Form data
    form_data = {}
    if current_data and 'data' in current_data:
        form_data = current_data['data']
    
    # Main layout
    with ui.column().classes('w-full min-h-screen bg-gray-50'):
        # Sticky header with back button and title
        with ui.row().classes('w-full bg-white shadow-sm p-4 items-center sticky top-0 z-20'):
            ui.button(icon='arrow_back',
                    on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                    color='primary').props('flat dense').classes('mr-2')
            
            with ui.column().classes('flex-1'):
                ui.label(f'Profile: {profile.get("name", "")}').classes('text-2xl font-bold text-gray-800')
                if current_data and current_data.get('status') == 'draft':
                    ui.label('Draft').classes('text-xs bg-yellow-100 text-yellow-800 px-2 py-0.5 rounded-full w-fit')
            
            with ui.row().classes('ml-auto gap-2'):
                if current_data and current_data.get('status') == 'draft':
                    ui.button('View All Drafts',
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria?tab=drafts'),
                            icon='drafts',
                            color='blue',
                            outline=True)
        
        # Initialize global variables
        global table_container, current_profile
        current_profile = profile
        
        # Main content area
        with ui.column().classes('w-full max-w-6xl mx-auto p-6 space-y-6') as main_container:
            # Initialize table container
            table_container = ui.column().classes('w-full')
            # Profile Information Section
            with ui.card().classes('w-full'):
                with ui.column().classes('w-full space-y-4'):
                    ui.label('Profile Information').classes('text-xl font-semibold')
                    
                    # Display profile name and description
                    with ui.card().classes('w-full bg-gray-50'):
                        with ui.column().classes('w-full space-y-2'):
                            ui.label('Profile Name').classes('text-sm font-medium text-gray-500')
                            ui.label(profile.get('name', '')).classes('text-lg font-medium')
                            
                            if profile.get('description'):
                                ui.label('Description').classes('mt-4 text-sm font-medium text-gray-500')
                                ui.label(profile.get('description', '')).classes('text-md')
            
            # Data Entry Section
            with ui.card().classes('w-full'):
                with ui.column().classes('w-full space-y-4'):
                    ui.label('Data Entry').classes('text-xl font-semibold')
                    
                    # Display headers for reference
                    if 'headers' in profile and profile['headers']:
                        with ui.expansion('View Headers', icon='list').classes('w-full'):
                            with ui.row().classes('flex flex-wrap gap-2 p-2'):
                                for i, header in enumerate(profile['headers']):
                                    ui.chip(header, color='primary', icon='label')
                    
                    # Entry Method Selection
                    with ui.card().classes('w-full bg-blue-50'):
                        with ui.row().classes('items-center w-full'):
                            ui.label('Select Entry Method:').classes('font-medium')
                            global entry_method
                            entry_method = ui.select(
                                options=['Manual Entry', 'Upload File'],
                                value='Manual Entry'
                            ).classes('ml-4 min-w-[200px]')
                    
                    # File Upload Section (initially hidden)
                    upload_container = ui.column().classes('w-full hidden')
                    
                    # Manual Entry Section (initially visible)
                    manual_entry_container = ui.column().classes('w-full')
                    
                    # Table for manual entry - ALWAYS VISIBLE when there's data
                    table_container = ui.column().classes('w-full overflow-x-auto')
                    
                    # Function to update the table based on entry method
                    global update_entry_ui
                    def update_entry_ui():
                        if entry_method.value == 'Upload File':
                            upload_container.classes(remove='hidden')
                            manual_entry_container.classes('hidden')
                            # Table is always visible when there's data
                            if entry_data and len(entry_data) > 0:
                                table_container.classes(remove='hidden')
                        else:
                            upload_container.classes('hidden')
                            manual_entry_container.classes(remove='hidden')
                            table_container.classes(remove='hidden')
                            render_manual_entry_table()
                    
                    entry_method.on('update:model-value', lambda: update_entry_ui())
                    
                    # File Upload Component
                    with upload_container:
                        with ui.card().classes('w-full p-4'):
                            ui.label('Upload File').classes('text-lg font-medium mb-4')
                            upload = ui.upload(
                                label='Choose a file (CSV, Excel)',
                                multiple=False,
                                on_upload=handle_file_upload,
                                auto_upload=True
                            ).classes('w-full')
                            
                            with upload:
                                ui.label('Drag and drop file here or click to select')
                            
                            # Add a status label
                            status_label = ui.label('Ready to upload file...').classes('mt-2 text-gray-600')
                    
                    # Manual Entry Controls
                    with manual_entry_container:
                        with ui.row().classes('w-full justify-between items-center'):
                            ui.label('Manual Data Entry').classes('text-lg font-medium')
                            ui.button('Add Row', 
                                    icon='add', 
                                    on_click=lambda: add_table_row(),
                                    color='primary')
                    
                    # Table for displaying and editing data - ALWAYS VISIBLE
                    with table_container:
                        print(f"DEBUG: Initializing table container")
                        print(f"DEBUG: current_profile: {current_profile}")
                        # Initialize the table with current profile headers
                        if current_profile and 'headers' in current_profile:
                            headers = current_profile['headers']
                            print(f"DEBUG: Found headers: {headers}")
                            if headers:
                                # Create initial empty row
                                global entry_data
                                entry_data = [{h: '' for h in headers}]
                                print(f"DEBUG: Created initial entry_data: {entry_data}")
                                render_manual_entry_table()
                            else:
                                print("DEBUG: Headers list is empty")
                                ui.label('No headers defined for this profile. Please contact your administrator.').classes('text-red-600 text-center')
                        else:
                            print("DEBUG: No current_profile or no headers")
                            ui.label('Profile information not found. Please refresh the page.').classes('text-red-600 text-center')
            
            # Document Upload Section
            with ui.card().classes('w-full'):
                ui.label('Supporting Documents').classes('text-xl font-semibold mb-4')
                
                # File upload component
                with ui.upload(
                    label='Upload Supporting Documents',
                    multiple=True,
                    on_upload=lambda e: handle_document_upload(e, form_data)
                ).classes('w-full') as upload:
                    ui.label('Drag and drop files here or click to select')
                
                # Display uploaded files
                if 'documents' in form_data and form_data['documents']:
                    with ui.row().classes('w-full flex-wrap gap-2'):
                        for doc in form_data['documents']:
                            with ui.card().classes('relative'):
                                ui.icon('description').classes('text-2xl mb-1')
                                ui.label(doc.get('name', 'Document')).classes('text-sm truncate')
                                ui.button(
                                    icon='delete',
                                    on_click=lambda d=doc: remove_document(d, form_data),
                                    color='red',
                                    size='sm',
                                    flat=True,
                                    dense=True
                                ).classes('absolute top-0 right-0')
            
            # Action buttons
            with ui.row().classes('w-full justify-end gap-4 mt-6'):
                ui.button('Cancel',
                        on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                        color='gray')
                
                ui.button('Save as Draft',
                        on_click=lambda: save_profile_draft(program_id, profile_id, form_data, False),
                        icon='save',
                        color='blue')
                
                ui.button('Submit',
                        on_click=lambda: save_profile_draft(program_id, profile_id, form_data, True),
                        icon='send',
                        color='green')

# Data storage for the table
entry_data = []
table_container = None  # Will be initialized in the page function
current_profile = None  # Will store the current profile data

def normalize_header(header, target_headers):
    """Normalize header by matching with target headers case-insensitively"""
    header_lower = header.strip().lower()
    for target in target_headers:
        if target.lower() == header_lower:
            return target
    return header  # Return original if no match found

def render_manual_entry_table():
    """Render or update the manual entry table"""
    global entry_data, table_container, current_profile
    
    print(f"DEBUG: render_manual_entry_table called")
    print(f"DEBUG: table_container exists: {table_container is not None}")
    print(f"DEBUG: current_profile: {current_profile}")
    
    if not table_container:
        print("DEBUG: No table container, returning")
        return
    
    # Clear existing table content
    table_container.clear()
    
    # Get headers from current profile
    headers = current_profile.get('headers', []) if current_profile else []
    print(f"DEBUG: Headers found: {headers}")
    
    if not headers:
        print("DEBUG: No headers found")
        with table_container:
            ui.label('No headers defined for this profile. Please contact your administrator.').classes('text-red-600 text-center')
        return
    
    # Ensure entry_data has at least one row
    if not entry_data or not isinstance(entry_data, list) or len(entry_data) == 0:
        entry_data = [{h: '' for h in headers}]
        print(f"DEBUG: Created initial entry_data: {entry_data}")
    
    print(f"DEBUG: Creating table with {len(entry_data)} rows and {len(headers)} headers")
    
    # Create the table with proper structure
    with table_container:
        # Create a grid-like table for manual entry
        with ui.card().classes('w-full p-4'):
            ui.label('Data Entry Table').classes('text-lg font-medium mb-4')
            
            # Table headers
            with ui.row().classes('w-full bg-gray-100 p-2 font-medium border-b'):
                for header in headers:
                    ui.label(header).classes('flex-1 text-center')
                ui.label('Actions').classes('w-24 text-center')
            
            # Table rows
            for row_idx, row_data in enumerate(entry_data):
                with ui.row().classes('w-full p-2 border-b hover:bg-gray-50'):
                    for header in headers:
                        # Create input field for each cell
                        input_field = ui.input(
                            value=row_data.get(header, ''),
                            on_change=lambda e, r=row_idx, h=header: update_cell_value(r, h, e.value)
                        ).classes('flex-1 mx-1')
                    # Delete button for this row
                    ui.button(
                        icon='delete',
                        on_click=lambda r=row_idx: remove_table_row_by_index(r),
                        color='red'
                    ).classes('w-24')
            
            # Add row button
            with ui.row().classes('w-full justify-center mt-4'):
                ui.button(
                    'Add New Row',
                    on_click=lambda: add_table_row(),
                    icon='add',
                    color='primary'
                )
    
    print("DEBUG: Table creation complete")

def add_table_row():
    """Add a new empty row to the table"""
    global entry_data, current_profile
    if not current_profile:
        return
    headers = current_profile.get('headers', []) or []
    entry_data.append({h: '' for h in headers})
    render_manual_entry_table()

def update_cell_value(row_idx, header, value):
    """Update a cell value in the entry_data"""
    global entry_data
    if 0 <= row_idx < len(entry_data):
        entry_data[row_idx][header] = value
        # Ensure table is visible when data exists
        if table_container:
            table_container.classes(remove='hidden')

def remove_table_row_by_index(row_idx):
    """Remove a specific row by index"""
    global entry_data
    if len(entry_data) > 1 and 0 <= row_idx < len(entry_data):
        entry_data.pop(row_idx)
        render_manual_entry_table()
    elif len(entry_data) <= 1:
        ui.notify('At least one row is required', color='warning')

def remove_table_row(e):
    """Remove a row from the table (legacy function)"""
    global entry_data
    if len(entry_data) > 1:  # Keep at least one row
        # Remove the last row (simple approach)
        entry_data.pop()
        render_manual_entry_table()
    else:
        ui.notify('At least one row is required', color='warning')

async def handle_file_upload(e: UploadEventArguments):
    """Handle file upload and parse data"""
    global entry_data, current_profile
    
    print("File upload triggered")  # Debug log
    
    if not hasattr(e, 'content') or not e.content:
        ui.notify('No file selected', color='warning')
        return
    
    # Get file content directly from the event
    content = e.content
    filename = getattr(e, 'name', 'uploaded_file')
    print(f"Processing file: {filename}")  # Debug log
    
    try:
        # Read content if it's a file-like object
        if hasattr(content, 'read'):
            content_bytes = content.read()
            content.seek(0)  # Reset file pointer
        else:
            content_bytes = content
        print(f"File read successfully, size: {len(content_bytes)} bytes")  # Debug log
        import pandas as pd
        from io import BytesIO
        
        # Read file based on type
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(BytesIO(content_bytes))
        elif filename.lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(BytesIO(content_bytes))
        else:
            raise ValueError('Unsupported file format. Please upload a CSV or Excel file.')
        
        if df.empty:
            ui.notify('The uploaded file is empty', color='warning')
            return
        
        # Get target headers from current profile
        target_headers = current_profile.get('headers', []) if current_profile else []
        
        # If no target headers, use the file headers
        if not target_headers:
            target_headers = list(df.columns)
            if current_profile:
                current_profile['headers'] = target_headers
        
        # Normalize headers to match target headers if they exist
        if target_headers:
            df.columns = [normalize_header(h, target_headers) for h in df.columns]
        
        # Convert to list of dicts
        entry_data = df.to_dict('records')
        
        # Ensure all rows have all headers
        for row in entry_data:
            for header in target_headers:
                if header not in row:
                    row[header] = ''
        
        # Automatically switch to manual entry view and show the table
        entry_method.value = 'Manual Entry'
        
        # Force the UI update to show manual entry view
        update_entry_ui()
        
        # Ensure table container is visible
        if table_container:
            table_container.classes(remove='hidden')
        
        # Render the table with the new data immediately
        render_manual_entry_table()
        
        ui.notify(f'Successfully loaded {len(entry_data)} rows from {filename}. Table is ready for editing!', color='positive')
    except Exception as ex:
        ui.notify(f'Error processing file: {str(ex)}', color='negative')
        import traceback
        print(traceback.format_exc())

def handle_document_upload(e, form_data):
    """Handle document upload and update form data"""
    if 'documents' not in form_data:
        form_data['documents'] = []
    
    for file in e.files:
        form_data['documents'].append({
            'name': file.name,
            'type': file.type,
            'size': file.size,
            'content': file.read().decode('utf-8') if file.type.startswith('text/') else file.read()
        })
    
    ui.notify(f'Uploaded {len(e.files)} files', color='positive')

def remove_document(doc, form_data):
    """Remove a document from the form data"""
    if 'documents' in form_data and doc in form_data['documents']:
        form_data['documents'].remove(doc)
        ui.notify('Document removed', color='info')

def submit_draft(draft, program_id):
    """Submit a draft to change its status from draft to submitted"""
    try:
        # Show confirmation dialog
        with ui.dialog() as dialog, ui.card():
            ui.label('Confirm Submission').classes('text-lg font-semibold mb-4')
            ui.label('Are you sure you want to submit this draft? This action cannot be undone.').classes('text-gray-600 mb-4')
            
            with ui.row().classes('w-full gap-2'):
                ui.button('Cancel', on_click=dialog.close).classes('bg-gray-500 text-white')
                ui.button('Submit', on_click=lambda: submit_draft_confirm(draft, program_id, dialog)).classes('bg-green-500 text-white')
    
    except Exception as e:
        ui.notify(f'Error preparing submission: {str(e)}', color='negative')

def submit_draft_confirm(draft, program_id, dialog):
    """Actually submit the draft after confirmation"""
    try:
        # Update the draft status to submitted
        if 'criteria_id' in draft:
            # Criteria submission
            criteria_submissions_col.update_one(
                {'_id': draft['_id']},
                {'$set': {
                    'status': 'submitted',
                    'submitted_at': datetime.now(timezone.utc),
                    'updated_at': datetime.now(timezone.utc)
                }}
            )
        else:
            # Extended profile submission
            extended_profile_submissions_col.update_one(
                {'_id': draft['_id']},
                {'$set': {
                    'status': 'submitted',
                    'submitted_at': datetime.now(timezone.utc),
                    'updated_at': datetime.now(timezone.utc)
                }}
            )
        
        dialog.close()
        ui.notify('Draft submitted successfully!', color='positive')
        
        # Refresh the page to show updated status
        ui.navigate.to(f'/program_admin/{program_id}/submissions')
        
    except Exception as e:
        ui.notify(f'Error submitting draft: {str(e)}', color='negative')

def delete_draft(draft, program_id):
    """Delete a draft after confirmation"""
    try:
        # Show confirmation dialog
        with ui.dialog() as dialog, ui.card():
            ui.label('Confirm Deletion').classes('text-lg font-semibold mb-4')
            ui.label('Are you sure you want to delete this draft? This action cannot be undone.').classes('text-red-600 mb-4')
            
            with ui.row().classes('w-full gap-2'):
                ui.button('Cancel', on_click=dialog.close).classes('bg-gray-500 text-white')
                ui.button('Delete', on_click=lambda: delete_draft_confirm(draft, program_id, dialog)).classes('bg-red-500 text-white')
    
    except Exception as e:
        ui.notify(f'Error preparing deletion: {str(e)}', color='negative')

def delete_draft_confirm(draft, program_id, dialog):
    """Actually delete the draft after confirmation"""
    try:
        # Delete the draft
        if 'criteria_id' in draft:
            # Criteria submission
            criteria_submissions_col.delete_one({'_id': draft['_id']})
        else:
            # Extended profile submission
            extended_profile_submissions_col.delete_one({'_id': draft['_id']})
        
        dialog.close()
        ui.notify('Draft deleted successfully!', color='positive')
        
        # Refresh the page to show updated status
        ui.navigate.to(f'/program_admin/{program_id}/submissions')
        
    except Exception as e:
        ui.notify(f'Error deleting draft: {str(e)}', color='negative')

def view_submission(submission, program_id):
    """View a submitted submission"""
    try:
        submission_type = 'Criteria' if 'criteria_id' in submission else 'Extended Profile'
        
        with ui.dialog() as dialog, ui.card().classes('w-full max-w-4xl'):
            ui.label(f'View {submission_type} Submission').classes('text-xl font-semibold mb-4')
            
            # Show submission details
            with ui.column().classes('w-full'):
                ui.label(f'Status: {submission.get("status", "Unknown")}').classes('text-lg mb-2')
                ui.label(f'Submitted by: {submission.get("submitted_by", "Unknown")}').classes('mb-2')
                ui.label(f'Submitted at: {submission.get("submitted_at", "").strftime("%Y-%m-%d %H:%M:%S") if submission.get("submitted_at") else "Unknown"}').classes('mb-4')
                
                # Show table data if available
                if 'data' in submission and 'table_data' in submission['data']:
                    table_data = submission['data']['table_data']
                    if table_data and len(table_data) > 0:
                        headers = list(table_data[0].keys()) if table_data else []
                        with ui.table(columns=headers, rows=table_data, row_key='id').classes('w-full'):
                            pass
                    else:
                        ui.label('No data available in this submission').classes('text-gray-500 italic')
                else:
                    ui.label('No data available in this submission').classes('text-gray-500 italic')
            
            with ui.row().classes('w-full justify-end mt-4'):
                ui.button('Close', on_click=dialog.close).classes('bg-gray-500 text-white')
    
    except Exception as e:
        ui.notify(f'Error viewing submission: {str(e)}', color='negative')

def edit_submitted_submission(submission, program_id):
    """Edit a submitted submission with confirmation"""
    try:
        submission_type = 'Criteria' if 'criteria_id' in submission else 'Extended Profile'
        
        with ui.dialog() as dialog, ui.card():
            ui.label('Confirm Edit').classes('text-lg font-semibold mb-4')
            ui.label(f'Are you sure you want to edit this {submission_type.lower()} submission?').classes('text-gray-600 mb-2')
            ui.label('Editing will change the status back to draft.').classes('text-orange-600 mb-4')
            
            with ui.row().classes('w-full gap-2'):
                ui.button('Cancel', on_click=dialog.close).classes('bg-gray-500 text-white')
                ui.button('Edit', on_click=lambda: edit_submission_confirm(submission, program_id, dialog)).classes('bg-orange-500 text-white')
    
    except Exception as e:
        ui.notify(f'Error preparing edit: {str(e)}', color='negative')

def edit_submission_confirm(submission, program_id, dialog):
    """Actually edit the submission after confirmation"""
    try:
        # Change status back to draft
        if 'criteria_id' in submission:
            # Criteria submission
            criteria_submissions_col.update_one(
                {'_id': submission['_id']},
                {'$set': {
                    'status': 'draft',
                    'updated_at': datetime.now(timezone.utc)
                }}
            )
        else:
            # Extended profile submission
            extended_profile_submissions_col.update_one(
                {'_id': submission['_id']},
                {'$set': {
                    'status': 'draft',
                    'updated_at': datetime.now(timezone.utc)
                }}
            )
        
        dialog.close()
        ui.notify('Submission changed to draft. You can now edit it.', color='positive')
        
        # Refresh the page to show updated status
        ui.navigate.to(f'/program_admin/{program_id}/submissions')
        
    except Exception as e:
        ui.notify(f'Error editing submission: {str(e)}', color='negative')

def download_submission(submission):
    """Download submission data as CSV"""
    try:
        if 'data' in submission and 'table_data' in submission['data']:
            table_data = submission['data']['table_data']
            if table_data and len(table_data) > 0:
                import pandas as pd
                import io
                
                # Convert to DataFrame
                df = pd.DataFrame(table_data)
                
                # Create CSV content
                csv_content = df.to_csv(index=False)
                
                # Create download link
                ui.download(
                    content=csv_content,
                    filename=f"submission_{submission.get('_id', 'data')}.csv",
                    label='Download CSV'
                )
                
                ui.notify('Download started!', color='positive')
            else:
                ui.notify('No data available for download', color='warning')
        else:
            ui.notify('No data available for download', color='warning')
            
    except Exception as e:
        ui.notify(f'Error downloading submission: {str(e)}', color='negative')

def save_profile_draft(program_id: str, profile_id: str, data: Optional[Dict], is_submit: bool):
    """Save profile data as draft or submit it"""
    try:
        # Get user from session
        current_user = app.storage.user.get('user')
        if not current_user or not current_user.get('email'):
            ui.notify('Please log in first', color='negative')
            ui.navigate.to('/')
            return
            
        # Include the table data in the form data
        global entry_data
        if not data:
            data = {}
        if 'table_data' not in data:
            data['table_data'] = []
        data['table_data'] = entry_data
        
        # Prepare submission data
        submission_data = {
            'program_id': ObjectId(program_id),
            'profile_id': ObjectId(profile_id),
            'data': data,
            'status': 'submitted' if is_submit else 'draft',
            'submitted_by': current_user['email'],
            'submitted_at': datetime.now(timezone.utc),
            'updated_at': datetime.now(timezone.utc),
            'created_at': datetime.now(timezone.utc),
            'created_by': current_user['email']
        }
        
        # Check for existing submission (both draft and submitted)
        existing = extended_profile_submissions_col.find_one({
            'profile_id': ObjectId(profile_id),
            'program_id': ObjectId(program_id)
        })
        
        if existing:
            # Update existing submission
            extended_profile_submissions_col.update_one(
                {'_id': existing['_id']},
                {'$set': submission_data}
            )
            submission_id = existing['_id']
            ui.notify(f'Profile {"submitted" if is_submit else "draft"} updated successfully!', color='positive')
        else:
            # Create new submission
            submission_id = extended_profile_submissions_col.insert_one(submission_data).inserted_id
            ui.notify(f'Profile {"submitted" if is_submit else "draft"} created successfully!', color='positive')
        
        # Log the action
        log_audit_action(
            'Profile Submission' if is_submit else 'Profile Draft',
            f'{"Submitted" if is_submit else "Saved as draft"} profile {profile_id} for program {program_id}',
            entity_type='Extended Profile',
            entity_id=str(submission_id)
        )
        
        if is_submit:
            ui.navigate.to(f'/program_admin/{program_id}/submissions')
        else:
            ui.navigate.to(f'/program_admin/{program_id}/profiles/{profile_id}')
            
    except Exception as e:
        ui.notify(f'Error saving profile: {str(e)}', color='negative')
        print(f'Error saving profile: {str(e)}')
        import traceback
        traceback.print_exc()

# Program Admin Fill Criteria Page
@ui.page('/program_admin/{program_id}/fill_criteria/{criteria_id}')
async def program_admin_fill_criteria_page(program_id: str, criteria_id: str, request: FastAPIRequest = None):
    """Program admin page to fill criteria data with file upload and manual entry options"""
    global current_user
    
    # Get user from session
    current_user = app.storage.user.get('user')
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Verify user is program admin for this program
    user = users_col.find_one({'email': current_user['email']})
    if not user or user.get('role') != 'Program Admin' or user.get('program_id') != program_id:
        ui.notify('Access denied. You can only access criteria for your assigned program.', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    import pandas as pd
    
    # Get program and criteria details
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/dashboard')
        return
        
    criteria = criterias_col.find_one({'_id': ObjectId(criteria_id)})
    if not criteria:
        ui.notify('Criteria not found', color='negative')
        ui.navigate.to(f'/program_admin/{program_id}/criteria')
        return
    
    # Get criteria headers
    headers = criteria.get('headers', [])
    if not headers:
        ui.notify('No headers defined for this criteria', color='warning')
    
    # Check for existing submission (prevent duplicate submissions)
    existing_submission = criteria_submissions_col.find_one({
        'criteria_id': criteria_id,
        'program_id': program_id,
        'status': 'submitted'
    })
    
    if existing_submission and not request.query_params.get('edit'):
        ui.notify('This criteria has already been submitted. You can only edit drafts.', color='warning')
    
    institution = institutions_col.find_one({'_id': ObjectId(program['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'

    # Initialize data structures
    df = pd.DataFrame()
    data_rows = []
    file_name = 'N/A'
    upload_method = 'manual'  # 'manual' or 'file'
    
    # Load drafts or existing submission
    drafts = list(criteria_submissions_col.find({
        'criteria_id': criteria_id,
        'program_id': program_id,
        'status': 'draft'
    }).sort('updated_at', -1))
    
    # Check if this is a draft or existing submission load
    draft_id = request.query_params.get('draft') if request else None
    submission_id = request.query_params.get('submission') if request else None
    
    current_data = None
    if draft_id:
        current_data = next((d for d in drafts if str(d['_id']) == draft_id), None)
        if current_data:
            ui.notify('Loaded draft data', color='positive')
    elif submission_id or existing_submission:
        current_data = existing_submission or criteria_submissions_col.find_one({
            '_id': ObjectId(submission_id),
            'criteria_id': criteria_id,
            'program_id': program_id
        })
        if current_data:
            ui.notify('Loaded existing submission', color='positive')
            
    # If we have existing data, prepare it for the form
    if current_data and 'data' in current_data:
        data_rows = current_data['data']
        if 'file_name' in current_data:
            file_name = current_data['file_name']
            upload_method = 'file'
    
    # Show drafts list if not loading a specific draft
    if not draft_id and not submission_id and not request.query_params.get('new') and not request.query_params.get('edit'):
        with ui.column().classes('w-full max-w-6xl mx-auto p-6'):
            # Header with title and new draft button
            with ui.row().classes('w-full justify-between items-center mb-6'):
                ui.label('📋 Drafts & Submissions').classes('text-2xl font-bold text-gray-800')
                ui.button('New Draft', 
                         on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?new=true'),
                         icon='add',
                         color='primary').classes('px-6 py-2')
            
            # Show existing submission if exists
            if existing_submission:
                with ui.card().classes('w-full mb-8 border-l-4 border-green-500'):
                    with ui.row().classes('w-full items-center'):
                        ui.icon('check_circle', color='green', size='28px')
                        ui.label('Submitted').classes('text-xl font-semibold ml-2 text-green-700')
                        ui.label(f'Submitted on: {existing_submission.get("submitted_at", "").strftime("%d %b %Y, %I:%M %p")}').classes('ml-auto text-gray-500')
                    
                    # Show preview of submitted data
                    if existing_submission.get('data'):
                        with ui.expansion('View Submission Data', icon='visibility').classes('w-full mt-4'):
                            with ui.table().classes('w-full border rounded-lg'):
                                # Header
                                with ui.row().classes('bg-gray-100 p-3 font-medium'):
                                    for header in headers:
                                        ui.cell(header).classes('font-semibold')
                                
                                # Data (show first 3 rows as preview)
                                for row in existing_submission.get('data', [])[:3]:
                                    with ui.row().classes('border-t p-3 hover:bg-gray-50'):
                                        for header in headers:
                                            ui.cell(str(row.get(header, ''))).classes('truncate max-w-xs')
                                
                                if len(existing_submission.get('data', [])) > 3:
                                    with ui.row().classes('border-t p-2 bg-gray-50'):
                                        ui.cell(f'+ {len(existing_submission.get("data", [])) - 3} more rows...').classes('text-gray-500')
                    
                    with ui.row().classes('w-full mt-4 gap-3'):
                        ui.button('View Full Submission', 
                                on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?submission={existing_submission["_id"]}'),
                                color='primary',
                                icon='visibility')
                        
                        if existing_submission.get('status') == 'submitted':
                            ui.button('Create New Version', 
                                    on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?edit={existing_submission["_id"]}'),
                                    color='blue',
                                    icon='edit_note')
            
            # Show drafts list if any
            if drafts:
                with ui.card().classes('w-full mt-8'):
                    with ui.row().classes('w-full items-center mb-4'):
                        ui.icon('drafts', color='primary', size='28px')
                        ui.label('Your Drafts').classes('text-xl font-semibold ml-2 text-gray-800')
                    
                    # Table for drafts
                    with ui.table().classes('w-full rounded-lg overflow-hidden'):
                        # Header
                        with ui.row().classes('bg-gray-100 p-4 font-medium'):
                            ui.cell('Created').classes('font-semibold')
                            ui.cell('Last Updated').classes('font-semibold')
                            ui.cell('Data Preview').classes('font-semibold')
                            ui.cell('Actions').classes('font-semibold text-right')
                        
                        # Rows
                        for draft in drafts:
                            with ui.row().classes('border-t hover:bg-gray-50 transition-colors'):
                                # Created and Updated times
                                with ui.cell().classes('p-4'):
                                    ui.label('Created:').classes('text-xs text-gray-500')
                                    ui.label(draft.get('created_at', '').strftime('%d %b %Y, %I:%M %p'))
                                
                                with ui.cell().classes('p-4'):
                                    ui.label('Last Updated:').classes('text-xs text-gray-500')
                                    ui.label(draft.get('updated_at', '').strftime('%d %b %Y, %I:%M %p'))
                                
                                # Data preview
                                with ui.cell().classes('p-4'):
                                    if draft.get('data'):
                                        preview = ', '.join([f"{k}: {v}" for k, v in draft['data'][0].items()][:2])
                                        ui.label(preview).classes('truncate max-w-xs')
                                        ui.label(f"{len(draft.get('data', []))} rows").classes('text-xs text-gray-500')
                                    else:
                                        ui.label('No data').classes('text-gray-400')
                                
                                # Actions
                                with ui.cell().classes('p-4 text-right'):
                                    with ui.row().classes('justify-end gap-2'):
                                        ui.button(icon='edit', 
                                                on_click=lambda d=draft: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?draft={d["_id"]}'),
                                                color='blue',
                                                flat=True,
                                                dense=True,
                                                tooltip='Edit Draft')
                                        ui.button(icon='delete', 
                                                on_click=lambda d=draft: criteria_submissions_col.delete_one({'_id': d['_id']}) or ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}'),
                                                color='red',
                                                flat=True,
                                                dense=True,
                                                tooltip='Delete Draft')
    
    def save_draft(is_submit=False):
        """Save the current form data as a draft or submit it"""
        try:
            # Create submission data dictionary
            submission_data = {
                'data': data_rows,
                'status': 'submitted' if is_submit else 'draft',
                'updated_at': datetime.datetime.utcnow(),
                'updated_by': current_user['email']
            }
            
            if upload_method == 'file':
                submission_data['file_name'] = file_name
        
            # Update existing or insert new
            if current_data and '_id' in current_data:
                criteria_submissions_col.update_one(
                    {'_id': current_data['_id']},
                    {'$set': submission_data}
                )
                submission_id = current_data['_id']
            else:
                submission_data['created_at'] = datetime.datetime.utcnow()
                submission_data['created_by'] = current_user['email']
                submission_id = criteria_submissions_col.insert_one(submission_data).inserted_id
            
            ui.notify('Data saved successfully!', color='positive')
            
            if is_submit:
                ui.navigate.to(f'/program_admin/{program_id}/criteria')
            else:
                ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?draft={submission_id}')
                
        except Exception as e:
            ui.notify(f'Error saving data: {str(e)}', color='negative')
    
    def handle_upload(e):
        """Handle file upload and parse data"""
        nonlocal data_rows, file_name, df
        try:
            file_content = e.content.read()
            file_name = e.name
            
            # Parse file based on extension
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                df = pd.read_excel(io.BytesIO(file_content))
            elif file_name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(file_content))
            else:
                raise ValueError('Unsupported file format. Please upload Excel (.xlsx, .xls) or CSV (.csv) files.')
            
            # Convert all columns to string and normalize headers
            df = df.astype(str)
            df.columns = [str(col).strip().lower() for col in df.columns]
            
            # Convert to list of dicts for the table
            data_rows = df.to_dict('records')
            
            # If no data rows, create one empty row
            if not data_rows:
                data_rows = [{}]
                
            ui.notify(f'Successfully loaded {len(data_rows)} rows from {file_name}', color='positive')
            
        except Exception as e:
            ui.notify(f'Error processing file: {str(e)}', color='negative')
    
    def render_data_entry_form():
        """Render the data entry form with headers, upload/manual options, and supporting docs"""
        with ui.column().classes('w-full max-w-5xl mx-auto gap-6'):
            # Section 1: Show Headers
            with ui.card().classes('w-full p-6'):
                ui.label('📋 Required Fields').classes('text-xl font-semibold mb-4')
                with ui.grid(columns=2, rows=(len(headers) + 1) // 2).classes('w-full gap-4'):
                    for header in headers:
                        with ui.card().classes('p-3 bg-gray-50'):
                            ui.label(header).classes('font-medium text-gray-700')
            
            # Section 2: Data Entry Method
            with ui.card().classes('w-full p-6'):
                ui.label('📤 Data Entry Method').classes('text-xl font-semibold mb-4')
                with ui.row().classes('w-full gap-6'):
                    # Manual Entry Option
                    with ui.column().classes('flex-1 border-2 rounded-lg p-4 transition-all'):
                        ui.radio(['Manual Entry', 'Upload File'], 
                                value='Manual Entry', 
                                on_change=lambda e: set_upload_method(e.value == 'Upload File'))
                        
                        # Manual Entry Form
                        if not upload_method == 'file':
                            with ui.column().classes('w-full mt-4 gap-4'):
                                if not data_rows:
                                    data_rows.append({})
                                
                                # Table for manual entry
                                with ui.table().classes('w-full border rounded-lg'):
                                    # Header
                                    with ui.row().classes('bg-gray-100 p-2'):
                                        for header in headers:
                                            ui.label(header).classes('flex-1 font-medium')
                                    
                                    # Rows
                                    for row_idx, row in enumerate(data_rows):
                                        with ui.row().classes('p-2 border-t hover:bg-gray-50'):
                                            for header in headers:
                                                value = str(row.get(header, ''))
                                                ui.input('', value=value, 
                                                        on_change=lambda e, r=row_idx, h=header: update_data(r, h, e.value)) \
                                                    .classes('flex-1 border rounded px-2 py-1')
                                
                                # Add Row Button
                                ui.button('+ Add Row', on_click=lambda: data_rows.append({}))
                    
                    # File Upload Option
                    with ui.column().classes('flex-1 border-2 rounded-lg p-4 transition-all', 
                                          'border-blue-500 bg-blue-50' if upload_method == 'file' else ''):
                        ui.radio(['Manual Entry', 'Upload File'], 
                                value='Upload File', 
                                on_change=lambda e: set_upload_method(e.value == 'Upload File'))
                        
                        if upload_method == 'file':
                            with ui.column().classes('w-full mt-4 gap-4'):
                                ui.upload(label='📤 Upload File', 
                                         on_upload=handle_upload,
                                         auto_upload=True).classes('w-full')
                                ui.label('Supported formats: .xlsx, .xls, .csv').classes('text-sm text-gray-500')
                                
                                # Preview Table
                                if data_rows:
                                    with ui.card().classes('w-full mt-4'):
                                        ui.label('📋 Preview').classes('font-medium mb-2')
                                        with ui.table().classes('w-full'):
                                            # Header
                                            with ui.row().classes('bg-gray-100 p-2'):
                                                for header in headers:
                                                    ui.label(header).classes('flex-1 font-medium')
                                            
                                            # Rows (limit to 5 for preview)
                                            for row in data_rows[:5]:
                                                with ui.row().classes('p-2 border-t'):
                                                    for header in headers:
                                                        ui.label(str(row.get(header, ''))).classes('flex-1 truncate')
                                            
                                            if len(data_rows) > 5:
                                                with ui.row().classes('p-2 border-t bg-gray-50'):
                                                    ui.label(f'+ {len(data_rows) - 5} more rows...').classes('text-gray-500')
            
            # Section 3: Supporting Documents (if required)
            if criteria.get('requires_documents', False):
                with ui.card().classes('w-full p-6'):
                    ui.label('📎 Supporting Documents').classes('text-xl font-semibold mb-4')
                    ui.label('Please upload any supporting documents required for this submission.')
                    
                    # Document upload area
                    with ui.column().classes('w-full border-2 border-dashed rounded-lg p-6 text-center'):
                        ui.icon('upload', size='48px').classes('text-gray-400 mx-auto mb-2')
                        ui.label('Drag & drop files here or click to browse').classes('text-gray-600 mb-2')
                        ui.button('Select Files', icon='folder_open').classes('mx-auto')
                    
                    # Uploaded files list
                    with ui.column().classes('w-full mt-4'):
                        ui.label('Uploaded Documents:').classes('font-medium mb-2')
                        # Example uploaded files (would be dynamic in real app)
                        with ui.row().classes('items-center gap-2 p-2 bg-gray-50 rounded'):
                            ui.icon('description', color='primary')
                            ui.label('document1.pdf')
                            ui.space()
                            ui.button(icon='delete', color='red').props('flat dense')
            
            # Action Buttons
            with ui.row().classes('w-full justify-end gap-4 mt-6'):
                ui.button('Cancel', 
                         on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                         color='gray')
                
                ui.button('Save as Draft', 
                         on_click=lambda: save_draft(data_rows, False), 
                         icon='save',
                         color='blue')
                
                ui.button('Submit', 
                         on_click=lambda: save_draft(data_rows, True), 
                         icon='send',
                         color='green')
    
    def update_data(row_idx, header, value):
        """Update data when input fields change"""
        nonlocal data_rows
        if row_idx >= len(data_rows):
            data_rows.extend([{} for _ in range(row_idx - len(data_rows) + 1)])
        if header not in data_rows[row_idx]:
            data_rows[row_idx][header] = ''
        data_rows[row_idx][header] = value
    
    def set_upload_method(is_file_upload):
        """Toggle between manual and file upload modes"""
        nonlocal upload_method
        upload_method = 'file' if is_file_upload else 'manual'

    # Main layout with fixed header and sidebar
    with ui.column().classes('w-full h-screen bg-gray-50'):
        # Sticky header with back button and title
        with ui.row().classes('w-full bg-white shadow-sm p-4 items-center sticky top-0 z-20'):
            ui.button(icon='arrow_back', 
                     on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'),
                     color='primary').props('flat dense').classes('mr-2')
            
            with ui.column().classes('flex-1'):
                ui.label(f'Fill Criteria: {criteria.get("name", "")}').classes('text-2xl font-bold text-gray-800')
                if current_data and current_data.get('status') == 'draft':
                    ui.label('Draft').classes('text-xs bg-yellow-100 text-yellow-800 px-2 py-0.5 rounded-full w-fit')
            
            with ui.row().classes('ml-auto gap-2'):
                if current_data and current_data.get('status') == 'draft':
                    ui.button('View All Drafts', 
                            on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}'),
                            icon='drafts',
                            color='blue',
                            outline=True)
        
        # Main content area with sidebar and form
        with ui.row().classes('w-full flex-1 overflow-hidden'):
            # Fixed Sidebar
            with ui.column().classes('h-[calc(100vh-4rem)] bg-white shadow-md w-64 fixed'):
                program_admin_sidebar(program_id, institution)
            
            # Scrollable content area
            with ui.column().classes('ml-64 w-[calc(100%-16rem)] h-[calc(100vh-4rem)] overflow-y-auto p-6'):
                # Show existing submission notice if exists
                if existing_submission and not draft_id:
                    with ui.card().classes('w-full bg-yellow-50 border-l-4 border-yellow-400 mb-6'):
                        with ui.row().classes('items-center w-full'):
                            ui.icon('warning', size='24px', color='yellow-700')
                            ui.label('You have already submitted this criteria.').classes('ml-2 text-yellow-800')
                            ui.button('View Submission', 
                                    on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?submission={existing_submission["_id"]}'),
                                    color='yellow-700', 
                                    outline=True).classes('ml-auto')
                
                # Render the appropriate view based on state
                if not current_data and not request.query_params.get('new'):
                    # Show drafts list or initial state
                    if drafts:
                        with ui.dialog() as drafts_dialog, ui.card().classes('p-4 w-full max-w-3xl'):
                            ui.label('📝 Your Drafts').classes('text-xl font-bold mb-4')
                            
                            # Create a table to show drafts
                            columns = [
                                {'name': 'created_at', 'label': 'Created', 'field': 'created_at', 'sortable': True},
                                {'name': 'updated_at', 'label': 'Last Updated', 'field': 'updated_at', 'sortable': True},
                                {'name': 'actions', 'label': 'Actions', 'field': 'actions'}
                            ]
                            
                            rows = []
                            for draft in drafts:
                                rows.append({
                                    'created_at': draft.get('created_at', 'N/A').strftime('%Y-%m-%d %H:%M'),
                                    'updated_at': draft.get('updated_at', 'N/A').strftime('%Y-%m-%d %H:%M'),
                                    'actions': str(draft['_id'])
                                })
                            
                            table = ui.table(columns=columns, rows=rows, row_key='actions').classes('w-full')
                            
                            # Add action buttons
                            with table.add_slot('body-cell-actions'):
                                table.add_slot('body-cell-actions', '''
                                    <q-td :props="props">
                                        <q-btn flat round dense color="primary" icon="edit" @click="$parent.$emit('edit', props.row.actions)" />
                                        <q-btn flat round dense color="negative" icon="delete" @click="$parent.$emit('delete', props.row.actions)" />
                                    </q-td>
                                ''')
                            
                            # Handle edit/delete actions
                            def on_edit(draft_id):
                                ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?draft={draft_id}')
                                drafts_dialog.close()
                            
                            def on_delete(draft_id):
                                criteria_submissions_col.delete_one({'_id': ObjectId(draft_id)})
                                ui.notify('Draft deleted', color='positive')
                                ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}')
                            
                            table.on('edit', on_edit)
                            table.on('delete', on_delete)
                            
                            with ui.row().classes('w-full justify-between mt-4'):
                                ui.button('Create New', on_click=drafts_dialog.close).classes('bg-blue-500 text-white')
                                ui.button('Cancel', on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/criteria'))
                        
                        # Show drafts dialog
                        ui.timer(0.1, lambda: drafts_dialog.open())
                    else:
                        # Show empty state with create new option
                        with ui.column().classes('items-center justify-center h-64 w-full bg-white rounded-lg border-2 border-dashed border-gray-300'):
                            ui.icon('description', size='48px').classes('text-gray-400 mb-4')
                            ui.label('No drafts found').classes('text-xl text-gray-500 mb-2')
                            ui.button('Create New Draft', 
                                     on_click=lambda: ui.navigate.to(f'/program_admin/{program_id}/fill_criteria/{criteria_id}?new=true'),
                                     icon='add').classes('mt-4 bg-blue-500 text-white')
                else:
                    # Show the data entry form
                    render_data_entry_form()
        ui.label(f'Program: {program.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem;'
        )
        criterias = list(criterias_col.find({'institution_id': program['institution_id'], 'scope_type': 'program_based'}))
        if criterias:
            for criteria in criterias:
                with ui.card().classes('beautiful-card').style('margin-bottom: 1.5rem; padding: 1.5rem; border-radius: 12px; box-shadow: var(--shadow);'):
                    ui.label(f"📋 {criteria.get('name', 'Unnamed Criteria')}").style('font-size: 1.2rem; font-weight: bold; color: var(--primary-color); margin-bottom: 0.5rem;')
                    if criteria.get('description'):
                        ui.label(criteria['description']).style('font-size: 1rem; color: var(--text-secondary); margin-bottom: 0.5rem;')
                    if criteria.get('deadline'):
                        ui.label(f"⏰ Deadline: {criteria['deadline'].strftime('%Y-%m-%d')}").style('font-size: 0.95rem; color: var(--warning-color); margin-bottom: 0.5rem;')
                    headers = criteria.get('headers', [])
                    def content():
                        # ... existing submissions page content ...
                        pass
                    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
                        program_admin_sidebar(program_id, institution)
                        with ui.column().style('margin-left: 200px; flex: 1; padding: 2rem; overflow-y: auto; height: 100vh;'):
                            content()
        data_table_container.style('display: block;')
        print("DEBUG: Data table container display set to block")
        with data_table_container:
            print("DEBUG: Creating table content")
            ui.label('📊 Data Table').style(
                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label(f'Parsed {len(data_rows)} rows from {file_name}').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            # Create editable data table
            if data_rows:
                print(f"DEBUG: Creating {len(data_rows)} row cards")
                # Create a grid of input fields for each row
                for row_index, row_data in enumerate(data_rows):
                    print(f"DEBUG: Creating row {row_index + 1}")
                    with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                        ui.label(f'Row {row_index + 1}').style(
                            'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                        )
        if not df.empty:
            print(f"DEBUG: Loaded {len(df)} rows from Excel file")
        else:
            print("DEBUG: No data found in the file")
        print(f"DEBUG: Excel parsing complete. Found {len(data_rows)} rows")
        # Show data table
        print(f"DEBUG: About to show data table with {len(data_rows)} rows")
        data_table_container.style('display: block;')
        print("DEBUG: Data table container display set to block")
        with data_table_container:
            print("DEBUG: Creating table content")
            ui.label('📊 Data Table').style(
                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label(f'Parsed {len(data_rows)} rows from {file_name}').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            # Create editable data table
            if data_rows:
                print(f"DEBUG: Creating {len(data_rows)} row cards")
                # Create a grid of input fields for each row
                for row_index, row_data in enumerate(data_rows):
                    print(f"DEBUG: Creating row {row_index + 1}")
                    with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                        ui.label(f'Row {row_index + 1}').style(
                            'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                        )
                        # Create input fields for each header
                        with ui.row().style('flex-wrap: wrap; gap: 1rem;'):
                            for header in headers:
                                with ui.column().style('min-width: 200px;'):
                                    ui.label(header).style(
                                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                                    )
                                    # Check if cell is empty and highlight it
                                    cell_value = row_data.get(header, '')
                                    is_empty = not cell_value.strip()
                                    input_field = ui.input(
                                        label='',
                                        value=cell_value,
                                        placeholder=f'Enter {header}'
                                    ).style(f'width: 100%; border: {"2px solid #dc3545" if is_empty else "1px solid #e9ecef"};')
                                    # Store reference to input field for later data collection
                                    row_data[f'input_{header}'] = input_field
                
                # Always show at least one empty row for manual data entry
                if not data_rows:
                    print("DEBUG: No data rows found, creating empty row for manual entry")
                    ui.label('📝 Manual Data Entry').style(
                        f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
                    )
                    ui.label('No data found in the file. You can fill the data manually below:').style(
                        'color: var(--text-secondary); margin-bottom: 1.5rem; text-align: center;'
                    )
                    
                    # Create empty row for manual entry
                    with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                        ui.label('Row 1 (Manual Entry)').style(
                            'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                        )
                        
                        # Create input fields for each header
                        with ui.row().style('flex-wrap: wrap; gap: 1rem;'):
                            for header in headers:
                                with ui.column().style('min-width: 200px;'):
                                    ui.label(header).style(
                                        'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                                    )
                                    
                                    input_field = ui.input(
                                        label='',
                                        value='',
                                        placeholder=f'Enter {header}'
                                    ).style('width: 100%; border: 2px dashed #e9ecef;')
                                    
                                    # Store reference to input field for later data collection
                                    if not data_rows:
                                        data_rows = [{}]
                                    data_rows[0][f'input_{header}'] = input_field
                
                print("DEBUG: Creating action buttons")
                # Action buttons (always show)
                with ui.row().style('gap: 1rem; margin-top: 1.5rem; justify-content: center;'):
                    ui.button('💾 Save as Draft', on_click=lambda: save_as_draft(data_rows, headers)).style(
                        'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                    )
                    ui.button('✅ Submit', on_click=lambda: submit_data(data_rows, headers)).style(
                        f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                    )
            
            print("DEBUG: Table creation complete, showing success notification")
            ui.notify(f'File "{file_name}" parsed successfully! {len(data_rows)} rows loaded. Please review and edit the data.', color='positive')
            # ...existing code...
    
    def submit_data(data_rows=None, headers=None):
        if not data_rows:
            ui.notify('No data to save. Please upload a file first.', color='warning')
            return
        
        try:
            # Collect data from input fields
            collected_data = []
            for row_data in data_rows:
                row_values = {}
                for header in headers:
                    input_field = row_data.get(f'input_{header}')
                    if input_field and hasattr(input_field, 'value'):
                        row_values[header] = input_field.value
                    else:
                        row_values[header] = input_field.value
                collected_data.append(row_values)
            
            print(f"DEBUG: Submitting data with {len(collected_data)} rows")
            
            # Save to database as submitted
            submission_doc = {
                'criteria_id': ObjectId(criteria_id),
                'program_id': ObjectId(program_id),
                'institution_id': ObjectId(criteria['institution_id']),
                'academic_year_id': criteria.get('academic_year_id'),
                'data': collected_data,
                'headers': headers,
                'status': 'submitted',
                'submitted_by': current_user['email'],
                'submitted_at': datetime.now(timezone.utc),
                'created_at': datetime.now(timezone.utc),
                'updated_at': datetime.now(timezone.utc)
            }
            
            # Insert into criteria_submissions collection
            result = criteria_submissions_col.insert_one(submission_doc)
            print(f"DEBUG: Data submitted to database with ID: {result.inserted_id}")
            
            ui.notify(f'Data submitted successfully! {len(collected_data)} rows processed. Submission ID: {str(result.inserted_id)}', color='positive')
            
        except Exception as e:
            ui.notify(f'Error submitting data: {str(e)}', color='negative')
            print(f"ERROR submitting data: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def save_manual_draft(manual_data, headers):
            """Save manual entry as draft"""
            try:
                # Collect data from manual input fields
                collected_data = []
                row_values = {}
                for header in headers:
                    input_field = manual_data.get(header)
                    if input_field and hasattr(input_field, 'value'):
                        row_values[header] = input_field.value
                    else:
                        row_values[header] = ''
                
                if any(row_values.values()):  # Only save if there's some data
                    collected_data.append(row_values)
                    
                    # Save to database as draft
                    draft_doc = {
                        'criteria_id': ObjectId(criteria_id),
                        'program_id': ObjectId(program_id),
                        'institution_id': ObjectId(criteria['institution_id']),
                        'academic_year_id': criteria.get('academic_year_id'),
                        'data': collected_data,
                        'headers': headers,
                        'status': 'draft',
                        'submitted_by': current_user['email'],
                        'submitted_at': datetime.now(timezone.utc),
                        'created_at': datetime.now(timezone.utc),
                        'updated_at': datetime.now(timezone.utc)
                    }
                    
                    # Insert into criteria_submissions collection
                    result = criteria_submissions_col.insert_one(draft_doc)
                    print(f"DEBUG: Manual draft saved to database with ID: {result.inserted_id}")
                    
                    ui.notify(f'Manual entry saved as draft! Draft ID: {str(result.inserted_id)}', color='positive')
                else:
                    ui.notify('Please enter some data before saving', color='warning')
                    
            except Exception as e:
                ui.notify(f'Error saving manual draft: {str(e)}', color='negative')
                print(f"ERROR saving manual draft: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def content():
        ui.label(f'📋 Fill Criteria: {criteria.get("name", "Unknown")}').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label(f'Program: {program.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        # Show headers for reference
        headers = criteria.get('headers', [])
        if headers:
            with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 2rem; border-radius: 10px;'):
                ui.label('📋 Spreadsheet Headers (Reference)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                with ui.row().style('flex-wrap: wrap; gap: 0.5rem;'):
                    for i, header in enumerate(headers):
                        ui.label(header).style(
                            'background: white; color: var(--text-primary); padding: 0.5rem 1rem; border-radius: 6px; border: 1px solid #e9ecef; font-weight: 500; font-size: 0.9rem;'
                        )
        
        # Define the file upload handler first
        def handle_file_upload(e, headers):
            # Parse the uploaded file and show data table
            try:
                # This would parse the file and extract data
                # For now, show a sample table
                data_table_container.style('display: block;')
                
                with data_table_container:
                    ui.label('📊 Data Preview').style(
                        f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                    )
                    
                    # Sample data table with editable cells
                    with ui.table(columns=headers, rows=[]).style('width: 100%;'):
                        # Table rows would be added here
                        pass
                
                ui.notify('File uploaded successfully!', color='positive')
            except Exception as ex:
                ui.notify(f'Error processing file: {str(ex)}', color='negative')
        
        # File upload section
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 2rem; margin-bottom: 2rem; border-radius: 10px; text-align: center;'):
            ui.label('📤 Upload Spreadsheet').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label('Upload your Excel/CSV file with data matching the headers above').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            
                    # File upload component
        file_upload = ui.upload(
            label='Choose File',
            on_upload=lambda e: handle_file_upload(e, headers)
        ).style(f'background: {main_color}; color: white; padding: 1rem; border-radius: 8px; border: none;')
        
        ui.label('Supported formats: .xlsx, .xls, .csv').style(
            'font-size: 0.9rem; color: var(--text-secondary); margin-top: 1rem;'
        )
        
        # Manual data entry form
        ui.separator().style('margin: 2rem 0; background: #e9ecef;')
        ui.label('📝 Manual Data Entry').style(
            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        # Create manual entry form
        manual_data = {}
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 1.5rem; margin-bottom: 1.5rem; border-radius: 8px;'):
            ui.label('Row 1 (Manual Entry)').style(
                'font-weight: bold; color: var(--text-primary); margin-bottom: 1rem; text-align: center;'
            )
            with ui.row().style('flex-wrap: wrap; gap: 1rem; justify-content: center;'):
                for header in headers:
                    with ui.column().style('min-width: 200px;'):
                        ui.label(header).style(
                            'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                        )
                        input_field = ui.input(
                            label='',
                            value='',
                            placeholder=f'Enter {header}'
                        ).style('width: 100%; border: 2px dashed #e9ecef;')
                        manual_data[header] = input_field
        
        # Action buttons for manual entry
        with ui.row().style('gap: 1rem; margin-bottom: 2rem; justify-content: center;'):
            ui.button('💾 Save Manual Entry as Draft', on_click=lambda: save_manual_draft(manual_data, headers)).style(
                'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
            )
            ui.button('✅ Submit Manual Entry', on_click=lambda: save_manual_draft(manual_data, headers)).style(
                f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
            )
        
        ui.separator().style('margin: 2rem 0; background: #e9ecef;')
        ui.label('📊 File Upload Results').style(
            f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
    
    # Data table section (initially hidden)
    data_table_container.style('display: none;')
    
    # Add the data table container to the UI
    data_table_container
    
    # Simple layout for this page
    with ui.column().style('padding: 2rem; max-width: 1200px; margin: 0 auto;'):
        content()

# Program Admin Fill Extended Profile Page
@ui.page('/program_admin/{program_id}/fill_profile/{profile_id}')
def program_admin_fill_profile_page(program_id: str, profile_id: str):
    """Program admin page to fill extended profile data"""
    global current_user
    
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Verify user is program admin for this program
    user = users_col.find_one({'email': current_user['email']})
    if not user or user.get('role') != 'Program Admin' or user.get('program_id') != program_id:
        ui.notify('Access denied. You can only access profiles for your assigned program.', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    
    # Get profile details
    profile = extended_profiles_col.find_one({'_id': ObjectId(profile_id)})
    if not profile:
        ui.notify('Extended profile not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get program details
    program = programs_col.find_one({'_id': ObjectId(program_id)})
    if not program:
        ui.notify('Program not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details for theming
    institution = institutions_col.find_one({'_id': ObjectId(profile['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content():
        ui.label(f'👤 Fill Extended Profile: {profile.get("name", "Unknown")}').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label(f'Program: {program.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        # Show headers for reference
        headers = profile.get('headers', [])
        if headers:
            with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 2rem; border-radius: 10px;'):
                ui.label('📋 Spreadsheet Headers (Reference)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                with ui.row().style('flex-wrap: wrap; gap: 0.5rem;'):
                    for i, header in enumerate(headers):
                        ui.label(header).style(
                            'background: white; color: var(--text-primary); padding: 0.5rem 1rem; border-radius: 6px; border: 1px solid #e9ecef; font-weight: 500; font-size: 0.9rem;'
                        )
        
        # File upload section
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 2rem; margin-bottom: 2rem; border-radius: 10px; text-align: center;'):
            ui.label('📤 Upload Spreadsheet').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label('Upload your Excel/CSV file with data matching the headers above').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            
            # File upload component
            file_upload = ui.upload(
                label='Choose File',
                on_upload=lambda e: handle_file_upload(e, headers)
            ).style(f'background: {main_color}; color: white; padding: 1rem; border-radius: 8px; border: none;')
        
        # Data table section (initially hidden)
        data_table_container = ui.column().style('display: none;')
        
        def handle_file_upload(e, headers):
            # Parse the uploaded file and show data table
            try:
                # This would parse the file and extract data
                # For now, show a sample table
                data_table_container.style('display: block;')
                
                with data_table_container:
                    ui.label('📊 Data Preview').style(
                        f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                    )
                    
                    # Sample data table with editable cells
                    with ui.table(columns=headers, rows=[]).style('width: 100%;'):
                        pass
                    
                    # Action buttons
                    with ui.row().style('gap: 1rem; margin-top: 1.5rem;'):
                        ui.button('💾 Save as Draft', on_click=lambda: save_as_draft()).style(
                            f'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                        )
                        ui.button('✅ Submit', on_click=lambda: submit_data()).style(
                            f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                        )
                
                ui.notify('File uploaded successfully! Please review and edit the data.', color='positive')
                
            except Exception as e:
                ui.notify(f'Error processing file: {str(e)}', color='negative')
        
        def save_as_draft():
            ui.notify('Data saved as draft successfully!', color='positive')
        
        def submit_data():
            ui.notify('Data submitted successfully!', color='positive')
    
    # Simple layout for this page
    with ui.column().style('padding: 2rem; max-width: 1200px; margin: 0 auto;'):
        content()

# Department Admin Fill Criteria Page
@ui.page('/department_admin/{department_id}/fill_criteria/{criteria_id}')
def department_admin_fill_criteria_page(department_id: str, criteria_id: str):
    """Department admin page to fill criteria data"""
    global current_user
    
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Verify user is department admin for this department
    user = users_col.find_one({'email': current_user['email']})
    if not user or user.get('role') != 'Department Admin' or user.get('department_id') != department_id:
        ui.notify('Access denied. You can only access profiles for your assigned department.', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    
    # Get criteria details
    criteria = criterias_col.find_one({'_id': ObjectId(criteria_id)})
    if not criteria:
        ui.notify('Criteria not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get department details
    department = departments_col.find_one({'_id': ObjectId(department_id)})
    if not department:
        ui.notify('Department not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details for theming
    institution = institutions_col.find_one({'_id': ObjectId(criteria['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content():
        ui.label(f'📋 Fill Criteria: {criteria.get("name", "Unknown")}').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label(f'Department: {department.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        # Show headers for reference
        headers = criteria.get('headers', [])
        if headers:
            with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 2rem; border-radius: 10px;'):
                ui.label('📋 Spreadsheet Headers (Reference)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                with ui.row().style('flex-wrap: wrap; gap: 0.5rem;'):
                    for i, header in enumerate(headers):
                        ui.label(header).style(
                            'background: white; color: var(--text-primary); padding: 0.5rem 1rem; border-radius: 6px; border: 1px solid #e9ecef; font-weight: 500; font-size: 0.9rem;'
                        )
        
        # File upload section
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 2rem; margin-bottom: 2rem; border-radius: 10px; text-align: center;'):
            ui.label('📤 Upload Spreadsheet').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label('Upload your Excel/CSV file with data matching the headers above').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            
            # File upload component
            file_upload = ui.upload(
                label='Choose File',
                on_upload=lambda e: handle_file_upload(e, headers)
            ).style(f'background: {main_color}; color: white; padding: 1rem; border-radius: 8px; border: none;')
            
            ui.label('Supported formats: .xlsx, .xls, .csv').style(
                'font-size: 0.9rem; color: var(--text-secondary); margin-top: 1rem;'
            )
        
        # Data table section (initially hidden)
        data_table_container = ui.column().style('display: none;')
        
        def handle_file_upload(e, headers):
            """Parse uploaded file and create editable data table"""
            try:
                # Debug logging
                print(f"DEBUG: File upload triggered for file: {e.name}")
                print(f"DEBUG: Headers received: {headers}")
                
                # Get the uploaded file
                uploaded_file = e.content
                file_name = e.name
                
                print(f"DEBUG: File size: {len(uploaded_file)} bytes")
                print(f"DEBUG: File extension: {file_name.split('.')[-1] if '.' in file_name else 'no extension'}")
                
                # Parse the file based on extension
                print(f"DEBUG: Starting file parsing for {file_name}")
                
                if file_name.endswith('.csv'):
                    print("DEBUG: Parsing CSV file")
                    # Parse CSV file
                    import csv
                    import io
                    
                    # Read CSV content
                    csv_content = uploaded_file.decode('utf-8')
                    print(f"DEBUG: CSV content length: {len(csv_content)} characters")
                    csv_reader = csv.DictReader(io.StringIO(csv_content))
                    
                    # Extract data rows
                    data_rows = []
                    for row in csv_reader:
                        # Ensure all headers are present, fill missing ones with empty string
                        clean_row = {}
                        for header in headers:
                            clean_row[header] = row.get(header, '')
                        data_rows.append(clean_row)
                    
                    print(f"DEBUG: CSV parsing complete. Found {len(data_rows)} rows")
                    print(f"DEBUG: Excel parsing complete. Found {len(data_rows)} rows")
                    # Show data table
                    print(f"DEBUG: About to show data table with {len(data_rows)} rows")
                    data_table_container.style('display: block;')
                    print("DEBUG: Data table container display set to block")
                    with data_table_container:
                        print("DEBUG: Creating table content")
                        ui.label('📊 Data Table').style(
                            f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                        )
                        ui.label(f'Parsed {len(data_rows)} rows from {file_name}').style(
                            'color: var(--text-secondary); margin-bottom: 1.5rem;'
                        )
                        # Create editable data table
                        if data_rows:
                            print(f"DEBUG: Creating {len(data_rows)} row cards")
                            # Try to get value from row, fallback to empty string
                            value = row.get(header, '')
                            if pd.isna(value):
                                value = ''
                            clean_row[header] = str(value)
                        data_rows.append(clean_row)
                    
                    print(f"DEBUG: Excel parsing complete. Found {len(data_rows)} rows")
                else:
                    print(f"DEBUG: Unsupported file format: {file_name}")
                    ui.notify('Unsupported file format. Please use .csv, .xlsx, or .xls files.', color='negative')
                    return
                
                # Show data table
                print(f"DEBUG: About to show data table with {len(data_rows)} rows")
                data_table_container.style('display: block;')
                print("DEBUG: Data table container display set to block")
                
                with data_table_container:
                    print("DEBUG: Creating table content")
                    ui.label('📊 Data Table').style(
                        f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                    )
                    
                    ui.label(f'Parsed {len(data_rows)} rows from {file_name}').style(
                        'color: var(--text-secondary); margin-bottom: 1.5rem;'
                    )
                    
                    # Create editable data table
                    if data_rows:
                        print(f"DEBUG: Creating {len(data_rows)} row cards")
                        # Create a grid of input fields for each row
                        for row_index, row_data in enumerate(data_rows):
                            print(f"DEBUG: Creating row {row_index + 1}")
                            with ui.card().style('background: white; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                                ui.label(f'Row {row_index + 1}').style(
                                    'font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                                )
                                
                                # Create input fields for each header
                                with ui.row().style('flex-wrap: wrap; gap: 1rem;'):
                                    for header in headers:
                                        with ui.column().style('min-width: 200px;'):
                                            ui.label(header).style(
                                                'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.25rem;'
                                            )
                                            
                                            # Check if cell is empty and highlight it
                                            cell_value = row_data.get(header, '')
                                            is_empty = not cell_value.strip()
                                            
                                            input_field = ui.input(
                                                label='',
                                                value=cell_value,
                                                placeholder=f'Enter {header}'
                                            ).style(f'width: 100%; border: {"2px solid #dc3545" if is_empty else "1px solid #e9ecef"};')
                                            
                                            # Store reference to input field for later data collection
                                            row_data[f'input_{header}'] = input_field
                        
                        print("DEBUG: Creating action buttons")
                        # Action buttons
                        with ui.row().style('gap: 1rem; margin-top: 1.5rem; justify-content: center;'):
                            ui.button('💾 Save as Draft', on_click=lambda: save_as_draft(data_rows, headers)).style(
                                'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                            )
                            ui.button('✅ Submit', on_click=lambda: submit_data(data_rows, headers)).style(
                                f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                            )
                    else:
                        print("DEBUG: No data rows found")
                        ui.label('No data found in the file. Please check the file format and headers.').style(
                            'color: var(--warning-color); text-align: center; padding: 2rem;'
                        )
                
                print("DEBUG: Table creation complete, showing success notification")
                ui.notify(f'File "{file_name}" parsed successfully! {len(data_rows)} rows loaded. Please review and edit the data.', color='positive')
                
            except Exception as e:
                ui.notify(f'Error processing file: {str(e)}', color='negative')
                import traceback
                traceback.print_exc()
        
        def save_as_draft(data_rows=None, headers=None):
            if not data_rows:
                ui.notify('No data to save. Please upload a file first.', color='warning')
                return
            
            if not data_rows:
                ui.notify('No data to save. Please upload a file first.', color='warning')
                return
            try:
                # Collect data from input fields
                collected_data = []
                for row_data in data_rows:
                    row_values = {}
                    for header in headers:
                        input_field = row_data.get(f'input_{header}')
                        if input_field and hasattr(input_field, 'value'):
                            row_values[header] = input_field.value
                        else:
                            row_values[header] = row_data.get(header, '')
                    collected_data.append(row_values)
                print(f"DEBUG: Saving draft with {len(collected_data)} rows")
                # Save to database as draft
                draft_doc = {
                    'criteria_id': ObjectId(criteria_id),
                    'program_id': ObjectId(program_id),
                    'institution_id': ObjectId(criteria['institution_id']),
                    'academic_year_id': criteria.get('academic_year_id'),
                    'data': collected_data,
                    'headers': headers,
                    'status': 'draft',
                    'submitted_by': current_user['email'],
                    'submitted_at': datetime.now(timezone.utc),
                    'created_at': datetime.now(timezone.utc),
                    'updated_at': datetime.now(timezone.utc)
                }
                # Insert into criteria_submissions collection
                result = criteria_submissions_col.insert_one(draft_doc)
                print(f"DEBUG: Draft saved to database with ID: {result.inserted_id}")
                ui.notify(f'Draft saved successfully! {len(collected_data)} rows stored. Draft ID: {str(result.inserted_id)}', color='positive')
            except Exception as e:
                ui.notify(f'Error processing file: {str(e)}', color='negative')
                try:
                    if not data_rows:
                        ui.notify('No data to save. Please upload a file first.', color='warning')
                        return
                    # Collect data from input fields
                    collected_data = []
                    for row_data in data_rows:
                        row_values = {}
                        for header in headers:
                            input_field = row_data.get(f'input_{header}')
                            if input_field and hasattr(input_field, 'value'):
                                row_values[header] = input_field.value
                            else:
                                row_values[header] = row_data.get(header, '')
                        collected_data.append(row_values)
                    print(f"DEBUG: Saving draft with {len(collected_data)} rows")
                    # Save to database as draft
                    draft_doc = {
                        'criteria_id': ObjectId(criteria_id),
                        'program_id': ObjectId(program_id),
                        'institution_id': ObjectId(criteria['institution_id']),
                        'academic_year_id': criteria.get('academic_year_id'),
                        'data': collected_data,
                        'headers': headers,
                        'status': 'draft',
                        'submitted_by': current_user['email'],
                        'submitted_at': datetime.now(timezone.utc),
                        'created_at': datetime.now(timezone.utc),
                        'updated_at': datetime.now(timezone.utc)
                    }
                    # Insert into criteria_submissions collection
                    result = criteria_submissions_col.insert_one(draft_doc)
                    print(f"DEBUG: Draft saved to database with ID: {result.inserted_id}")
                    ui.notify(f'Draft saved successfully! {len(collected_data)} rows stored. Draft ID: {str(result.inserted_id)}', color='positive')
                except Exception as e:
                    ui.notify(f'Error saving draft: {str(e)}', color='negative')
                    print(f"ERROR saving draft: {str(e)}")
                    import traceback
                    traceback.print_exc()
        return
    
    # Get institution details for theming
    institution = institutions_col.find_one({'_id': ObjectId(profile['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content():
        ui.label(f'👤 Fill Extended Profile: {profile.get("name", "Unknown")}').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label(f'Department: {department.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        # Show headers for reference
        headers = profile.get('headers', [])
        if headers:
            with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1.5rem; margin-bottom: 2rem; border-radius: 10px;'):
                ui.label('📋 Spreadsheet Headers (Reference)').style(
                    f'font-size: 1.2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                )
                with ui.row().style('flex-wrap: wrap; gap: 0.5rem;'):
                    for i, header in enumerate(headers):
                        ui.label(header).style(
                            'background: white; color: var(--text-primary); padding: 0.5rem 1rem; border-radius: 6px; border: 1px solid #e9ecef; font-weight: 500; font-size: 0.9rem;'
                        )
        
        # File upload section
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 2rem; margin-bottom: 2rem; border-radius: 10px; text-align: center;'):
            ui.label('📤 Upload Spreadsheet').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
            )
            ui.label('Upload your Excel/CSV file with data matching the headers above').style(
                'color: var(--text-secondary); margin-bottom: 1.5rem;'
            )
            
            # File upload component
            file_upload = ui.upload(
                label='Choose File',
                on_click=lambda e: handle_file_upload(e, headers)
            ).style(f'background: white; color: var(--text-primary); padding: 1rem; border-radius: 8px; border: none;')
        
        # Data table section (initially hidden)
        data_table_container = ui.column().style('display: none;')
        
        def handle_file_upload(e, headers):
            # Parse the uploaded file and show data table
            try:
                # This would parse the file and extract data
                # For now, show a sample table
                data_table_container.style('display: block;')
                
                with data_table_container:
                    ui.label('📊 Data Preview').style(
                        f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
                    )
                    
                    # Sample data table with editable cells
                    with ui.table(columns=headers, rows=[]).style('width: 100%;'):
                        pass
                    
                    # Action buttons
                    with ui.row().style('gap: 1rem; margin-top: 1.5rem;'):
                        ui.button('💾 Save as Draft', on_click=lambda: save_as_draft()).style(
                            f'background: #6c757d; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                        )
                        ui.button('✅ Submit', on_click=lambda: submit_data()).style(
                            f'background: {main_color}; color: white; padding: 0.75rem 1.5rem; border-radius: 8px; border: none; font-weight: 600;'
                        )
                
                ui.notify('File uploaded successfully! Please review and edit the data.', color='positive')
                
            except Exception as e:
                ui.notify(f'Error processing file: {str(e)}', color='negative')
        
        def save_as_draft():
            ui.notify('Data saved as draft successfully!', color='positive')
        
        def submit_data():
            ui.notify('Data submitted successfully!', color='positive')
    
    # Simple layout for this page
    with ui.column().style('padding: 2rem; max-width: 1200px; margin: 0 auto;'):
        content()

# Fill Criteria Page for Program/Department Admins
@ui.page('/fill_criteria/{criteria_id}')
def fill_criteria_page(criteria_id: str):
    """Page for filling criteria data with spreadsheet upload and table editing"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    
    # Get criteria details
    criteria = criterias_col.find_one({'_id': ObjectId(criteria_id)})
    if not criteria:
        ui.notify('Criteria not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details
    institution = institutions_col.find_one({'_id': ObjectId(criteria['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content(inst, main_color):
        ui.label(f'Fill Criteria: {criteria.get("name", "Unknown")}').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
        )
        
        # Working headers (from criteria) and grid data state
        headers = [h for h in criteria.get('headers', []) if str(h).strip()]
        required_headers = headers[:]  # treat all as required for now

        # Container references
        grid_container = ui.column().style('gap: 0.75rem;')
        action_bar = ui.row().style('gap: 0.75rem; margin-top: 0.5rem;')

        # Utility: build grid columns for AgGrid
        def build_columns(hdrs):
            return [{
                'field': h,
                'headerName': h,
                'editable': True,
                'sortable': True,
                'filter': True,
                'resizable': True,
            } for h in hdrs]

        # Utility: normalize rows to ensure all headers exist
        def normalize_rows(rows, hdrs):
            norm = []
            for r in rows:
                row = {}
                for h in hdrs:
                    row[h] = '' if r.get(h) is None else str(r.get(h))
                norm.append(row)
            return norm

        # Load existing draft for this user and criteria (latest)
        existing_rows = []
        try:
            from bson import ObjectId
            query = {
                'criteria_id': ObjectId(criteria_id),
                'submitted_by': current_user.get('email'),
                'status': 'draft',
            }
            draft_doc = criteria_submissions_col.find_one(query, sort=[('updated_at', DESCENDING)])
            if draft_doc:
                headers = draft_doc.get('headers', headers) or headers
                existing_rows = draft_doc.get('data', [])
        except Exception as e:
            # Safe fallback; keep UI functional
            existing_rows = []

        # Grid: either from existing draft or empty
        rows_state = normalize_rows(existing_rows, headers)

        # File upload: allows replacing/adding data
        with ui.card().style('background: white; border: 2px dashed #e9ecef; padding: 1rem; border-radius: 10px;'):
            ui.label('📤 Upload Spreadsheet (.xlsx/.xls/.csv)').style(
                f'font-weight: 600; color: {main_color}; margin-bottom: 0.25rem;'
            )
            ui.label('Headers must match the configured columns for this criteria.').style('color: var(--text-secondary); margin-bottom: 0.75rem;')

            async def on_upload(e):
                try:
                    file = e.files[0]
                    name = file.name.lower()
                    content = await file.read()
                    import pandas as pd, io
                    if name.endswith('.xlsx') or name.endswith('.xls'):
                        df = pd.read_excel(io.BytesIO(content))
                    elif name.endswith('.csv'):
                        df = pd.read_csv(io.BytesIO(content))
                    else:
                        raise ValueError('Unsupported file format. Use .xlsx, .xls or .csv')

                    # Ensure columns match headers
                    df_cols = [str(c).strip() for c in df.columns.tolist()]
                    if set(df_cols) != set(headers):
                        ui.notify('Uploaded columns do not match required headers. Please adjust your file.', color='warning')
                    # Reorder and coerce to string
                    safe_df = df.reindex(columns=headers)
                    new_rows = safe_df.fillna('').astype(str).to_dict(orient='records')

                    # Replace rows_state
                    grid_container.clear()
                    build_grid(new_rows)
                    ui.notify(f'Loaded {len(new_rows)} rows from file', color='positive')
                except Exception as ex:
                    ui.notify(f'Upload error: {ex}', color='negative')

            ui.upload(label='Choose File', on_upload=on_upload).style(
                f'background: {main_color}; color: white; padding: 0.6rem 1rem; border-radius: 8px; border: none;'
            )

        # Build the editable grid using AgGrid if available; fallback to simple table
        def build_grid(initial_rows):
            nonlocal rows_state
            rows_state = normalize_rows(initial_rows or [], headers)

            try:
                grid = ui.aggrid({
                    'defaultColDef': {
                        'editable': True,
                        'sortable': True,
                        'filter': True,
                        'resizable': True,
                    },
                    'columnDefs': build_columns(headers),
                    'rowData': rows_state,
                    'animateRows': True,
                    'rowSelection': 'multiple',
                }).style('height: 400px; width: 100%;')
            except Exception:
                # Fallback to simple table with inline editors (less capable)
                with ui.table(columns=headers, rows=rows_state).style('width: 100%;'):
                    pass

            # Row actions
            with action_bar:
                def add_row():
                    rows_state.append({h: '' for h in headers})
                    grid_container.clear(); build_grid(rows_state)
                def clear_all():
                    rows_state.clear(); grid_container.clear(); build_grid(rows_state)
                ui.button('➕ Add Row', on_click=add_row).props('flat color=primary')
                ui.button('🧹 Clear All', on_click=clear_all).props('flat color=warning')

            # Save/Submit actions
            def collect_rows() -> list:
                # Try to read from aggrid model if present, else use rows_state
                try:
                    # NiceGUI's aggrid stores data in .options['rowData']
                    return grid.options.get('rowData', rows_state)  # type: ignore
                except Exception:
                    return rows_state

            def validate_required(rows: list) -> Optional[str]:
                for i, r in enumerate(rows, start=1):
                    for h in required_headers:
                        if not str(r.get(h, '')).strip():
                            return f'Missing value for "{h}" in row {i}'
                return None

            def do_upsert(status: str):
                from bson import ObjectId
                doc = {
                    'criteria_id': ObjectId(criteria_id),
                    'institution_id': ObjectId(criteria.get('institution_id')) if criteria.get('institution_id') else None,
                    'academic_year_id': criteria.get('academic_year_id'),
                    'headers': headers,
                    'data': collect_rows(),
                    'status': status,
                    'submitted_by': current_user.get('email'),
                    'updated_at': datetime.now(timezone.utc),
                }
                # include program/department if available from session
                if current_user.get('program_id'):
                    doc['program_id'] = ObjectId(current_user['program_id'])
                if current_user.get('department_id'):
                    doc['department_id'] = ObjectId(current_user['department_id'])

                # Upsert by user+criteria+status=draft for drafts; insert new on submit
                if status == 'draft':
                    criteria_submissions_col.update_one(
                        {
                            'criteria_id': doc['criteria_id'],
                            'submitted_by': doc['submitted_by'],
                            'status': 'draft',
                        },
                        {
                            '$set': doc,
                            '$setOnInsert': {'created_at': datetime.now(timezone.utc)},
                        },
                        upsert=True,
                    )
                    return None
                else:
                    doc['submitted_at'] = datetime.now(timezone.utc)
                    return criteria_submissions_col.insert_one(doc)

            def on_save_draft():
                try:
                    rows = collect_rows()
                    # No validation for drafts; save raw
                    do_upsert('draft')
                    ui.notify(f'Saved draft with {len(rows)} rows', color='positive')
                except Exception as ex:
                    ui.notify(f'Error saving draft: {ex}', color='negative')

            def on_submit():
                try:
                    rows = collect_rows()
                    err = validate_required(rows)
                    if err:
                        ui.notify(err, color='warning'); return
                    res = do_upsert('submitted')
                    ui.notify('Submission created' + (f' (ID: {str(res.inserted_id)})' if res else ''), color='positive')
                except Exception as ex:
                    ui.notify(f'Error submitting: {ex}', color='negative')

            with ui.row().style('gap: 0.75rem; margin-top: 0.75rem; justify-content: flex-end;'):
                ui.button('💾 Save Draft', on_click=on_save_draft).style('background: #6c757d; color: white;')
                ui.button('✅ Submit', on_click=on_submit).style(f'background: {main_color}; color: white;')

        # Build initial grid (existing draft if any)
        with grid_container:
            build_grid(rows_state)
    
    # Simple layout for this page
    with ui.column().style('padding: 2rem; max-width: 1200px; margin: 0 auto;'):
        content(institution, main_color)

# 404 Error Page
@ui.page('/404')
def error_404_page():
    """404 Not Found error page"""
    def content():
        ui.label('😵 404 - Page Not Found').style(
            'font-size: 3rem; font-weight: bold; color: #dc3545; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label('The page you are looking for does not exist.').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        ui.label('It might have been moved, deleted, or you entered the wrong URL.').style(
            'font-size: 1rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        with ui.row().style('gap: 1rem; justify-content: center;'):
            ui.button('🏠 Go Home', on_click=lambda: ui.navigate.to('/')).style(
                'background: rgb(154, 44, 84); color: white; padding: 1rem 2rem; border-radius: 8px; border: none; font-weight: 600; font-size: 1.1rem;'
            )
            
            ui.button('🔙 Go Back', on_click=lambda: ui.run_javascript('history.back()')).style(
                'background: #6c757d; color: white; padding: 1rem 2rem; border-radius: 8px; border: none; font-weight: 600; font-size: 1.1rem;'
            )
    
    # Simple layout
    with ui.column().style('padding: 4rem 2rem; max-width: 800px; margin: 0 auto; text-align: center;'):
        content()

# Generic Error Page
@ui.page('/error')
def generic_error_page():
    """Generic error page for other errors"""
    def content():
        ui.label('⚠️ Something Went Wrong').style(
            'font-size: 3rem; font-weight: bold; color: #ffc107; margin-bottom: 1rem; text-align: center;'
        )
        
        ui.label('An unexpected error occurred. Please try again.').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        ui.label('If the problem persists, contact your administrator.').style(
            'font-size: 1rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        with ui.row().style('gap: 1rem; justify-content: center;'):
            ui.button('🔄 Try Again', on_click=lambda: ui.run_javascript('window.location.reload()')).style(
                'background: rgb(154, 44, 84); color: white; padding: 1rem 2rem; border-radius: 8px; border: none; font-weight: 600; font-size: 1.1rem;'
            )
            
            ui.button('🏠 Go Home', on_click=lambda: ui.navigate.to('/')).style(
                'background: #6c757d; color: white; padding: 1rem 2rem; border-radius: 8px; border: none; font-weight: 600; font-size: 1.1rem;'
            )
    
    # Simple layout
    with ui.column().style('padding: 4rem 2rem; max-width: 800px; margin: 0 auto; text-align: center;'):
        content()

# Change Password Page
# Department Admin Dashboard
@ui.page('/department_admin/{department_id}')
def department_admin_dashboard(department_id: str):
    """Department admin dashboard - only shows criteria and extended profiles"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    
    # Get department details
    department = schools_col.find_one({'_id': ObjectId(department_id), 'type': 'department'})
    if not department:
        ui.notify('Department not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details
    institution = institutions_col.find_one({'_id': ObjectId(department['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content(inst, main_color):
        ui.label(f'Department Admin Dashboard').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
        )
        ui.label(f'Department: {department.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem;'
        )
        
        # Only show criteria and extended profiles
        with ui.row().style('width: 100%; gap: 2rem;'):
            # Criteria Section
            with ui.card().classes('beautiful-card').style('flex: 1; padding: 2rem;'):
                ui.label('📊 Criteria Management').style(
                    f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;'
                )
                
                # Get criteria for this department
                department_criterias = list(criterias_col.find({
                    'institution_id': department['institution_id'],
                    'scope_type': 'department_based'
                }))
                
                if department_criterias:
                    for criteria in department_criterias:
                        with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                            ui.label(f"📋 {criteria.get('name', 'Unnamed')}").style(
                                'font-size: 1.1rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                            )
                            ui.label(f"Deadline: {criteria.get('deadline', 'No deadline').strftime('%Y-%m-%d') if criteria.get('deadline') else 'No deadline'}").style(
                                'font-size: 0.9rem; color: var(--warning-color); margin-bottom: 0.5rem;'
                            )
                            
                            # Fill criteria button
                            ui.button('📝 Fill Criteria', on_click=lambda c_id=str(criteria['_id']): ui.navigate.to(f'/department_admin/{department_id}/fill_criteria/{c_id}')).style(
                                f'background: {main_color}; color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                            )
                else:
                    ui.label('No criteria available for this department.').style(
                        'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
                    )
            
            # Extended Profiles Section
            with ui.card().classes('beautiful-card').style('flex: 1; padding: 2rem;'):
                ui.label('👤 Extended Profiles').style(
                    f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;'
                )
                
                # Get extended profiles for this department
                department_profiles = list(extended_profiles_col.find({
                    'institution_id': department['institution_id'],
                    'scope_type': 'department_based'
                }))
                
                if department_profiles:
                    for profile in department_profiles:
                        with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                            ui.label(f"📄 {profile.get('name', 'Unnamed')}").style(
                                'font-size: 1.1rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                            )
                            ui.label(f"Deadline: {profile.get('deadline', 'No deadline').strftime('%Y-%m-%d') if profile.get('deadline') else 'No deadline'}").style(
                                'font-size: 0.9rem; color: var(--warning-color); margin-bottom: 0.5rem;'
                            )
                            
                            # Fill profile button
                            ui.button('📝 Fill Profile', on_click=lambda p_id=str(profile['_id']): ui.navigate.to(f'/department_admin/{department_id}/fill_profile/{p_id}')).style(
                                f'background: {main_color}; color: white; padding: 0.5rem 1rem; border-radius: 6px; border: none; font-weight: 500;'
                            )
                else:
                    ui.label('No extended profiles available for this department.').style(
                        'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
                    )
    
    # Create Institution Admin-style sidebar for department admin
    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
        # Sidebar column with Institution Admin styling
        with ui.column().style('width: 280px; background: var(--surface); border-right: 2px solid var(--border); padding: 0; height: 100vh; overflow-y: auto;'):
            # Institution header card
            with ui.card().style('background: linear-gradient(135deg, var(--primary-color), var(--primary-dark)); color: white; padding: 1.5rem; margin: 1rem; border: none; border-radius: 12px; box-shadow: var(--shadow);'):
                # Institution logo placeholder
                ui.element('div').style('width: 60px; height: 60px; background: rgba(255,255,255,0.2); border-radius: 50%; margin: 0 auto 1rem; display: flex; align-items: center; justify-content: center; font-size: 1.5rem; font-weight: bold;').inner_html = '🏢'
                
                # Institution name
                ui.label(institution.get('name', 'Institution')).style('font-size: 1.2rem; font-weight: bold; text-align: center; margin-bottom: 0.5rem;')
                
                # User info section
                ui.separator().style('margin: 1rem 0; background: rgba(255,255,255,0.3);')
                
                # Get current user email
                current_user_email = app.storage.user.get('current_user', {}).get('email', 'Unknown User') if hasattr(app.storage, 'user') else 'Unknown User'
                ui.label(current_user_email).style('font-size: 0.9rem; text-align: center; opacity: 0.9; margin-bottom: 0.5rem;')
                ui.label('Department Admin').style('font-size: 0.9rem; text-align: center; opacity: 0.8;')
            
            # Navigation menu
            with ui.column().style('padding: 1rem; gap: 0.5rem;'):
                # Overview Section
                with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 1rem; margin-bottom: 1rem;'):
                    ui.label('📊 Overview').style('font-size: 1rem; font-weight: 600; color: var(--text-primary); margin-bottom: 0.5rem;')
                    
                    overview_items = [
                        ('📋 Criteria Management', lambda: ui.navigate.to(f'/department_admin/{department_id}')),
                        ('👤 Extended Profiles', lambda: ui.navigate.to(f'/department_admin/{department_id}')),
                        ('📈 Data Analytics', lambda: ui.notify('Coming soon!', color='info')),
                        ('📥 My Submissions', lambda: ui.navigate.to(f'/department_admin/{department_id}/submissions')),
                        ('⚙️ Settings', lambda: ui.notify('Coming soon!', color='info'))
                    ]
                    
                    for item_text, item_action in overview_items:
                        ui.button(item_text, on_click=item_action).style(
                            'width: 100%; justify-content: flex-start; background: white; color: var(--text-primary); border: 1px solid #e9ecef; padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 500; margin-bottom: 0.5rem;'
                        ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "#f8f9fa"; event.target.style.borderColor = "{main_color}";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "white"; event.target.style.borderColor = "#e9ecef";'))
                    
                    # Logout button
                    ui.separator().style('margin: 1rem 0; background: #e9ecef;')
                    ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                        f'width: 100%; justify-content: center; background: #dc3545; color: white; border: 1px solid #dc3545; '
                        f'padding: 0.75rem 1rem; border-radius: 8px; transition: all 0.3s ease; font-weight: 600;'
                    ).on('mouseenter', lambda e: ui.run_javascript(f'event.target.style.background = "#c82333"; event.target.style.borderColor = "#c82333";')).on('mouseleave', lambda e: ui.run_javascript(f'event.target.style.background = "#dc3545"; event.target.style.borderColor = "#dc3545";'))
        
        # Main content
        with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
            # Main content will be rendered here by other UI components
            pass

# School Admin Dashboard
@ui.page('/school_admin/{school_id}')
def school_admin_dashboard(school_id: str):
    """School admin dashboard - shows schools and programs"""
    add_beautiful_global_styles()
    
    # Check authentication
    if not check_auth():
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    from bson import ObjectId
    
    # Get school details
    school = schools_col.find_one({'_id': ObjectId(school_id), 'type': {'$ne': 'department'}})
    if not school:
        ui.notify('School not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details
    institution = institutions_col.find_one({'_id': ObjectId(school['institution_id'])})
    main_color = institution.get('theme_color', 'rgb(154, 44, 84)') if institution else 'rgb(154, 44, 84)'
    
    def content(inst, main_color):
        ui.label(f'School Admin Dashboard').classes('fade-in').style(
            f'font-size: 2rem; font-weight: bold; color: {main_color}; margin-bottom: 1rem;'
        )
        ui.label(f'School: {school.get("name", "Unknown")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 2rem;'
        )
        
        # Show programs in this school
        with ui.card().classes('beautiful-card').style('width: 100%; padding: 2rem;'):
            ui.label('🎓 Programs in School').style(
                f'font-size: 1.5rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem;'
            )
            
            # Get programs for this school
            school_programs = list(programs_col.find({'school_id': school_id}))
            
            if school_programs:
                for program in school_programs:
                    with ui.card().style('background: #f8f9fa; border: 1px solid #e9ecef; padding: 1rem; margin-bottom: 1rem; border-radius: 8px;'):
                        ui.label(f"📚 {program.get('name', 'Unnamed')}").style(
                            'font-size: 1.1rem; font-weight: bold; color: var(--text-primary); margin-bottom: 0.5rem;'
                        )
                        ui.label(f"Code: {program.get('code', 'N/A')}").style(
                            'font-size: 0.9rem; color: var(--text-secondary); margin-bottom: 0.5rem;'
                        )
                        ui.label(f"Duration: {program.get('duration', 'N/A')}").style(
                            'font-size: 0.9rem; color: var(--text-secondary);'
                        )
            else:
                ui.label('No programs found in this school.').style(
                    'color: var(--text-secondary); font-style: italic; text-align: center; padding: 2rem;'
                )
    
    # Create a simple sidebar for school admin
    with ui.row().style('height: 100vh; width: 100vw; background: var(--background);'):
        # Simple sidebar
        with ui.column().style('width: 250px; background: var(--surface); border-right: 2px solid var(--border); padding: 1rem;'):
            # Header
            with ui.card().style('background: linear-gradient(135deg, var(--primary-color), var(--primary-dark)); color: white; padding: 1.5rem; margin-bottom: 1rem; border: none;'):
                ui.label('🏫 School Admin').style('font-size: 1.2rem; font-weight: bold; text-align: center;')
                ui.label(f'{school.get("name", "Unknown")}').style('font-size: 1rem; text-align: center; opacity: 0.9;')
            
            # Navigation
            ui.button('📊 Dashboard', on_click=lambda: None).style(
                f'background: {main_color}; color: white; width: 100%; margin-bottom: 0.5rem; padding: 0.75rem; border-radius: 6px; border: none; font-weight: 500;'
            )
            
            # Logout button
            ui.button('🚪 Logout', on_click=lambda: ui.navigate.to('/')).style(
                'background: #dc3545; color: white; width: 100%; padding: 0.75rem; border-radius: 6px; border: none; font-weight: 500;'
            )
        
        # Main content
        with ui.column().style('flex: 1; padding: 2rem; overflow-y: auto;'):
            # Main content will be rendered here by other UI components
            pass

@ui.page('/change_password')
def change_password_page():
    """Page for users to change their password"""
    global current_user
    
    if not current_user or not current_user.get('email'):
        ui.notify('Please log in first', color='negative')
        ui.navigate.to('/')
        return
    
    # Get user details from database
    user = users_col.find_one({'email': current_user['email']})
    if not user:
        ui.notify('User not found', color='negative')
        ui.navigate.to('/')
        return
    
    # Get institution details for theming
    institution_id = user.get('institution_id')
    institution = None
    main_color = 'rgb(154, 44, 84)'  # Default color
    
    if institution_id:
        try:
            from bson import ObjectId
            institution = institutions_col.find_one({'_id': ObjectId(institution_id)})
            if institution:
                main_color = institution.get('theme_color', 'rgb(154, 44, 84)')
        except:
            pass
    
    def content():
        ui.label('🔐 Change Password').style(
            f'font-size: 2.5rem; font-weight: bold; color: {main_color}; margin-bottom: 2rem; text-align: center;'
        )
        
        ui.label(f'Welcome, {user.get("first_name", "")} {user.get("last_name", "")}').style(
            'font-size: 1.2rem; color: var(--text-secondary); margin-bottom: 0.5rem; text-align: center;'
        )
        
        ui.label(f'Role: {user.get("role", "Unknown")}').style(
            'font-size: 1rem; color: var(--text-secondary); margin-bottom: 2rem; text-align: center;'
        )
        
        with ui.card().style('max-width: 500px; margin: 0 auto; padding: 2rem; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.1);'):
            ui.label('Enter New Password').style(
                f'font-size: 1.3rem; font-weight: bold; color: {main_color}; margin-bottom: 1.5rem; text-align: center;'
            )
            
            new_password = ui.input(
                label='New Password',
                placeholder='Enter new password',
                password=True
            ).style('margin-bottom: 1rem;')
            
            confirm_password = ui.input(
                label='Confirm New Password',
                placeholder='Confirm new password',
                password=True
            ).style('margin-bottom: 1.5rem;')
            
            def submit_password_change():
                if not new_password.value or not confirm_password.value:
                    ui.notify('All fields are required', color='negative')
                    return
                
                if new_password.value != confirm_password.value:
                    ui.notify('New passwords do not match', color='negative')
                    return
                
                if len(new_password.value) < 6:
                    ui.notify('Password must be at least 6 characters long', color='negative')
                    return
                
                # Generate new salt and hash
                new_salt = secrets.token_hex(16)
                new_hash = hash_password(new_password.value, new_salt)
                
                try:
                    # Update user password
                    users_col.update_one(
                        {'_id': user['_id']},
                        {
                            '$set': {
                                'password_hash': new_hash,
                                'salt': new_salt,
                                'must_change_password': False,
                                'updated_at': datetime.datetime.now(datetime.timezone.utc)
                            }
                        }
                    )
                    
                    # Log the action
                    log_audit_action(
                        action='Password Changed',
                        details='User changed their password',
                        user_email=user['email'],
                        institution_id=institution_id
                    )
                    
                    ui.notify('Password changed successfully!', color='positive')
                    
                    # Store user in session and local storage
                    user_data = {
                        'email': user['email'],
                        'role': user.get('role', ''),
                        'institution_id': str(user.get('institution_id', '')),
                        'program_id': str(user.get('program_id', '')),
                        'department_id': str(user.get('department_id', '')),
                        'school_id': str(user.get('school_id', '')),
                        'name': user.get('name', '')
                    }
                    app.storage.user = {'user': user_data}
                    
                    # Force update the current_user global
                    global current_user
                    current_user = user_data
                    
                    # Log the login
                    log_audit_action(
                        'LOGIN',
                        f'User {user_data["email"]} logged in',
                        user_email=user_data['email'],
                        institution_id=user_data.get('institution_id')
                    )
                    
                    # Redirect based on user role
                    if user_data.get('role') == 'Institution Admin' and user_data.get('institution_id'):
                        ui.navigate.to(f"/institution_admin/{user_data['institution_id']}")
                    elif user_data.get('role') == 'Program Admin' and user_data.get('program_id'):
                        ui.navigate.to(f"/program_admin/{user_data['program_id']}")
                    elif user_data.get('role') == 'Department Admin' and user_data.get('department_id'):
                        ui.navigate.to(f"/department_admin/{user_data['department_id']}")
                    elif user_data.get('role') == 'School Admin' and user_data.get('school_id'):
                        ui.navigate.to(f"/school_admin/{user_data['school_id']}")
                    else:
                        ui.navigate.to('/dashboard')
                        
                except Exception as e:
                    ui.notify(f'Error changing password: {str(e)}', color='negative')
            
            ui.button('Change Password', on_click=submit_password_change).style(
                f'background: {main_color}; color: white; width: 100%; padding: 1rem; border-radius: 8px; border: none; font-weight: 600; font-size: 1.1rem;'
            )
            
            ui.button('Cancel', on_click=lambda: ui.navigate.to('/')).style(
                'background: #6c757d; color: white; width: 100%; padding: 0.75rem; border-radius: 8px; border: none; font-weight: 600; margin-top: 1rem;'
            )
    
    # Simple layout
    with ui.column().style('padding: 2rem; max-width: 800px; margin: 0 auto;'):
        content()

@ui.page('/logout')
def logout():
    """Handle user logout"""
    global current_user
    if app.storage.user.get('user'):
        log_audit_action(
            'LOGOUT',
            f'User {app.storage.user["user"].get("email")} logged out',
            user_email=app.storage.user['user'].get('email'),
            institution_id=app.storage.user['user'].get('institution_id')
        )
    # Clear user data properly
    app.storage.clear()
    current_user = None
    ui.navigate.to('/')
    return "Logging out..."

# Main application entry point
if __name__ == '__main__':
    # Run the application
    print("🚀 Starting IQAC Portal...")
    print("📊 Using MongoDB database...")
    
    print("🌐 Starting web interface...")
    ui.run(
        title='IQAC Portal',
        port=8083,
        reload=False,
        show=True,
        storage_secret='your-secret-key-here-12345'
    )
